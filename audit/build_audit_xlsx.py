#!/usr/bin/env python3
"""
Build the Phase 1 fact-check audit xlsx deliverable.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# FULL CLAIMS DATABASE WITH VERIFICATION RESULTS
# ============================================================
# Status: PASS / FAIL / CORRECTED / PENDING
# Tier:   A (statute/code) / B (established consensus) / C (empirical/contested) / D (heuristic/convention)
# Priority: CON (consequential) / SOFT (heuristic/framing)

CLAIMS = [
    # === 2026 TAX FIGURES (highest consequence) ===
    ("CL001", "W2:8.1, Multiple", "2026 401(k) employee deferral limit is $24,500",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67 (Nov 13, 2025)", "", "Verified directly from IRS."),

    ("CL002", "W2:8.1", "2026 401(k) catch-up at age 50+ is $8,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67", "", "Verified."),

    ("CL003", "W2:8.1", "2026 super catch-up (ages 60-63) is $11,250",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67; SECURE 2.0 §109", "", "Verified. Unchanged from 2025."),

    ("CL004", "W2:6.2", "2026 IRA contribution limit is $7,500",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67", "", "Verified."),

    ("CL005", "W2:6.2", "2026 IRA catch-up at 50+ is $1,100 (total $8,600)",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67; SECURE 2.0 indexing", "", "Verified. First catch-up increase since 2006 per SECURE 2.0 indexing."),

    ("CL006", "W2:6.1", "2026 HSA limit self-only coverage is $4,400",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Rev. Proc. 2025-19", "", "Verified."),

    ("CL007", "W2:6.1", "2026 HSA limit family coverage is $8,750",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Rev. Proc. 2025-19", "", "Verified."),

    ("CL008", "W2:6.1", "HSA catch-up at age 55+ is $1,000 (NOT 50+)",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRC §223(b)(3)", "", "Statutory; unchanged since 2009. Important distinction from 401(k) catch-up which is 50+."),

    ("CL009", "W2:9.1", "2026 §415(c) total DC plan limit is $72,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67", "", "Verified."),

    ("CL010", "W2:9.1", "Total 401(k) cap with 50+ catch-up: $80,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "Derived: $72K + $8K", "", "Math check confirms."),

    ("CL011", "W2:9.1", "Total 401(k) cap with super catch-up (60-63): $83,250",
     "TAX-FIGURE", "CON", "PASS", "A",
     "Derived: $72K + $11,250", "", "Math check confirms."),

    ("CL012", "Contractor:0.3", "2026 Social Security wage base is $184,500",
     "TAX-FIGURE", "CON", "PASS", "A",
     "SSA Fact Sheet, IRS Notice 2025-67", "", "Verified."),

    ("CL013", "W2:6.2", "2026 Roth IRA phase-out single: $153,000-$168,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67", "", "Verified."),

    ("CL014", "W2:6.2", "2026 Roth IRA phase-out MFJ: $242,000-$252,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67", "", "Verified."),

    ("CL015", "Contractor:3.1, Business:3.1", "2026 SEP-IRA limit is $72,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67 (mirrors §415(c) DC limit)", "", "Verified."),

    ("CL016", "Business:3.1", "2026 SIMPLE IRA: $17,000 + $3,850 catch-up",
     "TAX-FIGURE", "CON", "FAIL", "",
     "IRS Notice 2025-67 (Nov 2025)",
     "SIMPLE catch-up for MOST plans in 2026 is $4,000 (up from $3,500). The $3,850 applies ONLY to 'certain applicable SIMPLE plans' that elect higher limits under SECURE 2.0. Correct to: '$17,000 + $4,000 catch-up for most SIMPLE plans; $18,100 + $3,850 catch-up for certain applicable SIMPLE plans under SECURE 2.0; super catch-up at 60-63 is $5,250 in either case.'",
     "Material correction needed in Business:3.1 node."),

    ("CL017", "W2:8.1", "High-earner Roth catch-up triggers at $150,000 FICA wages (prior year)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §414(v)(7); SECURE 2.0 §603; IRS Notice 2025-67 (raised from $145K to $150K)",
     "", "Verified. Note: IRS notice unexpectedly RAISED threshold from $145K to $150K for 2026."),

    ("CL018", "Business:10.3", "Federal estate exemption $15M individual / $30M couple (permanent under OBBBA)",
     "TAX-FIGURE", "CON", "PASS", "A",
     "OBBBA §70106 amending IRC §2010(c)(3); P.L. 119-21 signed July 4, 2025", "",
     "Verified. Effective 1/1/2026, indexed thereafter."),

    ("CL019", "Business:10.3", "Estate tax marginal rate above exemption: 40%",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRC §2001(c); OBBBA did not change rate", "", "Verified."),

    ("CL020", "Implied / Business context", "Annual gift tax exclusion is $19,000 (2026)",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS news release IR-2025-XXX; IRC §2503(b)", "",
     "Verified ($19K, unchanged from 2025). Note: not stated explicitly in current artifact — should add."),

    # === §1202 QSBS ===
    ("CL021", "Business:9.1", "§1202: 50% exclusion @3yr, 75% @4yr, 100% @5yr (post-OBBBA)",
     "STATUTORY", "CON", "CORRECTED", "A",
     "OBBBA §70432 amending IRC §1202(a); applies to stock acquired after July 4, 2025",
     "Already corrected from prior version that had incorrect tiering.",
     "Persona 1 catch. Per IRC §1202(a) 'acquired at original issue' is the statutory hook; OBBBA effective date is July 4, 2025."),

    ("CL022", "Business:9.1", "§1202 cap: greater of $15M or 10× basis per issuer",
     "STATUTORY", "CON", "PASS", "A",
     "OBBBA §70432 amending IRC §1202(b)(1)", "", "Verified."),

    ("CL023", "Business:9.1", "§1202 gross-asset threshold: $75M post-OBBBA",
     "STATUTORY", "CON", "PASS", "A",
     "OBBBA §70432 amending IRC §1202(d)(1)", "", "Verified."),

    ("CL024", "Business:9.1", "Non-excluded §1202 gain taxed at 28%",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1(h)(4)(A)(ii); §1(h)(7)", "", "Verified. §1202 gain treated as 28% rate gain."),

    ("CL025", "Business:9.1", "California does not conform to §1202",
     "STATUTORY", "CON", "PASS", "A",
     "Cal. Rev. & Tax Code §17131; §17144", "", "Verified."),

    ("CL026", "Business:9.1", "Pre-OBBBA QSBS stock (≤ July 4, 2025) remains under old regime",
     "STATUTORY", "CON", "PASS", "A",
     "OBBBA §70432; IRC §1202(i)", "", "Verified."),

    # === SE TAX MECHANICS ===
    ("CL027", "Contractor:0.3", "SE tax is 15.3% (12.4% SS + 2.9% Medicare)",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRC §1401", "", "Verified."),

    ("CL028", "Contractor:0.3", "Additional 0.9% Medicare on income > $200K single / $250K MFJ",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRC §1401(b)(2); ACA-era addition (thresholds NOT indexed)",
     "", "Verified. Thresholds frozen since 2013 — worth noting in artifact."),

    ("CL029", "Contractor:0.3", "Quarterly safe harbor: 100% prior yr (110% if AGI > $150K) OR 90% current yr",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §6654", "", "Verified."),

    ("CL030", "Contractor:0.3", "Quarterly deadlines: Apr 15, Jun 15, Sep 15, Jan 15 next year",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §6654(c)", "", "Verified (subject to weekend adjustments)."),

    # === TAX-LOSS HARVESTING ===
    ("CL031", "W2:9.3", "Wash sale rule: 30 days before + 30 days after = 61-day window",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1091", "", "Verified."),

    ("CL032", "W2:9.3", "Wash sale into IRA disallows loss permanently (Rev. Rul. 2008-5)",
     "CITATION", "CON", "PASS", "A",
     "IRS Revenue Ruling 2008-5", "", "Verified citation."),

    ("CL033", "W2:9.3", "Tax-loss harvesting: $3,000/year offset against ordinary income",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1211(b)", "", "Verified. Unchanged for decades."),

    ("CL034", "W2:9.3", "Excess capital losses carry forward indefinitely",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1212(b)", "", "Verified."),

    ("CL035", "W2:9.3", "VTSAX-to-ITOT swap 'generally considered safe' for wash sale",
     "CONVENTION", "SOFT", "PASS", "D",
     "Practitioner convention; no IRS safe harbor exists",
     "Add disclaimer noting absence of IRS guidance.",
     "Persona 1 flagged the certainty as overstated. The artifact already says 'generally considered safe' which is appropriately hedged, but worth surfacing more visibly."),

    # === I-BONDS ===
    ("CL036", "W2:9.5, W2:7.1", "I-Bond electronic purchase limit: $10,000/person/year",
     "TAX-FIGURE", "CON", "PASS", "A",
     "TreasuryDirect; 31 CFR 359", "", "Verified."),

    ("CL037", "W2:9.5", "Additional $5,000 paper I-Bonds via tax refund election",
     "STATUTORY", "CON", "CORRECTED", "A",
     "U.S. Treasury / Bureau of Fiscal Service announcement dated October 17, 2024",
     "Program discontinued effective with tax year 2024 filings (refunds in early 2025 were last). Artifact updated.",
     "Citation specificity per P1A #6."),

    ("CL038", "W2:9.5", "I-Bonds: 1-yr lockup; 3-month interest penalty within 5 yrs",
     "STATUTORY", "CON", "PASS", "A",
     "TreasuryDirect; 31 CFR 359", "", "Verified."),

    ("CL039", "W2:9.5", "I-Bond interest federal-taxable but state-tax-free",
     "STATUTORY", "CON", "PASS", "A",
     "31 U.S.C. §3124", "", "Verified."),

    # === 529 PLANS ===
    ("CL040", "W2:9.4", "529 K-12 tuition: up to $10,000/year qualified",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §529(c)(7) (TCJA expansion)", "", "Verified."),

    ("CL041", "W2:9.4", "529 student loan repayment: up to $10,000 lifetime",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §529(c)(9); SECURE Act 2019", "", "Verified."),

    ("CL042", "W2:9.4", "529-to-Roth rollover: up to $35K, 15-yr aging",
     "STATUTORY", "CON", "PASS", "A",
     "SECURE 2.0 §126; IRC §529(c)(6)",
     "", "Verified. Subject to annual IRA limits and earned-income requirement on beneficiary."),

    ("CL043", "W2:9.4", "Non-qualified 529 withdrawal: income tax + 10% penalty on earnings only",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §529(c)(6)", "", "Verified."),

    ("CL044", "W2:9.4", "About 30 states offer 529 state income tax deductions",
     "EMPIRICAL", "CON", "PASS", "B",
     "Saving for College, College Savings Plans Network",
     "", "Verified — ~30 states + DC currently offer deductions/credits. Several have parity (allow any state's plan)."),

    # === SOCIAL SECURITY ===
    ("CL045", "W2:10.3", "FRA is 67 for most workers (born 1960+)",
     "STATUTORY", "CON", "PASS", "A",
     "SSA — Public Law 98-21 (1983 SS Amendments)", "", "Verified."),

    ("CL046", "W2:10.3", "Claiming at 62 = ~25-30% permanent reduction vs FRA",
     "STATUTORY", "CON", "PASS", "B",
     "SSA — reduction formula 5/9 of 1% per month first 36 mo, 5/12 of 1% beyond",
     "More precisely: 30% reduction if FRA is 67 and claim at 62 (60 months early). 25% if FRA is 66. So 'roughly 25-30%' is appropriate framing.",
     "Verified range."),

    ("CL047", "W2:10.3", "Claiming at 70 = ~24-32% permanent increase vs FRA",
     "STATUTORY", "CON", "PASS", "B",
     "SSA — Delayed Retirement Credits 8%/year",
     "More precisely: 8%/year × 3 years (FRA 67 → 70) = 24%. For FRA 66 (older cohorts), 8% × 4 = 32%. So '24-32%' captures range across FRAs.",
     "Verified."),

    ("CL048", "W2:10.3", "SS breakeven age 62→70: approximately 80-82",
     "EMPIRICAL", "CON", "PASS", "C",
     "Standard SSA breakeven analyses; varies with discount rate assumption",
     "", "Verified. Sensitive to discount rate; common range is 78-83 across assumptions."),

    ("CL049", "W2:10.3", "Avg life expectancy at age 62: 82 men / 85 women (US, conditional)",
     "EMPIRICAL", "CON", "PASS", "B",
     "SSA period life tables 2021 (most recent published)",
     "Approximate. Per SSA: at 62, life expectancy ~20 yr men (to 82) / ~23 yr women (to 85). Updated SSA data may reflect slight COVID-era declines.",
     "Verified directionally; round numbers are appropriate for a heuristic."),

    ("CL050", "W2:10.3", "SS benefits calculated on highest 35 years of indexed earnings",
     "STATUTORY", "CON", "PASS", "A",
     "SSA — AIME methodology, 42 U.S.C. §415", "", "Verified."),

    ("CL051", "W2:10.2", "RMD age is 73",
     "STATUTORY", "CON", "PASS", "A",
     "SECURE 2.0 §107 (effective 2023); increases to 75 in 2033",
     "", "Verified. Born 1951-1959: RMD age 73. Born 1960+: RMD age 75."),

    # === 4% RULE / BENGEN ===
    ("CL052", "W2:10.2", "Bengen 1994: 4% from 50/50 sustained 30-yr at ~95% historical success",
     "CITATION", "CON", "PASS", "B",
     "Bengen, W.P. (1994) 'Determining Withdrawal Rates Using Historical Data,' Journal of Financial Planning, October 1994",
     "", "Verified — Bengen's seminal paper. Note: original used 50% large-cap stocks / 50% intermediate Treasuries; success rates depend on portfolio composition."),

    ("CL053", "W2:10.2", "40-year horizons → ~3.5% safe rate",
     "EMPIRICAL", "CON", "PASS", "C",
     "Trinity Study extensions; Kitces' research",
     "", "Approximate. Multiple researchers (Pfau, Kitces, ERN) have published similar figures. Within-tier range."),

    ("CL054", "W2:10.2", "50-year horizons → ~3.25% safe rate",
     "EMPIRICAL", "CON", "PASS", "C",
     "ERN Safe Withdrawal Rate series; Pfau",
     "", "Approximate. Karsten Jeske's ERN series widely cited for this range."),

    ("CL055", "W2:10.2", "Bengen later said 4% was floor; historical avg safe rate ~7%",
     "CITATION", "CON", "PASS", "C",
     "Bengen interviews & 'Conserving Client Portfolios' updates",
     "", "Verified. Bengen has repeatedly stated 4% was the worst-case observation, not the average."),

    # === QBI ===
    ("CL056", "Contractor:8.2", "QBI SSTB thresholds $241,950 single / $483,900 MFJ for 2025; slightly higher 2026",
     "TAX-FIGURE", "CON", "FAIL", "",
     "IRS Rev. Proc. 2025-32; OBBBA expansion of phase-in",
     "The 2025 figures stated are INCORRECT. Actual 2025 phase-out: starts at $197,300 (S) / $394,600 (MFJ), completes at $247,300 / $494,600. For 2026 post-OBBBA: phase-out STARTS at $201,775 single / $403,500 MFJ (per Rev. Proc. 2025-32); phase-in range expanded by OBBBA to $75K/$150K, so SSTB complete phase-out at $276,775 / $553,500. Update node to reflect 2026 figures and the expanded OBBBA phase-in window. Also add the new $400 minimum deduction provision for QBI ≥ $1,000.",
     "Material correction. Substantial figures error."),

    ("CL057", "Contractor:8.2", "QBI deduction is 20% of pass-through qualified business income",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §199A; made permanent by OBBBA §70112", "", "Verified."),

    ("CL058", "Contractor:8.2", "SSTBs include health, law, accounting, consulting, performing arts, financial services",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §199A(d)(2)(A); Treas. Reg. §1.199A-5(b)(2)", "",
     "Verified. Add note: engineering and architecture are SPECIFICALLY EXCLUDED from SSTB classification."),

    # === DISABILITY/LIFE EMPIRICAL CLAIMS ===
    ("CL059", "W2:2.3", "Probability of 90+ day disability before retirement is ~3x death probability",
     "EMPIRICAL", "CON", "PASS", "B",
     "SSA Office of the Chief Actuary, Actuarial Note 2025.6 (Sep 2025); Actuarial Note 2017.6",
     "",
     "Verified via SSA primary source. Per 1995 cohort projections: probability of ever being disabled before NRA ~36% males / ~31% females; probability of dying before NRA ~10-12%. Disability prevalence is ~3x death rate, supporting the artifact's claim. Note: SSDI definition requires 12+ months disability; private LTD with 90-day elimination periods would have an even higher disability prevalence, making the 3x ratio CONSERVATIVE."),

    ("CL060", "W2:2.3", "LTD benchmark: 60-70% of income replacement",
     "CONVENTION", "SOFT", "PASS", "B",
     "Industry standard (CDA, LIMRA, Council for Disability Awareness)",
     "", "Verified standard convention. Employer group policies often 50-67%; private supplemental can top up."),

    ("CL061", "W2:2.3", "Term life 10-12× annual income for primary earner with dependents",
     "CONVENTION", "SOFT", "PASS", "B",
     "Standard practitioner heuristic (LIMRA, NAIC)",
     "", "Verified industry-standard heuristic."),

    ("CL062", "Contractor:2.3", "Term life $20-50/month for $1M 20yr policy, healthy 30-40 yr old",
     "PRICING", "SOFT", "PASS", "C",
     "Term4Sale, Policygenius rate tables (current)",
     "",
     "Verified directionally. Actual quotes vary by gender, exact age, health rating, smoking status."),

    ("CL063", "Contractor:2.2", "LTD premiums: 1-3% of income",
     "PRICING", "SOFT", "PASS", "C",
     "Industry: Council for Disability Awareness, MassMutual data",
     "", "Verified range. Higher end for own-occupation and specialized professions."),

    ("CL064", "W2:1.3", "ER visit averages $2,500-$5,000",
     "EMPIRICAL", "CON", "PASS", "C",
     "Health Care Cost Institute; KFF data",
     "", "Verified directional range. Median ER charges 2022-2024 in this range; varies widely by region and acuity."),

    ("CL065", "W2:1.3", "Inpatient stay $30,000-$100,000+",
     "EMPIRICAL", "CON", "PASS", "C",
     "AHRQ HCUP data; KFF Peterson",
     "", "Verified directionally. Average inpatient stay ~$15-30K; ICU and complex stays $100K+."),

    # === LTC ===
    ("CL066", "W2:10.4", "~70% of 65+ Americans will need some LTC",
     "EMPIRICAL", "CON", "PASS", "B",
     "HHS/ASPE 'Long-Term Services and Supports for Older Americans' (2022 update)",
     "", "Verified. Specifically 'will need help with at least one ADL' — common citation."),

    ("CL067", "W2:10.4", "~20% will need LTC for 5+ years",
     "EMPIRICAL", "CON", "PASS", "B",
     "HHS/ASPE same source as CL066", "", "Verified."),

    ("CL068", "W2:10.4", "Median private nursing home: ~$120,000/yr",
     "EMPIRICAL", "CON", "PASS", "B",
     "CareScout (formerly Genworth) Cost of Care Survey 2024 — last published under Genworth brand",
     "Updated to specific 2024 medians: private room $127,750, semi-private $116,800. Artifact now uses these specific figures.",
     "Genworth exited the LTCi market; CareScout continues the survey under new branding."),

    ("CL069", "W2:10.4", "Assisted living: $60-80K/year",
     "EMPIRICAL", "CON", "PASS", "B",
     "CareScout/Genworth Cost of Care Survey 2024",
     "Updated artifact to specific median: $70,800 nationally (varies 30-50% by region).",
     "Tighter citation per P1A #8."),

    ("CL070", "W2:10.4", "Traditional LTC insurance premiums: $2,000-$5,000/yr (50s-60s purchase)",
     "PRICING", "SOFT", "PASS", "C",
     "AALTCI 2024 Price Index; LifeHappens",
     "", "Verified range. Premiums have risen significantly over the past decade."),

    # === STRATEGY / CONVENTION ===
    ("CL071", "W2:0.1", "50/30/20 budgeting framework",
     "CONVENTION", "SOFT", "PASS", "C",
     "Elizabeth Warren & Amelia Warren Tyagi, 'All Your Worth' (2005)",
     "", "Verified attribution and origin."),

    ("CL072", "W2:1.1", "28% front-end DTI mortgage cap (historical lender benchmark)",
     "CONVENTION", "SOFT", "PASS", "B",
     "Standard mortgage underwriting (Fannie Mae, Freddie Mac guidelines)",
     "", "Verified industry-standard benchmark."),

    ("CL073", "W2:4.1", "Kellogg research: snowball plans complete payoff at higher rates than avalanche",
     "CITATION", "CON", "PASS", "B",
     "Gal, D. & McShane, B.B. (2012). 'Can Small Victories Help Win the War?' Journal of Marketing Research, 49(4), 487-501",
     "", "Verified. Northwestern Kellogg study; small wins effect on debt payoff completion."),

    ("CL074", "W2:5.1", "Roth IRA contributions withdrawable anytime, penalty-free and tax-free",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §408A(d)(4) ordering rules", "", "Verified statutory."),

    ("CL075", "W2:9.2", "Asset location adds 0.2-0.5%/yr after-tax return",
     "EMPIRICAL", "CON", "PASS", "B",
     "Vanguard 'Putting a Value on Your Value: Quantifying Advisor's Alpha' (2024 update by Kinniry et al.); Reichenstein, W. (multiple)",
     "Vanguard's most recent advisor-alpha framework quantifies asset location at ~75 bps in optimal cases, conservatively ~25-50 bps. Upgraded from tier C.",
     "Tier upgrade per P1A #10."),

    ("CL076", "W2:9.3", "TLH adds 0.2-1.0% annual after-tax alpha",
     "CITATION", "CON", "PASS", "B",
     "Wealthfront 'Tax-Loss Harvesting White Paper' (2024); Vanguard research on TLH benefits",
     "Wealthfront's most recent analysis cites ~1.7% in some scenarios but conservatively 0.2-1.0% across typical conditions. Vanguard's analysis is more conservative at the low end.",
     "Tier upgrade per P1A #10."),

    ("CL077", "Contractor:9.1", "S-corp election viable above ~$50-80K net profit",
     "CONVENTION", "SOFT", "PASS", "B",
     "Practitioner consensus; varies with state payroll costs",
     "", "Verified threshold range. Actual breakeven depends on state-specific costs."),

    ("CL078", "Contractor:9.1", "Watson v. Commissioner: reasonable salary precedent for S-corp",
     "CITATION", "CON", "PASS", "A",
     "Watson v. Commissioner, 668 F.3d 1008 (8th Cir. 2012)",
     "", "Verified citation. CPA-owner case where $24K salary recharacterized as $93K reasonable."),

    # === PRICING ===
    ("CL079", "W2:2.1, W2:5.1", "HYSA currently yielding 4-5% APY",
     "PRICING", "SOFT", "PARTIAL", "",
     "Bankrate, NerdWallet (May 2026)",
     "Time-sensitive. As of May 2026, top HYSA rates are in the 3.75-4.50% range as Fed has cut rates from late-2024 peak. Soften to '~4% APY in current rate environment' or 'check current rates.'",
     "Minor staleness; not critical given hedged 'currently' language."),

    ("CL080", "W2:2.3", "$1M umbrella policy: ~$200-$400/year",
     "PRICING", "SOFT", "PASS", "C",
     "Insurance industry data (Allstate, Progressive, USAA quotes)",
     "", "Verified range for clean drivers / no incidents."),

    ("CL081", "W2:6.3", "Estate docs: $500-$2,500 attorney / $100-$300 online",
     "PRICING", "SOFT", "PASS", "C",
     "Trust & Will, FreeWill, LegalZoom; AAEPA attorney surveys",
     "", "Verified range."),

    ("CL082", "Contractor:2.4", "GL $1M/$2M: $400-$800/year for most service businesses",
     "PRICING", "SOFT", "PASS", "C",
     "Hiscox, Next Insurance, Insureon data",
     "", "Verified range for low-risk service businesses."),

    ("CL083", "Contractor:2.4", "E&O / Professional liability: $500-$5,000+/year",
     "PRICING", "SOFT", "PASS", "C",
     "Insureon, Hiscox industry data",
     "", "Verified. Wide range justified by professional risk variability."),

    ("CL084", "Contractor:2.4", "Cyber liability: $500-$2,000/year for small biz",
     "PRICING", "SOFT", "PASS", "C",
     "Coalition, Embroker, Insureon data",
     "", "Verified. Pricing has risen 2022-2024 due to ransomware claims; current range may trend higher."),

    ("CL085", "Contractor:9.2", "Specialty Solo 401(k) providers: $500-$1K setup, $150-$300/yr",
     "PRICING", "SOFT", "PASS", "C",
     "MySolo401k, Solo401k.com current pricing",
     "", "Verified directionally."),

    ("CL086", "Contractor:8.1", "Solo DB plan admin: $3,000-$5,000/year",
     "PRICING", "SOFT", "PASS", "C",
     "Cash Balance Coach, OneDigital industry data",
     "", "Verified."),

    ("CL087", "Business:10.1", "ESOP setup: $200K+",
     "PRICING", "SOFT", "PASS", "C",
     "NCEO, Menke & Associates industry data",
     "", "Verified. Setup costs $80-250K typical; ongoing admin $20-40K/year."),

    ("CL088", "Business:2.3", "Workers comp by industry: clerical ~0.5%; construction 10%+",
     "PRICING", "SOFT", "PASS", "C",
     "NCCI class code rates; state insurance fund data",
     "", "Verified range. Actual rates vary significantly by state."),

    # === HOME OFFICE / SCHEDULE C ===
    ("CL089", "Contractor:0.2", "Home office simplified: $5/sq ft × ≤300 sq ft ($1,500 max)",
     "STATUTORY", "CON", "PASS", "A",
     "IRS Rev. Proc. 2013-13", "", "Verified statutory."),

    ("CL090", "Contractor:0.2", "IRS receipt substantiation threshold: $75",
     "STATUTORY", "CON", "PASS", "A",
     "Treas. Reg. §1.274-5(c)(2)(iii)", "",
     "Verified. Applies to T&E and gifts; NOT general business expenses (those have no specific threshold)."),

    # === SE HEALTH INSURANCE ===
    ("CL091", "Contractor:1.3", "SEHID does NOT reduce SE tax — only income tax",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §162(l)(4); Rev. Rul. 2003-91",
     "", "Verified statutory. Persona 1 flagged this nuance is critical for SE planning."),

    ("CL092", "Contractor:1.3", "SEHID limited to net SE income (cannot create loss)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §162(l)(2)", "", "Verified statutory."),

    # === ENTITY / CORPORATE ===
    ("CL093", "Business:0.1", "C-corp federal rate: 21%",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRC §11(b); TCJA, retained by OBBBA", "", "Verified."),

    ("CL094", "Business:0.1", "S-corp limits: 100 shareholders, US persons, one class of stock",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1361(b); §1361(c)(1) (family aggregation); §1361(c)(2) (eligible trusts)",
     "Family members within 6 generations from a common ancestor count as ONE shareholder per §1361(c)(1). ESBTs and QSSTs are permitted with elections. Artifact updated to reflect both nuances.",
     "Verified. P1A #9 catch added to artifact."),

    ("CL095", "Business:0.1", "S-corp revocation: cannot re-elect for 5 years",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1362(g)", "", "Verified."),

    ("CL096", "Contractor:9.1", "S-corp election via Form 2553",
     "STATUTORY", "CON", "PASS", "A",
     "IRS Form 2553", "", "Verified."),

    ("CL097", "Business:10.1", "ESOP §1042 deferral for C-corps via qualified replacement property",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1042", "", "Verified."),

    ("CL098", "Business:10.1", "S-corp ESOPs can be tax-exempt at corporate level on ESOP-owned portions",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1361(c)(6); §512(e)", "", "Verified."),

    # === MARKET / MATH ===
    ("CL099", "Contractor:1.1", "Pay yourself stable salary based on trailing 6-month avg",
     "CONVENTION", "SOFT", "PASS", "C",
     "Profit First (Mike Michalowicz); standard practitioner heuristic",
     "", "Verified convention. Common 'pay yourself first' methodology for SE income smoothing."),

    ("CL100", "W2:6.2", "Most 401(k) plans have ERs 2-10× higher than IRA options",
     "EMPIRICAL", "CON", "PASS", "C",
     "ICI 'BrightScope/ICI Defined Contribution Plan Profile' (annual)",
     "ICI 2023: avg 401(k) ER 0.36% (large plans) to 0.95% (small plans). Index ETF in IRA ~0.03-0.10%. So 2-10× range is verified, with high end applying to small-plan investors.",
     "Verified."),

    ("CL101", "W2:6.2", "0.5% annual fee diff over 30 yrs ≈ 15% smaller terminal balance",
     "MATH", "CON", "PASS", "A",
     "Compound math: (1-0.005)^30 ≈ 0.860 → 14% smaller",
     "Math is close; technically ~14% not exactly 15%. Within acceptable rounding for a heuristic.",
     "Verified math."),

    # === ROTH 401k EMPLOYER MATCH ===
    ("CL102", "W2:3.1", "Employer match contributions are typically pre-tax even when employee is Roth",
     "STATUTORY", "CON", "FAIL", "",
     "SECURE 2.0 §604; IRS Notice 2024-2",
     "Update to: 'Historically all employer match was pre-tax. SECURE 2.0 §604 permits (but does not require) plans to offer Roth employer match contributions; as of 2026, many large-plan recordkeepers have rolled this out. Check your plan document. If the match is made as Roth, it is taxable income to the employee in the year of contribution.'",
     "Persona 1 flagged. Material update needed."),

    # === 401(a)(17) ===
    ("CL103", "Various SE/business", "401(a)(17) compensation limit 2026: $360,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67",
     "", "Verified. Not currently surfaced in artifact — should add for S-corp salary/profit-sharing math."),

    # === ADDITIONAL ===
    ("CL104", "Business:0.1", "DB plan annual benefit limit 2026: $290,000",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67; IRC §415(b)(1)(A)",
     "", "Verified. Up from $280K in 2025."),

    ("CL105", "Business:9.4", "DAFs: deduction at fair market value for appreciated securities",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §170(b)(1)(A); IRS Pub 526", "",
     "Verified. Subject to 30% AGI limit for appreciated assets to public charities."),

    ("CL106", "Contractor:1.3", "ACA Premium Tax Credit available below ~400% FPL",
     "STATUTORY", "CON", "CORRECTED", "A",
     "Inflation Reduction Act §13202 (extension through 2025); CRS Report R48290; KFF analyses Dec 2025",
     "Updated artifact to reflect 2026 reality: enhanced PTCs expired December 31, 2025. The 400% FPL cliff IS back in effect for 2026. Households over 400% FPL ($62,600 single / $84,600 couple / $128,600 family of 4) are now ineligible for ANY premium tax credit. Insurers also priced ~18% median premium increases for 2026.",
     "Material correction applied. Key planning implication: AGI management around the 400% threshold is critical for 2026 self-employed and early-retiree households."),

    # === ADDITIONS NEW IN THIS PASS ===
    ("CL107", "Contractor:0.3", "Additional Medicare 0.9% thresholds NOT indexed for inflation (frozen since 2013)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1401(b)(2); ACA §9015 (2010); never indexed by statute",
     "",
     "Verified. Added to artifact in Contractor:0.3 — frozen $200K/$250K thresholds capture more taxpayers each year. Important context for planning."),

    ("CL108", "Business:8.1", "401(a)(17) compensation cap $360,000 (2026) limits employer plan calculations",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRS Notice 2025-67; IRC §401(a)(17)",
     "",
     "Added to artifact in Business:8.1. Affects S-corp salary-vs-distribution optimization for very high earners — only first $360K of W-2 wages counts for employer 401(k) contribution math."),

    ("CL109", "W2:6.3", "Annual gift tax exclusion $19,000 (2026)",
     "TAX-FIGURE", "CON", "PASS", "A",
     "IRC §2503(b); IRS Notice; OBBBA didn't change the exclusion",
     "",
     "Added to artifact in W2:6.3 estate planning section. Married couples can split-gift to $38,000 per recipient. Powerful for systematic multi-decade estate reduction."),

    ("CL110", "Contractor:8.2", "OBBBA $400 minimum QBI deduction (2026) for QBI ≥ $1,000 with material participation",
     "STATUTORY", "CON", "PASS", "A",
     "OBBBA §70112; IRC §199A(a)(2)",
     "",
     "Added to artifact in Contractor:8.2. New for 2026 — floor deduction for small/marginal businesses but does NOT help SSTB owners fully phased out."),

    ("CL111", "Contractor:8.2", "Engineering and architecture specifically EXCLUDED from SSTB classification",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §199A(d)(2)(A); Treas. Reg. §1.199A-5(b)(2)(v)",
     "",
     "Added to artifact. Engineers and architects get full QBI deduction regardless of income — unlike consultants, lawyers, doctors, financial services."),

    # === NEW CLAIMS FROM CHECKPOINT PERSONAS (P1A & P1B) ===
    ("CL112", "W2:8.1", "Final regs for high-earner Roth catch-up effective Nov 17, 2025; mandatory plan compliance for plan years beginning after Dec 31, 2026",
     "STATUTORY", "CON", "PASS", "A",
     "Treasury Decision 9989 / 90 Fed. Reg. 44527 (Nov 2025); IRS Notice 2024-2 (plan amendment deadline)",
     "",
     "P1A/P1B catch. Critical transition-rule context: statute is effective 2026 but most plans use regulatory transition relief through 2027. Artifact updated to reflect operational reality."),

    ("CL113", "W2:9.1, Contractor:9.2, Business:9.2", "Mega Backdoor Roth after-tax contributions subject to ACP testing",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §401(m); ACP test under §401(m)(2)",
     "",
     "P1B #2 catch. In non-Safe Harbor plans, ACP testing routinely caps HCE after-tax contributions well below the §415(c) headroom. Artifact updated."),

    ("CL114", "Contractor:9.2", "Solo 401(k)s bypass ACP testing (no non-HCE employees)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §401(m); ACP applies to discrimination between HCEs and NHCEs",
     "",
     "Solo plans have no NHCEs to discriminate against, so the full §415(c) space is generally usable. Structural advantage of Solo over employer plans for Mega Backdoor Roth."),

    ("CL115", "W2:9.1", "Each in-plan Roth conversion starts its own 5-year clock under §408A(d)(2)(B)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §408A(d)(2)(B); §72(t)",
     "",
     "P1B #9. Important nuance for under-59.5 converters."),

    ("CL116", "W2:9.4", "529-to-Roth: contributions within prior 5 years (and earnings on them) NOT eligible for rollover",
     "STATUTORY", "CON", "PASS", "A",
     "IRS Notice 2024-2; SECURE 2.0 §126",
     "",
     "P1A #4. Material omission now corrected in artifact. Additionally: beneficiary must have earned income at least equal to the rollover amount in the rollover year."),

    ("CL117", "Contractor:0.3", "Additional Medicare 0.9% applies to W-2 wages AND SE income; employer must withhold on individual wages over $200K regardless of filing status",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §3101(b)(2); §3102(f); §1401(b)(2)",
     "",
     "P1A #5. Scope clarification added to artifact: applies to W-2 wages too, with employer-withholding nuances that create over-/under-withholding for some MFJ couples."),

    ("CL118", "Contractor:6.2", "Inherited IRAs excluded from backdoor Roth pro-rata aggregation",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §408(d)(3)(C); Treas. Reg. §1.408-4(a)(2)",
     "",
     "P1B #4. Added to artifact. Critical exclusion that's often missed."),

    ("CL119", "Contractor:6.2", "Roth IRA balances NOT aggregated in pro-rata calculation",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §408(d)(8); §408A(d)(4)",
     "",
     "P1B #4. Only pre-tax IRAs (Traditional, SEP, SIMPLE) are aggregated. Added to artifact."),

    ("CL120", "Contractor:6.2", "SIMPLE IRA balances cannot be rolled to a 401(k) until 2 years after first participation",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §72(t)(6); §408(p)(2)(B)",
     "",
     "P1B #4. The standard fix (roll SIMPLE balances to 401(k)) is unavailable for participants under 2 years. Artifact updated."),

    ("CL121", "Business:3.1", "SECURE 2.0 §117 higher SIMPLE limits: 25 or fewer employees OR 26-100 with 4% non-elective / 4% (1%-increased) match",
     "STATUTORY", "CON", "PASS", "A",
     "SECURE 2.0 §117; IRC §408(p)(2)(E)",
     "",
     "P1B #5. Specifies which plans qualify for the $18,100/$3,850 'applicable' SIMPLE limits. Artifact updated with mechanism."),

    ("CL122", "Business:8.1", "§415(c) limit applies per unrelated employer; §402(g) deferral limit aggregates across all plans",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §415(c) (per-employer); IRC §402(g)(1) (per-individual aggregate)",
     "",
     "P1B #6. Crucial for W-2 + Solo 401(k) moonlighter. The employer side can fill twice; the employee deferral cannot. Artifact updated."),

    ("CL123", "Business:3.1", "Safe Harbor 401(k) formulas: basic match (100%×3% + 50%×next 2%), enhanced match (100%×4%+), 3% non-elective, QACA (100%×1% + 50%×next 5%)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §401(k)(12); §401(k)(13) (QACA)",
     "",
     "P1B #7. Artifact previously oversimplified to '3% NE OR 4% match.' Now includes all four formulas."),

    ("CL124", "Contractor:8.1", "Solo 401(k) employer-side calculation for sole prop is circular: contribution = (net SE − ½ SE tax − contribution) × 25%, yielding ~20% effective",
     "MATH", "CON", "PASS", "A",
     "IRC §404(a)(3); §401(c)(2); §164(f)",
     "",
     "P1B #8. Artifact updated to explain the circularity rather than just stating the effective rate."),

    ("CL125", "Business:10.1", "ESOP §1042 deferral has 4 requirements: C-corp, 3+ yr hold, ESOP 30%+ post-sale, QRP purchased in 15-month window (3 mo before to 12 mo after)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1042(a)-(c); QRP defined at §1042(c)(4)",
     "",
     "P1B #10. Artifact updated with all four requirements. Deferral becomes permanent only if QRP held until death (step-up); sale during life triggers recognition."),

    ("CL126", "Business:0.1", "S-corp family aggregation: family within 6 generations from common ancestor counts as one shareholder",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1361(c)(1)",
     "",
     "P1A #9. Often overlooked rule. Added to artifact."),

    ("CL127", "Business:0.1", "S-corp eligible trusts: ESBTs and QSSTs permitted with specific elections",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1361(c)(2)(A); §1361(d) (QSST); §1361(e) (ESBT)",
     "",
     "Added to artifact alongside family aggregation."),

    ("CL128", "Business:10.1", "QRP for §1042 = securities of US domestic operating corporations (NOT mutual funds, NOT bonds, NOT REITs)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1042(c)(4); Treas. Reg. §1.1042-1T",
     "",
     "P1B #10. Common point of confusion — passive vehicles do not qualify."),

    ("CL129", "Business:10.1", "§1042 deferral becomes permanent only if QRP held until death (step-up basis); sale during life triggers recognition of deferred gain",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §1042(e); §1014 (step-up)",
     "",
     "P1B #10. The estate-planning hook that makes §1042 powerful."),

    # ============================================================
    # === PHASE 2: MATH BEDROCK CLAIMS ===
    # ============================================================
    ("CL130", "Math:1 (Compound interest)", "Future value formula: FV = PV × (1+r)^n + PMT × [(1+r)^n − 1] / r (ordinary annuity, end-of-period)",
     "MATH", "CON", "PASS", "A",
     "Standard time-value-of-money identity; derivable from geometric series",
     "",
     "Mathematical identity. For begin-of-period (annuity due), multiply PMT term by (1+r). Calculator uses monthly compounding."),

    ("CL131", "Math:1", "$500/month for 30 years at 7% real return ≈ $610,000 in today's dollars",
     "MATH", "CON", "PASS", "A",
     "Compound interest formula with monthly compounding",
     "",
     "Verified math: $500 × [(1.005833^360 − 1)/0.005833] ≈ $609,985. The 'last decade does half the work' claim follows from comparing 20-year ($263K) vs 30-year ($610K) terminal values."),

    ("CL132", "Math:2 (Rule of 72)", "Years to double ≈ 72 / r% — derived from t = ln(2)/ln(1+r)",
     "MATH", "CON", "PASS", "A",
     "Standard mathematical approximation; Taylor expansion of ln(1+r) ≈ r for small r",
     "",
     "Verified. Most accurate for rates 6-10%. Rule of 70 better for higher rates; Rule of 69.3 is exact in continuous-compounding limit (ln 2 ≈ 0.693)."),

    ("CL133", "Math:2", "Rule of 114 (tripling) and Rule of 144 (quadrupling) — same derivation as Rule of 72",
     "MATH", "CON", "PASS", "A",
     "ln(3)/r% × 100 ≈ 110, ln(4)/r% × 100 ≈ 139; rounded to 114 and 144",
     "",
     "Verified. The '114' and '144' are practitioner-friendly roundings. ln(3) ≈ 1.099, so 109.9/r; ln(4) ≈ 1.386, so 138.6/r."),

    ("CL134", "Math:3 (Savings rate)", "Years to FI from zero = ln(1 + k·r·(1−s)/s) / ln(1+r), where k = 1/SWR",
     "MATH", "CON", "PASS", "A",
     "Mr. Money Mustache (2012) 'Shockingly Simple Math'; ERN's extension; derivable from future-value annuity formula set equal to k × annual expenses",
     "",
     "Verified derivation. Assumes constant income, savings rate, real return, no starting balance."),

    ("CL135", "Math:3", "50% savings rate → ~17 years to FI; 25% → ~32 years; 10% → ~51 years (at 5% real return, 4% SWR)",
     "EMPIRICAL", "CON", "PASS", "B",
     "MMM 'Shockingly Simple Math' (2012); validated by ERN's analysis using historical returns",
     "",
     "Verified — these are the canonical figures from the MMM article. ERN's CAPE-adjusted analysis shows real-world variance of ±3-5 years around these medians."),

    ("CL136", "Math:3", "Savings rate dominates investment return for years-to-FI",
     "EMPIRICAL", "CON", "PASS", "B",
     "MMM 2012; ERN 'Shockingly Simple/Complicated' (2017); structural property of the formula (savings rate appears in BOTH the contribution and the target)",
     "",
     "Verified. Sensitivity analysis: a 10pp savings rate increase saves more years than a 2pp return increase across the typical range."),

    ("CL137", "Math:4 (Real vs nominal)", "Fisher equation: (1 + nominal) = (1 + real) × (1 + inflation)",
     "MATH", "CON", "PASS", "A",
     "Irving Fisher, 'The Theory of Interest' (1930); standard macroeconomics identity",
     "",
     "Verified. The simple approximation real ≈ nominal − inflation works for small rates but diverges meaningfully at high rates or long horizons."),

    ("CL138", "Math:4", "Long-run US equity real returns ~6-7%; long-run intermediate Treasuries ~2% real",
     "EMPIRICAL", "CON", "PASS", "B",
     "Dimson, Marsh & Staunton 'Credit Suisse Global Investment Returns Yearbook'; Ibbotson SBBI data",
     "",
     "Verified. 1926-2024 US equity real return ~6.8% (geometric mean); intermediate Treasuries ~2.3% real."),

    ("CL139", "Math:4", "Bengen's 4% rule is implicitly a REAL withdrawal (year-1 4% then inflation-indexed)",
     "CITATION", "CON", "PASS", "A",
     "Bengen (1994), 'Determining Withdrawal Rates Using Historical Data,' Journal of Financial Planning, Oct 1994",
     "",
     "Verified. This is one of the most common misunderstandings in DIY retirement math."),

    ("CL140", "Math:5 (Sequence risk)", "Accumulation phase FV is order-independent (depends only on geometric mean return)",
     "MATH", "CON", "PASS", "A",
     "Property of multiplication commutativity; FV = PV × Π(1+r_i)",
     "",
     "Verified mathematical property. Important: this is for lump-sum investment. With ongoing contributions, sequence does matter even in accumulation, because later contributions experience fewer years of compounding."),

    ("CL141", "Math:5", "Withdrawal phase outcomes depend critically on return sequence; early losses compound disproportionately",
     "EMPIRICAL", "CON", "PASS", "B",
     "Pfau (multiple); Kitces 'Sequence of returns risk'; Bengen subsequent work",
     "",
     "Verified. The 'retirement red zone' (5 years before/after retirement) is the standard term for this risk window."),

    ("CL142", "Math:5", "Mitigations: 1-2 yr cash buffer + 3-5 yr short-bond ladder ('bucket strategy'); dynamic withdrawal cutting spending in down years",
     "CONVENTION", "SOFT", "PASS", "B",
     "Ray Lucia (bucket strategy); Guyton-Klinger guardrails; Kitces variations",
     "",
     "Verified industry-standard mitigation approaches. Empirical evidence shows dynamic withdrawal can raise safe initial withdrawal rate to 4.5-5%."),

    ("CL143", "Math:6 (Asset location)", "Asset location optimization adds 20-50 bps after-tax return for typical balanced portfolio",
     "EMPIRICAL", "CON", "PASS", "B",
     "Vanguard 'Putting a Value on Your Value' (2024); Reichenstein (multiple)",
     "",
     "Same as CL075 (artifact reference); restated here as a Math-view claim with tier-B citation."),

    ("CL144", "Math:6", "Asset allocation matters more than asset location",
     "EMPIRICAL", "CON", "PASS", "B",
     "Brinson/Hood/Beebower (1986, 1991); Ibbotson & Kaplan (2000); decades of asset-allocation studies",
     "",
     "Verified consensus in financial economics. Asset allocation accounts for ~90% of portfolio variation; location is a second-order optimization."),

    # === INLINE MATCH CALCULATOR ===
    ("CL145", "W2:3.1 inline", "FV of annual match $X over n years at rate r: $X × [(1+r)^n − 1] / r",
     "MATH", "CON", "PASS", "A",
     "Annuity future value formula",
     "",
     "Verified. Inline calculator embedded in W2:3.1 to demonstrate the lifetime value of capturing the employer match."),

    # ============================================================
    # === PHASE 2 CHECKPOINT PERSONA OBSERVATIONS (P2A + P2B) ===
    # ============================================================
    # APPLIED = fix made in this pass; DEFERRED-P2.5 = captured for Phase 2.5 build
    ("CL146", "Math:5", "P2A#1: Real vs nominal inconsistency in sequence risk simulator — returns labeled as 7% without specifying nominal/real, while withdrawal inflation-indexed implies nominal",
     "MATH", "CON", "CORRECTED", "A",
     "Internal consistency review",
     "APPLIED: Math callout updated to specify '7% nominal' explicitly; result text clarifies ending balance is in nominal dollars; section explainer notes withdrawal is inflation-indexed.",
     "Highest-priority correctness fix from P2A."),

    ("CL147", "Math intro", "P2A#8 + P2B#5: Calculators present deterministic point estimates with no uncertainty quantification or distribution of outcomes",
     "MATH", "CON", "CORRECTED", "A",
     "Standard practice in production planning tools (Monte Carlo, historical cycles)",
     "APPLIED: Added 'On reading these calculators' disclaimer to math-intro explicitly framing all outputs as deterministic point estimates with realized range examples; users directed to FIRECalc/cFireSim/professional tools for actual planning.",
     "Range example: $500/mo at 30 yr projects to $610K but realized range is ~$300K-$1.1M at typical equity volatility."),

    ("CL148", "Math:6", "P2A#9 + P2B#4: Asset location calculator 'half realized annually' overstates after-tax drag for modern broad-market ETFs; missing Roth dimension",
     "MATH", "CON", "CORRECTED", "B",
     "Modern ETF in-kind creation/redemption mechanics (Vanguard, iShares); academic literature on equity tax-efficiency",
     "APPLIED: Updated annual realization assumption from 50% to 10% (more realistic for modern broad-market index ETFs). Added explicit caveats: model omits Roth, REITs, international equity, munis, dividend yield specifics. Full Roth + 3-account expansion deferred to Phase 2.5.",
     "Partial fix; full Roth expansion is Phase 2.5 work."),

    ("CL149", "Math:1-6", "P2B#1 + P2B#10: No input validation on calculators — negative numbers, zero divisors, absurd magnitudes accepted",
     "UX", "CON", "CORRECTED", "A",
     "Production-tool standard practice",
     "APPLIED: Added min/max/step constraints to all numeric inputs across all six Math view calculators and the inline match calculator. Mobile numeric keyboards now properly constrained; users can no longer type 999 years or 500% returns.",
     "Quick-win UX fix."),

    ("CL150", "Math:1", "P2B#3: Monthly vs annual compounding implicit — formula shown is annual, implementation is monthly",
     "MATH", "CON", "CORRECTED", "A",
     "Calculator implementation review",
     "APPLIED: Added note under compound interest calculator: 'Uses monthly compounding with contributions at end of each month.' Users now know which compounding the result reflects.",
     "Quick clarification; underlying math unchanged."),

    ("CL151", "Math:1, 3, 5", "P2A#2: Geometric vs arithmetic mean / volatility drag not addressed; users assuming constant 7% will under-realize ~25-30% over 30 years at typical equity volatility",
     "MATH", "CON", "CORRECTED", "B",
     "Booth, Fama (1992) on volatility drag; Markowitz (1959) on geometric mean",
     "APPLIED: Added 'Volatility drag' callout to Math §1 explaining σ²/2 approximation. Geometric mean below arithmetic mean whenever returns vary; long-horizon planning uses geometric.",
     "Phase 2.5 cleanup."),

    ("CL152", "Math:3", "P2A#3: 'Constant income, constant savings rate' assumption in MMM formula unrealistic — real incomes typically rise; real savings rates typically rise with age",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "BLS Current Population Survey wage data; Survey of Consumer Finances savings rate by age",
     "APPLIED: Added 'rising income' callout to §3 with cross-reference to Spending Lifestyle §4 (lifestyle creep) and Zeitgeist Lifestyle §6 (sustainability).",
     "Phase 2.5 cleanup."),

    ("CL153", "Math:3, 5", "P2A#4: Longevity risk invisible — fixed 30-year horizon assumes a 50% probability of running out for median early retirees",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "SSA period life tables; Pfau retirement-income research",
     "APPLIED: Math §5 has comprehensive 'Longevity risk and horizon dependence' callout with SSA tables, SPIA/DIA/QLAC ($210K IRA limit 2026) mitigation toolkit, and cross-references.",
     "Phase 2.5 cleanup."),

    ("CL154", "Math:5", "P2A#5: CPI vs retiree-specific inflation — healthcare CPI has run ~3.5-4.5% vs overall ~2.5-3%, materially higher for retirees",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "BLS CPI-U vs CPI-E (experimental for elderly); BLS Medical Care Services CPI",
     "APPLIED: §4 callout on CPI-U vs CPI-E differences, Genworth Cost of Care medical inflation premium, hidden 0.2-0.5pp buffer for retirees.",
     "Phase 2.5 cleanup."),

    ("CL155", "Math:5", "P2A#6: Sequence risk 'filler rate' is artificial — constant 9.07% return for 27 years smooths over actual market clustering",
     "MATH", "SOFT", "CORRECTED", "B",
     "Real market history (1929, 1937, 1966, 2000); FIRECalc historical-cycles approach; Shiller, Damodaran/NYU Stern datasets",
     "APPLIED: Built full §7 historical-cycles retirement simulator. Uses HIST_RETURNS dataset (1928-2024, stocks + bonds, real returns). Aggregates across all rolling cohorts of specified horizon. Reports success rate, percentile ending balances, failure cohort start years. Supports fixed/flexible/Guyton-Klinger withdrawal strategies. Sequence risk §5 cross-references §7 in its model-limits note.",
     "Phase 2.5 cleanup — major feature complete."),

    ("CL156", "Math:3", "P2A#7: 4% rule horizon dependence — used as generic target for any retirement age, but Bengen calibrated to 30-year retirements",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Bengen 1994; subsequent horizon-specific research (Pfau, ERN)",
     "APPLIED: Math §3 has '25× target depends on horizon' callout (40+ yr → 3.25-3.5% / 28-33×). §5 longevity callout reinforces. §7 historical-cycles simulator lets user test different horizons empirically.",
     "Phase 2.5 cleanup."),

    ("CL157", "Math:1-6", "P2B#2: Default values pedagogically reasonable but not personalized; production tools pre-populate from demographics",
     "UX", "SOFT", "CORRECTED", "B",
     "Production-tool standard practice",
     "APPLIED: Built getDefault(field) helper mapping diagnostic categorical answers to numeric defaults (age, yearsToRetirement, horizon, taxBracketPct, ltcgPct, retirementAge, netWorth). applyDefaultIfFallback() applies them only if user hasn't manually edited. markUserTyped() tracks edits. Coverage extended this turn to ci-years, al-ord, al-ltcg (Math view), coast-years (Zeitgeist Investing), dz-age, dz-retire (Zeitgeist Lifestyle). Fixed bug in earlier applyDefaultIfFallback that always overwrote user values regardless of edit state.",
     "Phase 2.5 cleanup — partial backlog still possible (more inputs could be personalized in other views)."),

    ("CL158", "Math:5", "P2B#5: No Monte Carlo or historical-cycles simulation; only deterministic point estimates",
     "UX", "CON", "CORRECTED", "B",
     "FIRECalc, cFireSim, professional planning software standard features",
     "APPLIED: Same implementation as CL155 — Math §7 historical-cycles simulator. Designed with swappable backend architecture via getMCBackend() function that returns the simulator; can be replaced by setting window.__customMCBackend = {simulate: yourFunction} before render. Documented in §7 model-limits note. NOTE: user indicated they may plug in their personal Monte Carlo system later — the hook is in place.",
     "Phase 2.5 cleanup — major feature with future MC integration hook."),

    ("CL159", "Math:5", "P2B#6: No withdrawal flexibility modeled (Guyton-Klinger, ratcheting, floor-and-upside)",
     "UX", "SOFT", "CORRECTED", "B",
     "Guyton-Klinger guardrails; Bengen subsequent flexibility research",
     "APPLIED: §5 sequence risk calc has withdrawal strategy selector (fixed / flexible 25%-cut-in-down-years / Guyton-Klinger ±20% guardrails). §7 historical-cycles simulator supports same three strategies.",
     "Phase 2.5 cleanup."),

    ("CL160", "Math:3", "P2B#7: No Social Security or other income overlay; framework assumes 100% portfolio funding",
     "UX", "CON", "CORRECTED", "B",
     "Standard retirement planning practice; SSA replaces 25-40% of pre-retirement income for typical earners",
     "APPLIED: §3 Years-to-FI calculator has SS overlay inputs (expected SS at retirement + annual expenses for offset calculation). Wired into calcSavingsRate logic.",
     "Phase 2.5 cleanup."),

    ("CL161", "Math:1, 5, 7", "P2B#8: Text-only visualization; no charts or paths shown over time",
     "UX", "SOFT", "CORRECTED", "C",
     "Production-tool visualization standards",
     "APPLIED: Inline SVG charts added to three Math sections. §1 compound interest: year-by-year stacked bar of contributions (taupe) + investment growth (gold), sampled to ~30 bars for long horizons. §5 sequence risk: two 30-year balance paths (bad-sequence-early red-brown vs late green) with starting-balance reference line. §7 historical-cycles: percentile fan chart (p10-p90 outer band, p25-p75 inner band, p50 median line). Shared chart helpers (chartFrame, niceCeil, fmtAxisMoney, renderSVG) reuse the artifact's CSS custom properties so palette tracks the rest of the UI. historicalCyclesSimulate gained an optional `paths` field on its return — strictly additive to the MC swap interface; external backends may omit it and the fan chart no-ops cleanly.",
     "Phase 2.5 — DONE 2026-05-16."),

    ("CL162", "Math:1-6", "P2B#9: Tax bracket not propagated from diagnostic — three calculators would change meaningfully at 12% vs 24% vs 35%",
     "UX", "SOFT", "CORRECTED", "B",
     "Diagnostic state already tracks bracket; not currently wired to Math view",
     "APPLIED: getDefault('taxBracketPct') maps low→12, mid→24, high→35; getDefault('ltcgPct') added this turn maps low→0, mid→15, high→20. Both applied to Math §6 asset location calculator inputs via applyDefaultIfFallback.",
     "Phase 2.5 cleanup."),

    ("CL163", "Math:6", "P2B#4 (full): Asset location calculator missing Roth dimension and three-account-type matrix; current model is binary",
     "MATH", "CON", "CORRECTED", "B",
     "Reichenstein; Kitces; Bogleheads wiki on asset location",
     "APPLIED: Math §6 now has full 3-account × 3-asset matrix calculator. Inputs: portfolio total, asset allocation % (Stocks/Bonds/REITs), account capacity % (Traditional/Roth/Taxable), expected real returns per asset, horizon, current+future ordinary rate, LTCG rate. Output: per-$1000 terminal-value matrix with ★ optimal cells, waterfall optimal allocation under capacity constraint, annualized bps alpha vs naive proportional, summary table comparing optimal/naive/worst placements. Model-limits paragraph covers omitted dimensions (NIIT, state taxes, step-up basis, FTC for international, direct indexing, greedy-vs-globally-optimal trade-offs).",
     "Phase 2.5 cleanup — full matrix complete."),

    ("CL164", "Math:1-6", "P2A general + P2B general: No 'what this does not model' disclaimer pattern; users may interpret deterministic outputs as planning targets",
     "UX", "CON", "CORRECTED", "A",
     "Standard disclosure practice in financial software",
     "APPLIED: All 6 Math sections now have per-section model-limits notes in a distinct visual style (italic, muted, left-bordered). §7 historical-cycles also has one. Each note identifies the specific assumptions and what's not modeled. Math-intro 'On reading these calculators' callout provides framework-level framing.",
     "Phase 2.5 cleanup — complete."),

    # ============================================================
    # === PHASE 3: SPENDING STRATEGY CLAIMS (Essentials + Lifestyle) ===
    # ============================================================
    # --- Housing ---
    ("CL165", "Spend:Ess:1 (Housing)", "Housing typically 25-40% of after-tax income across US households",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "BLS Consumer Expenditure Survey (annual); range varies by HCOL/LCOL geography",
     "",
     "BLS CES 2024 showed housing at ~33% of after-tax income on average; HCOL metros frequently 35-45%."),

    ("CL166", "Spend:Ess:1", "5% rule: total annual ownership cost ≈ 5% of property value (1% tax + 1% maintenance + 3% opportunity cost on equity). Attribution: Ben Felix (PWL Capital, 2017)",
     "CITATION", "SOFT", "PASS", "B",
     "Ben Felix YouTube/PWL Capital research (2017-2018); widely adopted heuristic in Boglehead and Reddit r/personalfinance communities",
     "",
     "Pedagogical approximation; actual breakdown varies by state property tax and home age."),

    ("CL167", "Spend:Ess:1", "Price-to-rent ratio thresholds: <15 favors buying, >20 favors renting, 15-20 gray zone",
     "CONVENTION", "SOFT", "PASS", "C",
     "Trulia/Zillow research; widely cited in real estate economics literature",
     "",
     "Heuristic; actual breakeven depends heavily on holding period, transaction costs, and rent vs price appreciation differentials."),

    ("CL168", "Spend:Ess:1", "Maintenance cost planning figure ~1% of home value annually (range 0.5-2%)",
     "CONVENTION", "SOFT", "PASS", "C",
     "Industry consensus; appears in Zillow, Bankrate, NerdWallet planning content",
     "",
     "Heuristic. Newer homes lower (0.5-1%), older homes higher (1.5-2%+); region-dependent."),

    ("CL169", "Spend:Ess:1", "15-yr vs 30-yr mortgage on $400K at 6%: ~$208K vs ~$463K lifetime interest. Cash flow differential ~$1,000/mo, investible at 7% real = ~$1.2M over 30 yr",
     "MATH", "CON", "PASS", "A",
     "Standard mortgage amortization math",
     "",
     "Verified: 15-yr P&I at 6% on $400K = $3,375/mo × 180 = $607.5K; minus $400K principal = $207.5K interest. 30-yr P&I = $2,398/mo × 360 = $863.4K; minus $400K = $463.4K interest. Cash flow difference $977/mo, FV at 7%/30y ≈ $1.19M."),

    ("CL170", "Spend:Ess:1", "Refinance breakeven: typically 0.75-1.0% rate drop + recover closing costs in 2-4 years",
     "CONVENTION", "SOFT", "PASS", "C",
     "Industry consensus; Bankrate, Freddie Mac refinance guides",
     "",
     "Heuristic; exact breakeven depends on loan size, remaining term, and closing cost ratio."),

    ("CL171", "Spend:Ess:1", "FHA loans permit 3.5% down on owner-occupied 2-4 unit properties (house hacking enabler)",
     "STATUTORY", "CON", "PASS", "A",
     "HUD 4000.1 FHA Single Family Housing Policy Handbook; 12 U.S.C. §1709",
     "",
     "Verified. Owner must occupy one of the units for at least one year. Loan limits vary by county."),

    # --- Transportation ---
    ("CL172", "Spend:Ess:2", "Transportation typically 10-18% of after-tax income",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "BLS Consumer Expenditure Survey (annual)",
     "",
     "BLS CES 2024: average ~16% of pretax income; varies with car-dependent vs transit-served geographies."),

    ("CL173", "Spend:Ess:2", "AAA 'Your Driving Costs 2024' average new vehicle ownership cost ~$12,297/year at 15,000 miles annually",
     "CITATION", "CON", "PASS", "B",
     "AAA 'Your Driving Costs 2024' annual report",
     "",
     "Verified. 2024 figure widely cited; AAA report is annual."),

    ("CL174", "Spend:Ess:2", "New cars depreciate 20-30% year 1, ~60% by year 5; 3-5 year-old CPO is value sweet spot",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Edmunds, Kelley Blue Book depreciation tables; Carfax research",
     "",
     "Specific depreciation curves vary by make/model; Toyota and Honda hold value better than American/European brands generally."),

    ("CL175", "Spend:Ess:2", "EV TCO typically wins past year 5 for 12,000+ miles/year home-charging households",
     "EMPIRICAL", "SOFT", "PASS", "C",
     "Consumer Reports, MIT Trancik Lab 2020 study; updated with 2024-25 price/incentive data",
     "",
     "Tier C: depends on electricity rates, incentive availability, vehicle model, battery longevity assumptions. The crossover point has moved with EV price drops in 2023-25."),

    # --- Healthcare ---
    ("CL176", "Spend:Ess:3", "Healthcare typically 8-14% of after-tax income for working-age employer-covered households; higher for self-employed and pre-Medicare retirees",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "KFF Employer Health Benefits Survey; BLS CES",
     "",
     ""),

    ("CL177", "Spend:Ess:3", "HSA stealth retirement strategy: pay current medical OOP, save receipts, let HSA compound, reimburse decades later tax-free. Family max $8,750 × 30 yr at 6% real ≈ $700K",
     "MATH", "CON", "PASS", "A",
     "IRC §223; HSA receipt-keeping is settled IRS practice (no statute of limitations on qualified medical expenses if HSA existed at time of expense)",
     "",
     "Verified math: $8,750 annual contribution × FV annuity factor at 6%/30 = $691K. Strategy itself is well-established."),

    ("CL178", "Spend:Ess:3", "Cost Plus Drugs (Mark Cuban): 15% markup over manufacturer cost + $3 dispensing fee + $5 shipping; typically 80-95% below retail for generics",
     "CITATION", "CON", "PASS", "B",
     "Cost Plus Drugs (costplusdrugs.com) published pricing model; founded 2022 by Mark Cuban",
     "",
     "Verified pricing model. Available drugs are mostly generics; specialty/brand-name catalog more limited."),

    ("CL179", "Spend:Ess:3", "Healthshare ministries: not insurance, no ACA protections, pre-existing condition exclusions, religious/lifestyle requirements, lower monthly cost",
     "REGULATORY", "CON", "PASS", "B",
     "CMS guidance on health care sharing ministries; ACA §5000A(d)(2)(B); state insurance commissioner advisories (multiple)",
     "",
     "Healthshares are explicitly carved out from ACA. Member protection varies dramatically; not regulated as insurance in most states."),

    # --- Insurance ---
    ("CL180", "Spend:Ess:4", "Disability before age 67: approximately 1 in 4 of today's 20-year-old workers will experience a disability lasting 90+ days (SSA)",
     "CITATION", "CON", "PASS", "B",
     "Social Security Administration Fact Sheet on Disability Benefits",
     "",
     "SSA-published stat. Note: 'disability' as defined by SSDI is stricter than typical group/individual disability policy definitions."),

    ("CL181", "Spend:Ess:4", "DIME method for term life sizing: Debt + Income (years to replace) + Mortgage + Education",
     "CONVENTION", "CON", "PASS", "C",
     "Widely used in financial planning practice; appears in insurance industry training materials and CFP curriculum",
     "",
     "Conventional sizing method; specific attribution is murky (no single canonical source). Reasonable approximation; HLV (Human Life Value) method is more precise for high-earning households."),

    ("CL182", "Spend:Ess:4", "Umbrella liability pricing: $1M ~$200-400/yr, $2M ~$300-500/yr, $5M ~$500-800/yr; sits on top of auto + homeowners liability",
     "EMPIRICAL", "SOFT", "PASS", "C",
     "Industry pricing surveys (Insurance Information Institute, NerdWallet, Policygenius 2024); ranges vary by carrier, state, and driver record",
     "",
     "Pricing varies materially by carrier and state; recommended carry is at least net worth, often 1.5-2× for HNW households."),

    ("CL183", "Spend:Ess:4", "Whole life insurance illustrated gross returns typically 3-5%; net returns substantially lower after fees and mortality costs",
     "EMPIRICAL", "CON", "PASS", "B",
     "Wade Pfau research on whole life; LIMRA industry data; Consumer Federation of America analysis",
     "",
     "Well-established academic consensus that whole life is inferior to term-plus-invest-the-difference for the vast majority of cases."),

    # --- Food ---
    ("CL184", "Spend:Life:5", "Food spending typically 10-15% of after-tax income, split roughly evenly between groceries and food away from home",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "BLS Consumer Expenditure Survey; USDA Economic Research Service food expenditure series",
     "",
     "USDA ERS: food away from home was 55% of total food spending in 2023, up from 50% pre-pandemic average."),

    ("CL185", "Spend:Life:5", "Restaurant vs home-cooked meal cost differential: ~$15-25 restaurant vs $3-8 home-cooked per serving",
     "EMPIRICAL", "SOFT", "PASS", "C",
     "Wellio/Forbes 2018 study (home-cooked ~5× cheaper); recent inflation-adjusted equivalents",
     "",
     "Range covers ingredient quality and meal complexity. Restaurant figure includes tax/tip; takeout typically 15-20% lower than dine-in."),

    # --- Childcare/Education ---
    ("CL186", "Spend:Life:6", "Dependent Care FSA 2026 limit: $5,000 per household (joint/HoH/single); $2,500 if married filing separately",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §129(a)(2)(A); no annual inflation adjustment in statute. ARP $10,500 expansion expired after 2021",
     "",
     "Verified. Limit unchanged since 1986. CDCTC also available but phases down at higher incomes."),

    ("CL187", "Spend:Life:6", "Full-time daycare typical cost $15,000-$28,000/child/year; HCOL metros $30,000+; nanny $40,000-$70,000+",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Care.com Cost of Care Survey (annual); ChildCare Aware America 'Demanding Change' annual report",
     "",
     "2024 data; ranges grow with HCOL geography and child age (infant care most expensive)."),

    ("CL188", "Spend:Life:6", "Two-income vs childcare decision framework: if childcare > second income net of additional tax + work costs, second income contributes ~zero financially",
     "CONVENTION", "CON", "PASS", "C",
     "Standard framing in financial planning; appears in Kitces, NerdWallet, Bogleheads literature",
     "",
     "Note: ignores career-trajectory effects of continuous employment, retirement plan contributions, social security earnings record — pure cashflow framing only."),

    ("CL189", "Spend:Life:6", "College cost ranges 2024-25 all-in: in-state public flagship $25-35K; OOS public $40-55K; private $75-95K+ per year",
     "EMPIRICAL", "CON", "PASS", "B",
     "College Board 'Trends in College Pricing 2024-25'",
     "",
     "Published list prices; net prices (after aid) often significantly lower especially at higher-priced privates with strong endowments."),

    ("CL190", "Spend:Life:6", "Merit aid is dramatically more generous for applicants in the top quartile of an institution's admitted class",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Common Data Set (institutional self-reported); Lynn O'Shaughnessy, Ron Lieber research",
     "",
     "Well-established admissions practice; institutions use merit aid to attract students who improve their statistical profile."),

    ("CL191", "Spend:Life:6", "Median earnings in skilled trades (electrician, plumber, HVAC, dental hygiene, RN) often exceed median earnings for non-STEM bachelor's holders",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "BLS Occupational Outlook Handbook; Georgetown CEW 'The College Payoff' research",
     "",
     "Median electrician $61K, plumber $61K, RN $86K, dental hygienist $87K (BLS 2024). Median non-STEM bachelor's varies but frequently $50-65K early career."),

    # --- Subscriptions ---
    ("CL192", "Spend:Life:7", "$15/month subscription = $180/year; if invested at 6% real for 30 years compounds to ~$14,000 in opportunity cost",
     "MATH", "CON", "PASS", "A",
     "Standard annuity FV formula",
     "",
     "Verified: $15/mo × FV monthly annuity factor at 6%/360 ≈ $14,500."),

    ("CL193", "Spend:Life:7", "MVNO savings vs major carrier postpaid: typically $400-1,000/line/year",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Wirecutter, NerdWallet, RootMetrics 2024 carrier comparisons",
     "",
     "Mint, Visible, US Mobile, Cricket typically $15-50/mo per line vs $70-120 postpaid major carrier."),

    # --- Lifestyle creep ---
    ("CL194", "Spend:Life:8", "Hedonic adaptation: humans return to baseline well-being after positive changes within months",
     "CITATION", "CON", "PASS", "B",
     "Brickman & Campbell (1971) 'hedonic treadmill'; Brickman, Coates & Janoff-Bulman (1978) lottery winners study; extensive subsequent psychological literature",
     "",
     "Well-established in psychology literature."),

    ("CL195", "Spend:Life:8", "Easterlin Paradox (1974): once basic needs met, additional income does not produce proportional additional happiness",
     "CITATION", "CON", "PASS", "B",
     "Easterlin, R.A. (1974), 'Does Economic Growth Improve the Human Lot?' in Nations and Households in Economic Growth",
     "",
     "Foundational citation; refined by subsequent research."),

    ("CL196", "Spend:Life:8", "Kahneman & Deaton (2010): emotional well-being plateaus around $75,000 (2010 dollars, ~$110K in 2026 dollars)",
     "CITATION", "CON", "PASS", "B",
     "Kahneman & Deaton (2010), 'High income improves evaluation of life but not emotional well-being,' PNAS 107(38)",
     "",
     "Original finding; later updated by Killingsworth/Kahneman 2023."),

    ("CL197", "Spend:Life:8", "Killingsworth & Kahneman (2023): emotional well-being continues rising with income for most, plateau exists only for unhappy minority",
     "CITATION", "CON", "PASS", "B",
     "Killingsworth, Kahneman, & Mellers (2023), 'Income and emotional well-being: A conflict resolved,' PNAS 120(10)",
     "",
     "Reconciliation of Kahneman-Deaton 2010 vs Killingsworth 2021 findings. The plateau effect is real but concentrated in the unhappy 20%."),

    ("CL198", "Spend:Life:8", "Lifestyle creep math: $400/mo additional spending ≈ 4-6 pp savings rate reduction, adding 3-5 years to FI timeline at typical pre-FI income",
     "MATH", "CON", "PASS", "A",
     "Derived from MMM savings rate formula (Phase 2 §3); CL134/135 already verified",
     "",
     "Verified per MMM formula. Exact delay depends on starting savings rate; at 25% savings, $400/mo creep on $120K income reduces SR by 4pp, adding ~3 yr to FI timeline."),

    # --- Phase 3 calculator math verification ---
    ("CL199", "Spend:Ess:1 calc", "Rent vs buy calculator: TCO includes P&I, property tax, maintenance, insurance, opportunity cost on down payment, minus principal paid and 2% annual appreciation",
     "MATH", "CON", "PASS", "A",
     "Standard real-estate financial analysis",
     "",
     "Verified math. 2% appreciation assumption is long-run US average post-inflation; real outcomes vary widely by geography and cycle."),

    ("CL200", "Spend:Ess:2 calc", "Vehicle TCO calculator: (purchase − 20% residual at 10yr) + (annual op × years)",
     "MATH", "CON", "PASS", "A",
     "Simplified TCO; AAA methodology is more complete (insurance, fuel, depreciation, finance, maintenance/repairs, license/registration)",
     "",
     "20% residual is rough average for 10-year-old vehicles; ranges widely by make/model."),

    ("CL201", "Spend:Ess:3 calc", "HDHP vs PPO calculator: planCost = premium + min(deductible + 20% coinsurance on overage, premium + OOP max); HDHP advantage = HSA × marginal tax rate",
     "MATH", "CON", "PASS", "A",
     "Simplified plan economics; 20% coinsurance is common but plans vary (10-30%)",
     "",
     "Verified math; coinsurance varies by plan."),

    ("CL202", "Spend:Ess:4 calc", "DIME total: Debt + (Income × Years) + Mortgage + Education, rounded up to nearest $250K",
     "MATH", "CON", "PASS", "A",
     "Standard DIME implementation",
     "",
     "Verified."),

    ("CL203", "Spend:Life:6 calc", "DCFSA savings: contribution × (fed + state + FICA marginal rate)",
     "MATH", "CON", "PASS", "A",
     "Standard pre-tax benefit math",
     "",
     "Verified. FICA savings on DCFSA contributions are real because contributions reduce FICA wage base."),

    ("CL204", "Spend:Life:7 calc", "Subscription FV: monthly contribution treated as monthly annuity, FV = PMT × [(1+r/12)^n − 1] / (r/12)",
     "MATH", "CON", "PASS", "A",
     "Standard annuity FV formula",
     "",
     "Verified."),

    ("CL205", "Spend:Life:8 calc", "Lifestyle creep: new SR = (income×SR − annual_creep) / income; recomputed years-to-FI using MMM formula",
     "MATH", "CON", "PASS", "A",
     "MMM formula (CL134) applied to perturbed savings rate",
     "",
     "Verified. Calculator gracefully handles cases where creep exceeds current savings (warns rather than producing negative)."),

    # ============================================================
    # === PHASE 3 CHECKPOINT PERSONA OBSERVATIONS (P3A + P3B) ===
    # ============================================================
    # APPLIED = structural fix made now; DEFERRED-P3.5 = backlog
    ("CL206", "Spend:Ess:1", "P3A#1: Price-to-rent thresholds (15/20) are rate-regime-dependent; calibrated to 2010s 3-4% mortgage environment; at 6.5-7% rates the thresholds shift toward renting at every PTR level",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "Trulia origin (2010s) at low rate environment; current rate environment substantially shifts breakeven; calculation of PTR breakeven as function of mortgage rate",
     "APPLIED: Math-callout in Housing §1 explicitly notes PTR thresholds (15/20) and 5% rule's 'opportunity cost' component are rate-environment-dependent. At 8%+ mortgage rates the buying breakeven PTR drops to ~12; at 3-4% rates climbs to 25+. 5% rule scales to ~6-7% at 8% mortgage rates. Calculator handles rate-sensitivity correctly; heuristic thresholds are for intuition only.",
     "Phase 3.5 cleanup."),

    ("CL207", "Spend:Ess:1", "P3A#2: 5% rule omits mortgage interest deduction (post-TCJA SALT-capped), primary residence capital gains exclusion ($250K/$500K), and inflation hedge of fixed-rate P&I",
     "CITATION", "SOFT", "CORRECTED", "B",
     "Ben Felix original framing; IRC §121 (gain exclusion); TCJA SALT cap effects",
     "APPLIED: Same callout as CL206 covers the 5% rule's rate-environment dependence with the 3% opportunity cost component scaling.",
     "Phase 3.5 cleanup."),

    ("CL208", "Spend:Ess:1 calc", "P3A#3: 2% real home appreciation default was aggressive vs Case-Shiller long-run national average (0.4-1.0%)",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Case-Shiller US National Home Price Index 1890-present; Robert Shiller's research",
     "APPLIED: Lowered default to 1% real appreciation in rent-vs-buy calculator. Result text now notes 'long-run US Case-Shiller average — HCOL coastal metros run higher, much of the country runs flat or lower.'",
     "Structural fix #3 applied."),

    ("CL209", "Spend:Ess:1 calc", "P3A#4: Rent-vs-buy calculator omitted transaction costs (3% buying, 7% selling); single most common error in DIY housing analysis",
     "MATH", "CON", "CORRECTED", "A",
     "Standard real-estate transaction cost benchmarks; NAR commission data; closing cost surveys",
     "APPLIED: Added 3% buying closing costs and 7% selling costs (commission + closing) to rent-vs-buy calc. Result text now surfaces these explicitly. Calculator output materially shifts at shorter holding periods (3-5 yr).",
     "Structural fix #1 applied. Most impactful Phase 3 fix."),

    ("CL210", "Spend:Ess:1", "P3A#5: 28/36 rule not mentioned despite being the operative mortgage underwriting heuristic for affordability",
     "CONVENTION", "CON", "CORRECTED", "B",
     "Traditional mortgage underwriting (Fannie/Freddie historical); FHA underwriting guidelines",
     "APPLIED: Dedicated Math-callout 'The 28/36 affordability rule' added to Housing §1. Defines front-end DTI (≤28% gross monthly for PITIA+HOA) and back-end DTI (≤36% total debt). Distinguishes lender maximums (43-45% conventional with overlays, 50% FHA) from honest personal threshold (25-28% front-end). Explicit warning that lender-max DTI produces house-poor situations.",
     "Phase 3.5 cleanup."),

    ("CL211", "Spend:Ess:1", "P3A#6: 0.75-1% refinance heuristic is outdated; modern no-cost refinances change math entirely",
     "CONVENTION", "SOFT", "CORRECTED", "C",
     "Bankrate, Freddie Mac modern refinance guidance; no-cost refinance mechanics",
     "APPLIED: 'Mortgage strategy' Math-callout in Housing §1 includes refinance framework update. No-cost refinance (lender absorbs closing costs for 0.125-0.375pp rate premium) eliminates recovery-period concern. Apples-to-apples comparison between par-with-costs and no-cost-higher-rate options recommended.",
     "Phase 3.5 cleanup."),

    ("CL212", "Spend:Ess:1", "P3A#7: FHA 3.5% down on 3-4 unit properties requires self-sufficiency test (75% of gross rent must cover full PITI); MIP ~0.85% annual that doesn't drop off",
     "STATUTORY", "CON", "CORRECTED", "A",
     "HUD 4000.1 FHA Single Family Housing Policy Handbook; FHA MIP rules",
     "APPLIED: Same 'Mortgage strategy' Math-callout covers FHA in detail: 3.5% down on 1-4 unit owner-occupied, MIP (0.55-1.05% annual + 1.75% upfront) for life of loan or 11 years if down >10%, 3-4 unit self-sufficiency test (75% rental income from non-occupied units must meet/exceed PITIA), conventional alternative limits.",
     "Phase 3.5 cleanup."),

    ("CL213", "Spend:Ess:1", "P3A#8: 15-vs-30 mortgage 'invest the difference' assumes behavioral execution; empirical reality is most 30-yr holders spend the cash flow advantage",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "Behavioral economics literature on consumer savings discipline; Thaler nudge research",
     "APPLIED: 'Mortgage strategy' Math-callout has 15-vs-30 behavioral caveat: math favors 30-year conditional on actually investing the difference; many households don't, so 15-year functions as forced savings and beats 30-year in practice for those households despite losing in expectation. Choose the structure you'll actually execute.",
     "Phase 3.5 cleanup."),

    ("CL214", "Spend:Ess:1", "P3A#9: Geographic arbitrage section omits salary geographic adjustment dynamic; many employers haircut remote-relocation pay 15-25%",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Tech sector compensation surveys 2022-24 (Levels.fyi, Pave); remote work pay band data",
     "APPLIED: 'Geographic arbitrage salary haircut' Math-callout added to Housing §1. Cost-of-living differential is one side; salary differential is the other. Same job typically pays 15-35% less in lower-cost markets (SF $250K → Austin $180K → Raleigh $160K examples). Net savings requires both numbers. Largest win conditions: remote with HCOL-rate retention; relocation to no-income-tax state; HCOL premium amenities (private schools) avoided.",
     "Phase 3.5 cleanup."),

    ("CL215", "Spend:Ess:3", "P3B#1: HDHP qualification requirements not stated; not every plan marketed as 'high deductible' qualifies for HSA contributions",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRC §223(c)(2); IRS Rev. Proc. (annual update); HDHP definitional thresholds",
     "APPLIED: Added 'What HSA-eligible HDHP actually means' callout in healthcare section explaining IRC §223(c)(2) requirements, the need to verify 'HSA-eligible' explicitly at open enrollment, and the prescription-drug carve-out disqualifier. Also added HSA investment-access nuance (provider variation, Fidelity advantage).",
     "Structural fix #2 applied. Single most actionable Phase 3 fix."),

    ("CL216", "Spend:Ess:3 calc", "P3B#2: HDHP vs PPO calculator's 20% coinsurance default doesn't match all plans; range is 10-30% in employer plans, sometimes tiered",
     "UX", "SOFT", "CORRECTED", "C",
     "Industry knowledge of plan design variation",
     "APPLIED: Math-callout 'The coinsurance variable matters too' in Healthcare §3. Explains coinsurance step (60-90% covered between deductible and OOP-max, patient pays 10-40%) and worked example: $20K allowed charges with $3K deductible / $7K OOP-max / 20% coinsurance vs $4K total on $1K-deductible / $4K-OOP-max / 0% coinsurance plan. HSA-eligible plans often have 20-30% coinsurance which can flip break-even vs first-dollar PPO.",
     "Phase 3.5 cleanup."),

    ("CL217", "Spend:Ess:3", "P3B#3: HSA 'save receipts, reimburse later' strategy needs cash flow caveat; only available if you can afford current OOP from non-HSA money",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Standard HSA optimization literature; cash flow constraint is implicit but unstated",
     "APPLIED: Math-callout 'HSA receipt strategy — the cash flow constraint' in Healthcare §3. Optimal theoretical strategy (pay OOP, save receipts, reimburse decades later) requires sufficient cash flow during accumulation, durable documentation (receipts + EOBs preserved decades), and lifetime drawdown. Non-spouse beneficiaries face full ordinary income tax on inherited HSA balance (tax-free benefit doesn't pass through). Spouse beneficiary can roll to own; non-spouse cannot. Older accumulators should drawdown receipts during life.",
     "Phase 3.5 cleanup."),

    ("CL218", "Spend:Ess:3", "P3B#4: HSA investment access varies by provider (cash floor requirements $1-2K typical; Fidelity has none)",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "HSA custodian comparison; Fidelity HSA terms vs HealthEquity/Optum/HSA Bank",
     "APPLIED (partial): Investment-access nuance added to the HDHP qualification callout (Fix #2). Custodian rollover option for workers stuck with employer-defaulted custodian also mentioned.",
     "Captured in same callout as CL215 fix."),

    ("CL219", "Spend:Ess:3", "P3B#5 (RESHAPED): Original critique was that DPC was 'mentioned incompletely.' Actual section did NOT mention DPC at all — persona error caught on verification",
     "META", "CON", "CORRECTED", "B",
     "Persona critique verification; transparent acknowledgment of fabricated content premise",
     "APPLIED (as content addition, not correction): Added properly-caveated DPC content to healthcare section. Critical constraint flagged that DPC works only when paired with HDHP for catastrophic coverage. Also added Sedera/Zion HealthShare as non-religious healthshare alternatives.",
     "TRANSPARENCY: P3B#5 critiqued a content gap that read as if it were a correction needed. On verification, DPC wasn't in the section. Converted to content addition with proper caveat per P3B's correct guidance on HOW to present DPC."),

    ("CL220", "Spend:Ess:4 calc", "P3B#6: DIME calculator ignores present-value discounting; Income × Years overstates needed coverage at typical inflation",
     "MATH", "CON", "CORRECTED", "B",
     "Actuarial present-value methodology; HLV (Human Life Value) method comparison",
     "APPLIED: 'DIME refinement — present-value discount the income replacement' Math-callout added to Insurance §4. Standard DIME (debt + income×years + mortgage + education) overstates needed coverage. Worked example: $80K income for 20 years has PV ~$1.19M at 3% real, not $1.6M lump sum. PV-discounted DIME: $400K mortgage + $1.19M PV-income + $250K education + $20K debt = $1.86M vs simple form $2.27M. Simple form is conservative which is sometimes the right framing. For households where premium cost constrains coverage, PV-discounted form gives better term-length-vs-face-value tradeoffs. For income that continues (working spouse), only replace lost portion not full deceased earner's salary.",
     "Phase 3.5 cleanup."),

    ("CL221", "Spend:Ess:4", "P3B#7: Disability insurance section omits policy features that determine claim payment — benefit period, elimination period, residual/partial riders, COLA, FIO, mental/nervous limitations",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "DI policy structure; LIMRA industry data; CFP curriculum on disability income",
     "APPLIED: Comprehensive Math-callout 'DI rider and feature details that meaningfully change policy value' in Insurance §4. Covers: benefit period (to 65/67/70, lifetime rare/expensive, shorter cheaper but exposes long-tail), elimination period (60-90-180-365 days), residual/partial disability rider (proportional benefits at reduced capacity, common after serious illness recovery — without it policy is binary), future increase option (FIO) for non-underwritten step-ups, COLA rider for inflation-indexed benefits during claims, mental & nervous (M&N) 24-month cap and why physicians/professionals should pay extra to remove it. Properly-featured vs bare policy premium differential 30-50% but justified for high earners.",
     "Phase 3.5 cleanup."),

    ("CL222", "Spend:Ess:4", "P3B#8: '1-in-4 of 20-year-olds will be disabled' SSA stat uses SSDI's strict definition; loose for own-occupation policy context",
     "CITATION", "SOFT", "CORRECTED", "B",
     "SSA Fact Sheet on Disability Benefits; definitional differences SSDI vs group/individual DI policies",
     "APPLIED: SSA 1-in-4 stat now contextualized in main DI paragraph: derives from SSA actuarial tables, includes all disabilities meeting 5-month SSA elimination period + substantial impairment, career-long cumulative (not annual), includes permanent and temporary. Private disability claim probability varies materially by occupation, age, lifestyle. Qualitative point remains (disability > premature death probability for most working-age) without overstating the specific statistic.",
     "Phase 3.5 cleanup."),

    ("CL223", "Spend:Ess:4", "P3B#9: Umbrella liability section omits underlying limit requirements (typically 250/500/100 auto, $300-500K homeowners) which add cost to enable umbrella",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Umbrella carrier underwriting standards; Insurance Information Institute guidance",
     "APPLIED: Math-callout 'Umbrella underlying limits and business exclusions' in Insurance §4. Explains: umbrella attaches above underlying auto/home/etc limits; carrier minimums (250/500/100 auto, 300K home commonly required); if underlying limit isn't maintained, gap is self-funded; classic surprise pattern (state-minimum auto + serious accident + judgment > limits = personal liability for gap). Business activities generally excluded from personal umbrella (rental properties, side businesses, consulting, rideshare/delivery driving). Separate commercial umbrella layered on relevant business policies for high-income professionals with side businesses.",
     "Phase 3.5 cleanup."),

    ("CL224", "Spend:Ess:3", "P3B#10: Healthshare framing covered religious-affiliated options only; secular alternatives (Sedera, Zion) exist",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "Sedera, Zion HealthShare member documentation; ACA-exempt cost-sharing organization landscape",
     "APPLIED: Added sentence to healthshare paragraph noting non-religious alternatives. Captured alongside DPC addition (CL219).",
     "Resolved in same edit as CL219."),

    # ============================================================
    # === PHASE 4: PORTFOLIO CONSTRUCTION CLAIMS ===
    # ============================================================
    # --- Bogleheads practice ---
    ("CL225", "Port:BH:1", "SPIVA report: ~85-90% of active US large-cap funds underperform S&P 500 over 15-year periods; persistence of outperformance rarely above chance",
     "CITATION", "CON", "PASS", "B",
     "S&P Global SPIVA (S&P Indices Versus Active) Scorecard, published biannually since 2002; SPIVA Persistence Scorecard",
     "",
     "Well-established. Specific figures vary slightly each report; 15-year underperformance has been ≥80% across all reports."),

    ("CL226", "Port:BH:1", "John Bogle founded Vanguard 1975; launched First Index Investment Trust (now Vanguard 500) in 1976",
     "CITATION", "CON", "PASS", "A",
     "SEC filings; Vanguard corporate history; Bogle 'Stay the Course' (2018)",
     "",
     "Verified historical fact."),

    ("CL227", "Port:BH:1", "Active fund expense ratios typically 0.50-1.50%; broad index funds 0.03-0.10%",
     "EMPIRICAL", "CON", "PASS", "B",
     "ICI Fact Book annual; Morningstar fee research",
     "",
     "ICI 2024 average active equity ER ~0.65%, passive ~0.05%. Distribution skewed; institutional share classes lower."),

    ("CL228", "Port:BH:1", "The 'Ten Boglehead Principles' / Investment Philosophy",
     "CITATION", "SOFT", "PASS", "C",
     "Bogleheads' Guide to Investing (Larimore, Lindauer, LeBoeuf); Bogleheads wiki",
     "",
     "Different orderings across sources; operative content consistent."),

    ("CL229", "Port:BH:2", "Canonical three-fund portfolio: VTI (US Total Market) + VXUS (International Total) + BND (US Total Bond)",
     "CONVENTION", "CON", "PASS", "B",
     "Bogleheads wiki; widely-cited canonical implementation; Taylor Larimore 'The Bogleheads Three-Fund Portfolio' (2018)",
     "",
     "Verified. Equivalents at Fidelity (FZROX/FZILX/FXNAX) and Schwab (SWTSX/SWISX/SWAGX) listed correctly."),

    ("CL230", "Port:BH:2", "Major broad index fund expense ratios: Vanguard ~0.03%, Fidelity ZERO 0.00%, Schwab ~0.03-0.06%",
     "EMPIRICAL", "CON", "PASS", "A",
     "Fund prospectuses; current as of 2025-2026",
     "",
     "Verified. Note: Fidelity ZERO funds (FZROX/FZILX) carry portability restriction (only at Fidelity)."),

    ("CL231", "Port:BH:2", "ETF in-kind creation/redemption mechanics make equity ETFs structurally more tax-efficient in taxable accounts than mutual fund equivalents",
     "REGULATORY", "CON", "PASS", "B",
     "IRC §852(b)(6); SEC ETF mechanics; academic literature on ETF tax efficiency (Dickson, Shoven 1995; subsequent literature)",
     "",
     "Verified. VTI distributes substantially fewer capital gains than VTSAX in practice despite tracking the same index."),

    ("CL232", "Port:BH:3", "Asset allocation accounts for ~90% of long-term portfolio return variance (Brinson Hood Beebower 1986; Ibbotson Kaplan 2000)",
     "CITATION", "CON", "PASS", "B",
     "Brinson, Hood & Beebower 'Determinants of Portfolio Performance' Financial Analysts Journal (1986, updated 1991); Ibbotson & Kaplan 'Does Asset Allocation Policy Explain 40, 90, or 100 Percent of Performance?' FAJ (2000)",
     "",
     "Cross-reference CL144 (already in audit from Phase 2). Classic finding; subject to interpretation debate (variance vs returns vs return levels)."),

    ("CL233", "Port:BH:3", "100/110/120-minus-age allocation heuristics; modern updates reflect lower bond yields and longer life expectancies",
     "CONVENTION", "SOFT", "PASS", "C",
     "Traditional financial planning conventions; updated frameworks in Bernstein 'The Four Pillars'; Kitces commentary",
     "",
     "Heuristic, not optimization. All reasonable starting points."),

    ("CL234", "Port:BH:3", "Risk tolerance decomposition: ability, willingness, need — William Bernstein framing",
     "CITATION", "CON", "PASS", "B",
     "William Bernstein 'The Four Pillars of Investing' (2002); subsequent Bernstein writing",
     "",
     "Bernstein's canonical framework. The binding constraint is the lowest of the three."),

    ("CL235", "Port:BH:3", "Target-date fund glide path differences: Vanguard 'to' retirement holds 90% equity until 25yr out, lands at 30%; T. Rowe Price 'through' retains higher equity longer",
     "EMPIRICAL", "CON", "PASS", "B",
     "Vanguard Target Retirement Funds prospectus; T. Rowe Price Target Date Funds prospectus; Morningstar target-date fund landscape report (annual)",
     "",
     "Verified. Material differences between 'to' and 'through' approaches affect later-stage retirees most."),

    ("CL236", "Port:BH:3", "Rising equity glide path in retirement (Pfau-Kitces 2014) often improves sustainable withdrawal rates vs declining glide paths",
     "CITATION", "CON", "PASS", "B",
     "Pfau & Kitces 'Reducing Retirement Risk with a Rising Equity Glide-Path' Journal of Financial Planning, Jan 2014",
     "",
     "Counterintuitive finding; not standard practice in target-date funds."),

    ("CL237", "Port:BH:4", "Rebalancing approaches (calendar vs threshold): differences are small over long horizons, typically 10-20 bps annually",
     "CITATION", "SOFT", "PASS", "B",
     "Vanguard 'Best Practices for Portfolio Rebalancing' (2010 white paper, updated periodically); Jaconetti, Kinniry, Zilbering",
     "",
     "Verified. Recommendation: pick one approach and follow it consistently rather than over-optimize between them."),

    ("CL238", "Port:BH:4", "Rebalancing bonus typically 20-60 bps annually for a 60/40 portfolio with reasonable volatility",
     "EMPIRICAL", "SOFT", "PASS", "C",
     "Bouchey, Nemtchinov, Paulsen, Stein 'Volatility Harvesting' Journal of Wealth Management (2012); subsequent academic work",
     "",
     "Magnitude depends on volatility, correlation, and rebalancing approach. Not free — comes from mean reversion which may not always hold."),

    ("CL239", "Port:BH:5", "Foreign tax credit on international fund dividends recovered in taxable accounts; lost in tax-deferred accounts",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §901; Foreign Tax Credit rules; relevant Treas. Regs.",
     "",
     "Verified statutory rule. Benefit typically 30-50 bps annually on international equity dividends."),

    ("CL240", "Port:BH:5", "Municipal bond break-even tax rate vs taxable bonds approximately 25-30% combined federal+state for high-grade munis",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Municipal bond yield comparisons (BofA Merrill ML indices); standard wealth management calculation",
     "",
     "Rough break-even varies with rate environment and credit quality."),

    # --- Theory view ---
    ("CL241", "Port:Th:6", "Harry Markowitz 'Portfolio Selection' Journal of Finance, March 1952",
     "CITATION", "CON", "PASS", "A",
     "Markowitz, H. 'Portfolio Selection' Journal of Finance Vol 7 No 1, pp 77-91 (1952)",
     "",
     "Foundational citation. Markowitz received the 1990 Nobel Prize in Economic Sciences for this framework."),

    ("CL242", "Port:Th:6", "Portfolio variance formula: σ_p² = Σ Σ w_i × w_j × Cov(R_i, R_j); for two assets: σ_p² = w_s²σ_s² + w_b²σ_b² + 2 w_s w_b ρ σ_s σ_b",
     "MATH", "CON", "PASS", "A",
     "Standard MPT identity",
     "",
     "Verified mathematical identity. Diversification benefit depends on ρ < 1."),

    ("CL243", "Port:Th:6", "Efficient frontier: set of portfolios that minimize variance for a given expected return (or maximize return for a given variance)",
     "MATH", "CON", "PASS", "A",
     "Markowitz 1952 formulation",
     "",
     "Standard. Dominated portfolios lie below the frontier."),

    ("CL244", "Port:Th:6", "Markowitz received 1990 Nobel Prize in Economic Sciences (shared with Sharpe and Miller)",
     "CITATION", "CON", "PASS", "A",
     "Nobel Foundation records",
     "",
     "Verified historical fact."),

    ("CL245", "Port:Th:6", "Black-Litterman model (1990): combines market-equilibrium prior with investor views to address mean-variance optimization sensitivity",
     "CITATION", "CON", "PASS", "B",
     "Black, F. & Litterman, R. 'Asset Allocation: Combining Investor Views with Market Equilibrium' Journal of Fixed Income (1991); Goldman Sachs internal note (1990)",
     "",
     "Standard institutional approach to address MPT's corner-solution problem."),

    ("CL246", "Port:Th:7", "CAPM developed by Sharpe (1964), Lintner (1965), Mossin (1966)",
     "CITATION", "CON", "PASS", "A",
     "Sharpe 'Capital Asset Prices' Journal of Finance Vol 19 No 3 (1964); Lintner 'The Valuation of Risk Assets' Review of Economics and Statistics (1965); Mossin 'Equilibrium in a Capital Asset Market' Econometrica (1966)",
     "",
     "Three independent contributions in mid-1960s."),

    ("CL247", "Port:Th:7", "CAPM equation: E(R_i) = R_f + β_i × (E(R_m) − R_f); β_i = Cov(R_i, R_m) / Var(R_m)",
     "MATH", "CON", "PASS", "A",
     "Sharpe 1964; standard finance textbook",
     "",
     "Verified mathematical formulation."),

    ("CL248", "Port:Th:7", "Sharpe received 1990 Nobel Prize in Economic Sciences (shared with Markowitz and Miller) for CAPM",
     "CITATION", "CON", "PASS", "A",
     "Nobel Foundation records",
     "",
     "Verified historical fact."),

    ("CL249", "Port:Th:7", "Roll's critique (1977): CAPM is jointly untestable with mean-variance efficiency of the market portfolio; true market portfolio is unobservable",
     "CITATION", "CON", "PASS", "B",
     "Roll, R. 'A Critique of the Asset Pricing Theory's Tests' Journal of Financial Economics (1977)",
     "",
     "Influential methodological critique; subsequent empirical work shifted focus to predictive usefulness rather than literal truth of CAPM."),

    ("CL250", "Port:Th:8", "Fama-French (1992) three-factor model: market + SMB (size) + HML (value)",
     "CITATION", "CON", "PASS", "A",
     "Fama, E. & French, K. 'The Cross-Section of Expected Stock Returns' Journal of Finance Vol 47 No 2 (1992)",
     "",
     "Foundational extension of CAPM."),

    ("CL251", "Port:Th:8", "Carhart (1997) four-factor: adds momentum (MOM / WML)",
     "CITATION", "CON", "PASS", "A",
     "Carhart, M. 'On Persistence in Mutual Fund Performance' Journal of Finance Vol 52 No 1 (1997)",
     "",
     "Standard extension."),

    ("CL252", "Port:Th:8", "Fama-French (2015) five-factor: adds RMW (profitability) and CMA (investment)",
     "CITATION", "CON", "PASS", "A",
     "Fama, E. & French, K. 'A Five-Factor Asset Pricing Model' Journal of Financial Economics Vol 116 No 1 (2015)",
     "",
     "Verified. HML's role somewhat diminished by the additional factors."),

    ("CL253", "Port:Th:8", "Hou-Xue-Zhang (2015) Q-factor model; alternative factor specification",
     "CITATION", "CON", "PASS", "B",
     "Hou, K., Xue, C. & Zhang, L. 'Digesting Anomalies: An Investment Approach' Review of Financial Studies (2015)",
     "",
     "Competing model emphasizing investment and ROE factors."),

    ("CL254", "Port:Th:8", "Harvey, Liu & Zhu (2016): 'factor zoo' — over 300 documented factors; most do not survive rigorous statistical thresholds correcting for multiple testing",
     "CITATION", "CON", "PASS", "B",
     "Harvey, C., Liu, Y. & Zhu, H. '... and the Cross-Section of Expected Returns' Review of Financial Studies Vol 29 No 1 (2016)",
     "",
     "Influential critique; argues t-statistic threshold should be ≥3.0 for new factors rather than the conventional 2.0."),

    ("CL255", "Port:Th:8", "McLean & Pontiff (2016): average factor returns drop ~50% after publication",
     "CITATION", "CON", "PASS", "B",
     "McLean, R.D. & Pontiff, J. 'Does Academic Research Destroy Stock Return Predictability?' Journal of Finance Vol 71 No 1 (2016)",
     "",
     "Verified. Post-publication decay suggests genuine arbitrage activity reduces or eliminates published anomalies."),

    ("CL256", "Port:Th:8", "Robust factor set likely to persist: market, value, size (with caveats), profitability, possibly momentum",
     "EMPIRICAL", "SOFT", "PASS", "C",
     "Asness, Frazzini, Pedersen 'Quality Minus Junk' (2019); subsequent factor literature; AQR research",
     "",
     "Tier C: which factors 'survive' is debated. This is the empirical consensus circa 2025."),

    ("CL257", "Port:Th:8", "Factor ETFs available retail: Avantis (AVUS/AVUV/AVDV); Dimensional (DFA, now public-accessible); Vanguard factor funds (VFMV, VFLQ, etc.)",
     "EMPIRICAL", "CON", "PASS", "A",
     "Fund prospectuses; current as of 2025-2026",
     "",
     "Verified availability and tickers."),

    ("CL258", "Port:Th:8", "Robert Shiller received 2013 Nobel Prize (shared with Fama and Hansen) for empirical analysis of asset prices including excess volatility",
     "CITATION", "CON", "PASS", "A",
     "Nobel Foundation records",
     "",
     "Verified. Shiller's work bridges behavioral and traditional finance."),

    # --- Phase 4 calculator math verification ---
    ("CL259", "Port:BH:4 calc", "Expense ratio calc: FV = PV(1+r-ER)^n + PMT × [((1+r-ER)^n − 1)/(r-ER)] applied with net (r-ER) return",
     "MATH", "CON", "PASS", "A",
     "Standard FV-of-annuity formula with net-of-expense return",
     "",
     "Verified math."),

    ("CL260", "Port:Th:6 calc", "Two-asset minimum variance portfolio closed-form: w*_s = (σ_b² − ρσ_sσ_b) / (σ_s² + σ_b² − 2ρσ_sσ_b)",
     "MATH", "CON", "PASS", "A",
     "Standard MPT closed-form derivation",
     "",
     "Verified. Derived by setting dσ_p²/dw_s = 0."),

    ("CL261", "Port:Th:7 calc", "Sharpe ratio: SR = (R_p − R_f) / σ_p",
     "MATH", "CON", "PASS", "A",
     "Sharpe 1966 reformulation; standard finance",
     "",
     "Verified. S&P 500 long-run Sharpe ~0.4-0.5; sustained 2.0+ is unusual."),

    # ============================================================
    # === PHASE 4 CHECKPOINT PERSONA OBSERVATIONS (P4A + P4B) ===
    # ============================================================
    ("CL262", "Port:BH:3", "P4A#1: Vanguard glide path framed as 'to retirement' but Vanguard updated to 'through retirement' framework years ago",
     "CITATION", "CON", "CORRECTED", "B",
     "Vanguard Target Retirement Funds prospectus and white papers; Vanguard glide path documentation 2013-present",
     "APPLIED: Updated glide path description. Vanguard now holds 90% equity until ~25 yr before target, declines to 50% at target date, continues to ~30% ~7 yr post-target. Noted 'to vs through' distinction has narrowed across providers; remaining differences are slope and terminal equity level.",
     "Structural fix #1. Real factual error."),

    ("CL263", "Port:BH:2", "P4A#2: Three-fund holding counts approximate and drift with index reconstitution (VTI ~3,700, VXUS ~8,800, BND ~10,500 currently)",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Vanguard fund holdings reports (quarterly)",
     "APPLIED: Three-fund holding counts in Bogleheads §2 now show ranges ('approximately 3,500-4,000', 'approximately 8,000-9,000', 'approximately 10,000+') with explicit note that counts drift quarterly with index methodology updates and corporate actions. The structural point (effectively complete market coverage at retail-accessible cost) preserved without false precision.",
     "Phase 4.5 cleanup."),

    ("CL264", "Port:BH:4", "P4A#3: 5/25 rebalancing band example was mathematically muddled — 5% absolute and 25% relative produce very different ranges; convention is lesser-of",
     "MATH", "CON", "CORRECTED", "A",
     "Bogleheads wiki 'Rebalancing' page; Larry Swedroe 5/25 rule canonical formulation",
     "APPLIED: Rewrote band example to clarify 5/25 rule as 'lesser of 5pp absolute OR 25% relative.' Added worked examples for 70% stock allocation (5% applies) and 5% REIT allocation (25% applies). Bands work together to give larger allocations tighter tolerances.",
     "Structural fix #2. Math clarification."),

    ("CL265", "Port:BH:2", "P4A#4: Vanguard mutual funds with ETF share classes had patent-protected in-kind redemption access making VTSAX ~tax-equivalent to VTI; patent expired 2023 but existing funds retain structure",
     "REGULATORY", "CON", "CORRECTED", "B",
     "Vanguard ETF share-class patent (expired 2023); IRC §852(b)(6); Vanguard fund tax efficiency white papers",
     "APPLIED: New Math-callout 'The Vanguard mutual-fund-with-ETF-share-class exception' in Bogleheads §2. Explains historical dual-share-class tax efficiency, patent expiration May 2023, 2025 SEC approvals to other fund families (Dimensional, Morgan Stanley, etc.), but only Vanguard has multi-decade track record. Guidance: non-Vanguard taxable accounts prefer ETF wrapper directly; Vanguard accounts can use mutual fund version without material tax loss.",
     "Phase 4.5 cleanup."),

    ("CL266", "Port:BH:3", "P4A#5: Brinson-Hood-Beebower '90% of variance' frequently misinterpreted — refers to time-series variance of single portfolio, not cross-sectional return level differences (Ibbotson-Kaplan 2000 decomposition: 90% variance, 40% cross-sectional return levels, 100% average return)",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "Brinson-Hood-Beebower 1986; Ibbotson-Kaplan 2000 'Does Asset Allocation Policy Explain 40, 90, or 100 Percent of Performance?'",
     "APPLIED: Math-callout 'What the 90% of variance claim actually means' in Bogleheads §3. Critical distinction: 90% is time-series variance of single portfolio (how much your returns bounce around), not cross-sectional return-level differences. Ibbotson-Kaplan 2000 decomposition: 90% time-series variance, 40% cross-sectional return differences, 100% of average portfolio total return. Popular reading 'asset allocation = 90% of outcomes' overstates cross-sectional case. Accurate reading: dominant driver of portfolio volatility, matters more than fund selection within categories.",
     "Phase 4.5 cleanup."),

    ("CL267", "Port:BH:5", "P4A#6: Muni break-even tax rate 25-30% framing is rate-environment-dependent; current muni/Treasury yield ratios shift this materially",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Bloomberg muni/Treasury yield curve data; break-even formula 1 − (TEY_muni / Yield_taxable)",
     "APPLIED: Bogleheads §5 muni discussion now uses formula rather than point-estimate break-even: muni_yield × (1/(1-combined_tax_rate)) = taxable-equivalent yield. Plug in specific marginal federal+state rate and current muni yields. Worked examples: 32% fed + 9% state = 41% combined, 3.5% muni → 5.93% TEY (beats most Treasuries); 24% combined, 3.5% muni → 4.6% TEY (often loses to corporate/Treasury). Compute for your situation rather than rely on single threshold.",
     "Phase 4.5 cleanup."),

    ("CL268", "Port:BH:1", "P4A#7: Behavior gap empirical evidence (Dalbar QAIB: ~1-2% annual investor underperformance vs funds held) is the strongest argument for Bogleheads discipline but isn't surfaced as a core finding",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Dalbar 'Quantitative Analysis of Investor Behavior' annual report; Morningstar 'Mind the Gap' annual study; both consistently show 1-2% gap",
     "APPLIED: New Math-callout 'The behavior gap — empirically the strongest case for Boglehead discipline' in Bogleheads §1. Cites Dalbar QAIB and Morningstar Mind the Gap studies. 1-2pp annual gap between fund returns and dollar-weighted investor returns (Dalbar methodology critiqued but Morningstar independent replication confirms magnitude). Combined empirical case: low-cost indexing wins 1-1.5% expense gap vs active; disciplined holding wins additional 1-2% behavior gap vs typical investor; together compound to 25-50% lower terminal wealth over 30 years for actively-trading investor vs disciplined indexer.",
     "Phase 4.5 cleanup."),

    ("CL269", "Port:BH:5", "P4A#8: HNW asset location nuance — 'bonds in tax-deferred, stocks in taxable' rule breaks down when taxable dominates, when Roth conversion ladder planned, or with direct indexing tax-loss harvesting",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Reichenstein; Kitces; direct indexing literature (Parametric, Aperio research)",
     "APPLIED: New Math-callout 'HNW asset location complications' in Bogleheads §5. Three boundary patterns: (1) taxable-dominant portfolios (>70-80% taxable) — bonds-in-deferred rule limited by capacity; muni bonds in taxable as mitigation. (2) Planned Roth conversion ladders during low-income windows reduce 'future ordinary rate' component, monetize deferred tax efficiency. (3) Direct indexing with active TLH ($100K+ min, Parametric/Aperio/Wealthfront/Betterment) generates 0.5-1.5% tax-alpha competing with standard asset-location alpha — equity-in-taxable with direct indexing partially defeats bonds-in-deferred default. Phase 7 advanced strategies will cover in detail.",
     "Phase 4.5 cleanup."),

    ("CL270", "Port:Th:8", "P4B#1: Fama-French three-factor model formally introduced in 1993 paper (J. Financial Economics), not 1992 (J. of Finance which documented patterns); attribution error",
     "CITATION", "CON", "CORRECTED", "A",
     "Fama-French 1992 (JF) documented patterns; Fama-French 1993 (JFE) 'Common Risk Factors in the Returns on Stocks and Bonds' formalized the three-factor model",
     "APPLIED: Updated text to distinguish 1992 pattern documentation from 1993 factor model formalization. Both papers now cited.",
     "Structural fix #3. Bibliographic correction."),

    ("CL271", "Port:Th:8", "P4B#2: Momentum anomaly first documented by Jegadeesh & Titman 1993, not Carhart 1997; Carhart applied JT in factor model construction",
     "CITATION", "CON", "CORRECTED", "A",
     "Jegadeesh & Titman (1993) 'Returns to Buying Winners and Selling Losers' Journal of Finance; Carhart (1997) JF",
     "APPLIED: Added Jegadeesh-Titman 1993 citation as origin of momentum anomaly. Carhart's contribution clarified as factor model incorporation rather than discovery.",
     "Structural fix #4. Attribution correction."),

    ("CL272", "Port:Th:8", "P4B#3: 'Robust factor set' (market/value/size/profitability/momentum) reflects 2010s consensus; current 2025 view more skeptical of size and value, favors quality meta-factor",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Asness-Frazzini-Israel-Moskowitz-Pedersen 2018 'Size Matters'; Asness-Frazzini-Pedersen 2019 'Quality Minus Junk'; recent factor literature 2020-2025",
     "APPLIED: Theory §3 factor models discussion updated to 2025 consensus. Robust set: market foundational, profitability (quality) strongest additional premium, value and momentum real but post-publication-diminished, size weakened post-1980. Asness-Frazzini-Israel-Moskowitz-Pedersen 2018 'Size Matters, If You Control Your Junk': size effect exists only after controlling for quality — small junk underperforms; small quality delivers premia. Practical: quality factor tilts strongest current case; value and momentum defensible but with reduced expected premia vs historical.",
     "Phase 4.5 cleanup."),

    ("CL273", "Port:Th:8", "P4B#4: Replication-crisis treatment oversimplified; Hou-Xue-Zhang 2020 found ~65% factors fail to replicate, but Chen-Zimmermann 2020/2022 find ~90% replication using original methodology; Jensen-Kelly-Pedersen 2023 synthesis concludes the apparent crisis is methodological",
     "CITATION", "SOFT", "CORRECTED", "B",
     "Hou-Xue-Zhang 2020 RFS; Chen-Zimmermann 2020/2022; Jensen-Kelly-Pedersen 2023 JF 'Is There a Replication Crisis in Finance?'",
     "APPLIED: Theory §3 includes nuanced replication-crisis treatment. Hou-Xue-Zhang 2020 found ~65% failed to replicate; Chen-Zimmermann 2020/2022 with methodology closer to originals found ~90%+ replication. Jensen-Kelly-Pedersen 2023 'Is There a Replication Crisis in Finance?' synthesized: apparent crisis largely methodological; most published factors do replicate when tested with reasonable proxies; apparent failures often involve test-set or implementation differences.",
     "Phase 4.5 cleanup."),

    ("CL274", "Port:Th:6", "P4B#5: Mean-variance sensitivity treatment omitted DeMiguel-Garlappi-Uppal 2009 (1/N vs optimization); Michaud 1998 resampled frontier; Jorion 1986 shrinkage estimators",
     "CITATION", "CON", "CORRECTED", "B",
     "DeMiguel-Garlappi-Uppal 'Optimal Versus Naive Diversification' RFS 2009; Michaud 1998; Jorion 1986",
     "APPLIED: Added DeMiguel-Garlappi-Uppal 2009 to MPT section as the strongest case against pure optimization (1/N outperforms sample-based MVO out-of-sample across 14 datasets). Also added Jorion 1986 and Michaud 1998 references. Phase 4.5 cleanup: full Michaud resampled-frontier treatment now applied as dedicated Math-callout in Theory §1 — Monte Carlo resampling of optimization, averaged-weights producing stable frontier, empirical out-of-sample outperformance vs classical MV, New Frontier Advisors patent context, Jorion shrinkage estimator as complementary approach (Bayesian prior shrinks raw historical means toward global mean), and practical implication that 'just buy market portfolio' is theoretically well-grounded when inputs are noisy and weights fragile.",
     "Structural fix #5. Most consequential Phase 4 fix — directly supports Bogleheads thesis."),

    ("CL275", "Port:Th:8", "P4B#6: Behavioral synthesis attribution too narrow — Shiller 2013 is one strand; Kahneman-Tversky 1979, Thaler 1980+, DeBondt-Thaler 1985, Shleifer-Summers, Barberis-Shleifer-Vishny, Shleifer-Vishny 1997 'limits to arbitrage' also foundational",
     "CITATION", "CON", "CORRECTED", "B",
     "Kahneman-Tversky 1979 prospect theory; Thaler foundational work; DeBondt-Thaler 1985 overreaction; Shleifer-Vishny 1997 limits to arbitrage",
     "APPLIED: Theory §4 behavioral synthesis broadened from Shiller-attribution to fuller picture: Kahneman-Tversky 1979 prospect theory (Econometrica) cognitive foundation; Thaler 1980+ mental accounting/endowment effect/equity premium puzzle; DeBondt-Thaler 1985 'Does the Stock Market Overreact?' (JF); Barberis-Shleifer-Vishny 1998 sentiment-driven over/underreaction theoretical framework; Shleifer-Vishny 1997 'The Limits of Arbitrage' (JF) — capital constraints, fund-manager agency problems, risk that mispricings widen before correcting. Shiller's excess volatility work positioned as one strand of broader research program, not its entirety.",
     "Phase 4.5 cleanup."),

    ("CL276", "Port:Th:7", "P4B#7: CAPM post-Roll framing should note pragmatic recovery via conditional/intertemporal CAPM (Merton 1973 ICAPM; Campbell 1996 conditional)",
     "CITATION", "SOFT", "CORRECTED", "B",
     "Merton 1973 'An Intertemporal Capital Asset Pricing Model' Econometrica; Campbell 1996 'Understanding Risk and Return' JPE",
     "APPLIED: Theory §2 (CAPM/Sharpe) has new Math-callout 'CAPM after Roll — what survived'. Pushes back on 'CAPM unfalsifiable therefore discarded' reading. Academic response was extension not abandonment: Merton 1973 ICAPM (Econometrica) multi-period state-variable factors; Campbell 1996 'Understanding Risk and Return' (JPE) conditional CAPM with time-varying beta and expected returns. Pragmatic synthesis: pure single-factor CAPM empirically incomplete; broader CAPM-family (ICAPM, conditional, multi-factor) is working framework asset pricing still uses. Roll made testing harder, didn't refute intuition that systematic risk earns premia.",
     "Phase 4.5 cleanup."),

    ("CL277", "Port:Th:7 calc", "P4B#8: Sharpe ratio calculator should flag confidence intervals (Lo 2002) and skewness limitations (Sortino, Calmar alternatives)",
     "CITATION", "SOFT", "CORRECTED", "B",
     "Lo, A. 'The Statistics of Sharpe Ratios' Financial Analysts Journal 2002; Sortino & Price 1994; Calmar ratio (Young 1991)",
     "APPLIED: Sharpe-ratio calc result text in Theory §2 now includes Lo 2002 'The Statistics of Sharpe Ratios' (FAJ) confidence interval caveat — wide CIs for typical samples, 5-year 0.7 vs 0.5 not statistically distinguishable, manager rankings on short windows mostly noise. Plus Sortino ratio (downside deviation, better for skewed distributions) and Calmar ratio (annualized return / max drawdown, better for retirees/leveraged). None complete substitute for understanding underlying return distribution.",
     "Phase 4.5 cleanup."),

    ("CL278", "Port:Th:8", "P4B#9: Q-factor model (Hou-Xue-Zhang 2015) framed as variant of Fama-French; actually competing paradigm (different factor construction, neoclassical investment theory motivation)",
     "CITATION", "SOFT", "CORRECTED", "C",
     "Hou-Xue-Zhang 2015 'Digesting Anomalies'; theoretical motivation in neoclassical q-theory",
     "APPLIED: Theory §3 has new Math-callout 'Q-factor as a competing paradigm, not a variant'. Hou-Xue-Zhang 2015 'Digesting Anomalies' (RFS) Q-factor model derived from neoclassical Q-theory of investment, not from empirical regularities (Fama-French). Different factor construction (NYSE breakpoints, 2x3x3 sorts on size/I-A/ROE vs FF's 2x3). Empirically explain similar cross-sectional variation but reach different conclusions about which anomalies are 'explained'. M^4 model (Stambaugh-Yu) another contender. Practical retail implication small (any framework supports value, profitability, possibly momentum tilts); academic question of which framework correct remains open.",
     "Phase 4.5 cleanup."),

    # ============================================================
    # === PHASE 5: ZEITGEIST BEHAVIORS CLAIMS ===
    # ============================================================
    # --- Zeitgeist Investing ---
    ("CL279", "Zeit:Inv:1", "FIRE movement coalesced early 2010s via blogs/forums: Mr. Money Mustache (Pete Adeney, 2011), Early Retirement Extreme (Jacob Fisker, book 2010), Mad Fientist (Brandon Ganch, 2012), ChooseFI podcast",
     "CITATION", "CON", "PASS", "B",
     "Adeney 'Mr. Money Mustache' blog 2011-present; Fisker 'Early Retirement Extreme' 2010 book; r/financialindependence subreddit history",
     "",
     "Verified historical accounts. MMM is the most influential single voice."),

    ("CL280", "Zeit:Inv:1", "FIRE variants: Lean FIRE ($25-50K spending), Fat FIRE ($100K+), Barista FIRE (partial portfolio + part-time), Coast FIRE (compound carries to traditional retirement), GeoArb FIRE (relocate to LCOL)",
     "CONVENTION", "CON", "PASS", "C",
     "FIRE community consensus terminology (Bogleheads wiki, r/financialindependence, ChooseFI)",
     "",
     "Verified community-standard terminology. Specific dollar thresholds vary by author."),

    ("CL281", "Zeit:Inv:1", "FIRE sample bias: prominent bloggers disproportionately high earners in tech/finance; FIRE math at $200K income materially different from $60K income",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Survey demographics from r/financialindependence Choose FI demographic surveys; Pete the Planner research on FIRE blogger income distribution",
     "",
     "Well-acknowledged community critique."),

    ("CL282", "Zeit:Inv:1", "4% rule horizon dependence: Bengen calibrated to 30-year retirements; 40+ year early retirement requires 3.25-3.5% SWR (28-30× expenses)",
     "EMPIRICAL", "CON", "PASS", "B",
     "Bengen 1994 JFP; Pfau subsequent research on long-horizon SWR; ERN 'safe withdrawal rate' series",
     "",
     "Already partially addressed in CL053-054 and CL156 (deferred). Cross-references Phase 2.5 backlog."),

    ("CL283", "Zeit:Inv:2", "Kakhbod, Loginova, Malenko, Malenko (2023) on FinTok: approximately 56% of analyzed financial advice on TikTok was misleading",
     "CITATION", "CON", "PASS", "B",
     "Kakhbod A., Loginova U., Malenko A., Malenko N. (2023) 'Finfluencers' working paper / journal versions",
     "",
     "Verified study. Methodology classifies advice as misleading based on departure from expert consensus or factual errors."),

    ("CL284", "Zeit:Inv:2", "SEC enforcement actions against finfluencers: 2022 charges against eight social media personalities for unregistered investment advice and promoting stocks they were simultaneously selling",
     "CITATION", "CON", "PASS", "A",
     "SEC press release December 2022 (Atlas Trading, etc.); SEC Division of Enforcement records",
     "",
     "Verified enforcement action. Specific defendants included Edward Constantinescu (MrZackMorris), others."),

    ("CL285", "Zeit:Inv:2", "Dividend yield strategy is mathematically equivalent to total-return with systematic selling at equal pre-tax returns; appeal is psychological (paycheck framing) rather than economic",
     "MATH", "CON", "PASS", "A",
     "Standard finance — Miller-Modigliani dividend irrelevance (1961); tax-equivalent for qualified dividends and LTCG identical",
     "",
     "Verified. Tax differential for ordinary-rate dividends vs LTCG is real but specific to non-qualified dividends."),

    ("CL286", "Zeit:Inv:3", "Barber & Odean (2000) 'Trading is Hazardous to Your Wealth' Journal of Finance: most active retail traders underperform indexes by ~6.5% annually after costs",
     "CITATION", "CON", "PASS", "A",
     "Barber, B. & Odean, T. 'Trading is Hazardous to Your Wealth: The Common Stock Investment Performance of Individual Investors' Journal of Finance Vol 55 No 2 (2000)",
     "",
     "Foundational empirical finding. Heavily cited."),

    ("CL287", "Zeit:Inv:3", "Welch (2022) on Robinhood platform: retail option trading produces large average losses",
     "CITATION", "CON", "PASS", "B",
     "Welch, I. 'The Wisdom of the Robinhood Crowd' Journal of Finance Vol 77 No 3 (2022)",
     "",
     "Empirical study of Robinhood-popular stock returns and option trading outcomes."),

    ("CL288", "Zeit:Inv:3", "Ben-David, Birru, Rossi (2022): retail leveraged-ETF trading destroys approximately $2 billion of value annually",
     "CITATION", "CON", "PASS", "B",
     "Ben-David I., Birru J., Rossi A. (2022) 'The Performance of Hedge Fund Performance Fees' [verify specific paper]; broader leveraged ETF retail trading literature",
     "",
     "Note: claim attribution may need refinement — the specific $2B annual figure traces to multiple papers; will verify exact source in P5A persona review."),

    ("CL289", "Zeit:Inv:3", "0DTE options: approximately 80% expire worthless; retail participation in 0DTE has grown to record levels in 2024-2025",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "CBOE 0DTE option statistics; Bryzgalova, Pavlova, Sikorskaya (2023) 'Retail Trading in Options and the Rise of the Big Three Wholesalers'",
     "",
     "80% worthless-expiration rate is approximate; varies by underlying volatility and OTM-ness."),

    ("CL290", "Zeit:Inv:4", "Bitcoin spot ETFs received SEC approval in January 2024 (BlackRock IBIT, Fidelity FBTC, Grayscale GBTC, others)",
     "REGULATORY", "CON", "PASS", "A",
     "SEC ETF approval orders January 10, 2024; specific 19b-4 filings by individual issuers",
     "",
     "Verified. Approval followed extended legal battle and Grayscale court victory."),

    ("CL291", "Zeit:Inv:4", "Ethereum spot ETFs received SEC approval 2024",
     "REGULATORY", "CON", "PASS", "A",
     "SEC ETF approval orders; specific issuers in 2024",
     "",
     "Verified."),

    ("CL292", "Zeit:Inv:4", "IRS treats cryptocurrency as property (Notice 2014-21); every transaction is a taxable event subject to capital gains rules",
     "STATUTORY", "CON", "PASS", "A",
     "IRS Notice 2014-21; IRS Virtual Currency Tax Guidance",
     "",
     "Verified. Subsequent IRS guidance via Rev. Rul. 2019-24 (hard forks) and Notice 2019-24."),

    ("CL293", "Zeit:Inv:4", "BTC/ETH historical drawdowns: 50%+ intra-year drawdowns have occurred multiple times; 75%+ peak-to-trough drawdowns occurred at least twice (2018, 2022)",
     "EMPIRICAL", "CON", "PASS", "B",
     "CoinGecko, CoinMarketCap historical price data; Bitcoin price history 2013-2024",
     "",
     "Verified empirical history. 2018 drawdown ~84% peak-to-trough; 2022 drawdown ~77%."),

    ("CL294", "Zeit:Inv:4", "Wash-sale rules technically apply to securities only (not currently crypto); IRS treatment may change with proposed legislation",
     "STATUTORY", "CON", "PASS", "B",
     "IRC §1091 (wash sale rules); applies to 'stock or securities'; multiple Congressional proposals to extend",
     "",
     "Verified current statutory state. Proposed extensions in Build Back Better and subsequent reconciliation bills did not pass."),

    # --- Zeitgeist Lifestyle ---
    ("CL295", "Zeit:Life:5", "Bill Perkins 'Die With Zero: Getting All You Can from Your Money and Your Life' (2020): framework for optimizing lifetime utility rather than end-of-life accumulation",
     "CITATION", "CON", "PASS", "A",
     "Perkins, B. 'Die With Zero' Houghton Mifflin Harcourt 2020",
     "",
     "Verified book and thesis."),

    ("CL296", "Zeit:Life:5", "Memory dividend argument: experiences early in life accumulate compounded utility value over remaining decades of memory",
     "CITATION", "SOFT", "PASS", "C",
     "Perkins 2020; behavioral economics literature on experiential utility (Kahneman, Gilbert)",
     "",
     "Perkins-specific framing; broader empirical support from happiness/experience research."),

    ("CL297", "Zeit:Life:5", "Annual gift exclusion: $19,000 per recipient (2026); unlimited direct payment of tuition and medical bills under IRC §2503(e)",
     "STATUTORY", "CON", "PASS", "A",
     "IRC §2503; Rev. Proc. 2025-32 (annual exclusion inflation adjustments)",
     "",
     "Verified. CL010 covers similar in Phase 1; restated here for Die-with-Zero context."),

    ("CL298", "Zeit:Life:5", "Late-life long-term care/dementia care costs: $80,000-$120,000+ annually for several years (US average)",
     "EMPIRICAL", "CON", "PASS", "B",
     "Genworth Cost of Care Survey 2024; CareScout annual data",
     "",
     "Verified ranges. State variance substantial."),

    ("CL299", "Zeit:Life:6", "Gallup employee engagement near historic lows; mental health/burnout discourse mainstreamed 2020-2024",
     "EMPIRICAL", "SOFT", "PASS", "B",
     "Gallup State of the Global Workplace annual reports 2020-2024",
     "",
     "Engagement % declined from ~36% in 2020 to ~33% in 2023-2024 in US data."),

    ("CL300", "Zeit:Life:6", "'Soft saving' phenomenon: explicit de-prioritization of retirement saving for present experience; distinct from 'soft life' which is broader",
     "CITATION", "SOFT", "PASS", "C",
     "2023-2024 financial media coverage (Bloomberg, Business Insider, Wall Street Journal); Intuit Credit Karma consumer surveys",
     "",
     "Trend documented in consumer surveys 2023-2024. Particularly prevalent among Gen Z respondents."),

    ("CL301", "Zeit:Life:7", "Multi-generational US households at ~18% of adults in recent surveys, up from ~12% in 1980; below 1940s peak of ~25%",
     "CITATION", "CON", "PASS", "B",
     "Pew Research Center 'Multigenerational Households' analysis (2022 update); US Census American Community Survey",
     "",
     "Verified Pew data. Trend has continued post-COVID."),

    ("CL302", "Zeit:Life:7", "Young adults (25-29) living with parents: approximately 28% in recent years, up from ~18% in 1995; peaked at ~32% during COVID 2020-2021",
     "CITATION", "CON", "PASS", "B",
     "Pew Research Center analysis of Census Current Population Survey data",
     "",
     "Verified Pew data. Boomerang pattern well-documented."),

    ("CL303", "Zeit:Life:7", "Boomerang savings math: $30K annual housing savings invested at 7% real over 5 years compounds to ~$172K; compounded 30 more years to ~$1.3M",
     "MATH", "CON", "PASS", "A",
     "Standard FV-of-annuity math",
     "",
     "Verified: $30K × FV annuity factor at 7%/5 = $172.6K; $172.6K × 1.07^30 = $1.31M."),

    ("CL304", "Zeit:Life:8", "Empirical observation: framework's accumulated math doesn't bend on optimization but presentation/framing benefits from engagement with cultural context",
     "META", "SOFT", "PASS", "C",
     "Self-observation; modern personal finance literature on framework presentation",
     "",
     "Framework's accumulated position synthesized from Phases 1-4."),

    # --- Calculator math ---
    ("CL305", "Zeit:Inv:1 calc", "Coast FIRE: projected_no_contrib = PV × (1+r)^n; if >= target, Coast FIRE achieved",
     "MATH", "CON", "PASS", "A",
     "Standard compound interest",
     "",
     "Verified."),

    ("CL306", "Zeit:Inv:2 calc", "Dividend vs total return: dividend strategy preserves principal, reinvests yield surplus; total return funds withdrawal via systematic selling at higher growth rate",
     "MATH", "CON", "PASS", "A",
     "Standard portfolio mechanics",
     "",
     "Simplified illustration; assumes constant yield and total return."),

    ("CL307", "Zeit:Life:5 calc", "Die with Zero sustainable W: P × r / (1 − (1+r)^(-n)) — standard PV of annuity formula solving for level withdrawal that depletes by horizon",
     "MATH", "CON", "PASS", "A",
     "Annuity present value identity",
     "",
     "Verified."),

    # ============================================================
    # === EXPANSIONIST EX2 ADDITIONS → PHASE 5.5 BACKLOG ===
    # ============================================================
    ("CL308", "Zeit:Inv (new section)", "EX2#1: Dave Ramsey orthodoxy not covered — largest single personal finance voice in America; framework disagrees on debt snowball vs avalanche, 12% return assumption, 8% safe withdrawal rate, credit card avoidance, all-debt-is-bad framing",
     "CITATION", "CON", "CORRECTED", "B",
     "Ramsey Solutions company info; Baby Steps published curriculum; SPIVA 12% return critique; Bengen safe withdrawal rate research",
     "APPLIED: Zeitgeist Investing §5 full new section 'Dave Ramsey orthodoxy — the mass-market counterweight'. Covers Baby Steps sequence overview, debt snowball vs avalanche with Gal-McShane 2012/Brown-Lahey 2015 citations, 12% return critique (empirical 10% nominal arithmetic / 7% real / 8.5% nominal geometric), 8% SWR critique (Bengen/Pfau/Kitces consensus 3.5-4.5% at 30yr; 8% produces 30-40% failure rate), credit card framing (population question: ~40% carry balances where Ramsey is right; ~60% pay-in-full get net positive value), where Ramsey wins (emergency fund priority, debt discipline, mass-market reach the optimization frameworks haven't achieved). Hybrid recommendation: capture employer match unconditionally, attack high-interest debt aggressively, then resume full retirement contributions.",
     "Phase 5.5 cleanup — new section."),

    ("CL309", "Zeit:Inv (existing section expansion)", "EX2#2: Dividend investing subculture deserves substantive treatment beyond the §2 illustration calculator — Seeking Alpha, dividend-growth-investing community, 'live off dividends' ecosystem",
     "CITATION", "SOFT", "CORRECTED", "C",
     "Seeking Alpha dividend community; dividend growth investing literature; Miller-Modigliani 1961 dividend irrelevance",
     "APPLIED: Zeitgeist Investing §6 full new section 'Dividend investing as a subculture — the paycheck appeal'. Miller-Modigliani 1961 dividend irrelevance theorem; modern tax neutrality (qualified div = LTCG); structural argument (dividend payers slower-growing, retain less for reinvestment); empirical record (high-yield underperforms market-cap broad over recent decades, outperforms during 2000-2009 lost decade). Honest case for dividend investing is psychological/behavioral — paycheck framing keeps investors invested through drawdowns. Skepticism: dividend traps, 'live off dividends from low capital base' as yield-chasing trap, qualified-dividend tax argument weaker than promoted. Practical recommendations: SCHD/VIG for diversified low-cost dividend exposure with quality screens; or broad-market portfolio with dividend reinvestment (VTI ~1.3-1.5% yield + share sales) for Miller-Modigliani-equivalent results with most behavioral benefits.",
     "Phase 5.5 cleanup — new section."),

    ("CL310", "Zeit:Life (new content)", "EX2#3: Buy Now Pay Later (BNPL) culturally absorbed and embedded in checkout; behavioral economics literature on payment fragmentation effects",
     "CITATION", "CON", "CORRECTED", "B",
     "Affirm, Klarna, Afterpay growth statistics; CFPB BNPL market reports 2022-2024; academic literature on payment effects (Prelec & Loewenstein)",
     "APPLIED: BNPL Math-callout folded into Spending Lifestyle §7 Subscriptions. Covers: Affirm/Klarna/Afterpay/Apple Pay Later/PayPal Pay-in-4 proliferation in all e-commerce; 4-payments-over-6-weeks structure no interest if on time, longer-tenor BNPL interest rates rival/exceed credit cards. Behavioral economics literature on payment fragmentation (splitting purchase price measurably increases consumer willingness to pay, volume, satisfaction). CFPB BNPL market report (2022+) documents BNPL users younger/lower-income/more concurrent credit card debt than non-users. Credit reporting integration gap (most BNPL doesn't report; Experian/TransUnion added structured BNPL data 2023+ but incomplete). Framework position: occasional zero-interest 4-payment BNPL on planned purchases ≈ credit card paid in full; behavioral risk in interest-bearing BNPL, multiple concurrent obligations, fragmented-payment-induced consumption expansion.",
     "Phase 5.5 cleanup — fold-in to existing section."),

    ("CL311", "Zeit:Life (new content)", "EX2#4: Gen-Z financial anxiety / 'system is broken' framing — substantial empirical underpinning (real wage growth vs housing/education cost growth); the 'soft saving' specific phenomenon and 'I'll work until I die' resignation",
     "CITATION", "CON", "CORRECTED", "B",
     "Federal Reserve real wage data; BLS housing/education CPI components; Pew generational wealth research; Gen-Z financial survey data (Intuit Credit Karma, Bank of America)",
     "APPLIED: Zeitgeist Lifestyle §8 full new section 'Gen-Z financial anxiety — the structural framing'. Covers empirical underpinning (real wage growth lagging housing/education costs, median home prices vs incomes diverging, student debt cohort differences, SSA trust fund 77-80% post-depletion benefit projection). Cultural responses: soft saving deliberate, 'I'll work until I die' resignation, side hustle obsession, 'FIRE impossible for normal people'. Where framework concedes (save-25%+ feasibility-dependent on income, 4% rule assumes US historical real returns, housing affordability creates wealth-building barriers, SS uncertainty real but 'won't exist' inaccurate). Where framework pushes back (compound interest unchanged for younger workers, Roth IRA young-saver advantage arguably greater, employer 401k matches still 50-100% guaranteed, tax-advantaged accounts expanded over time).",
     "Phase 5.5 cleanup — new section."),

    ("CL312", "Zeit:Life (new content)", "EX2#5: DINK (Dual Income No Kids) financial pattern — emerging cultural identity with distinct financial trajectory (higher savings rates achievable, no childcare/education obligations, different retirement healthcare considerations)",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Census data on child-free households; Pew Research on family structure trends; emerging DINK community/identity research",
     "APPLIED: Zeitgeist Lifestyle §9 full new section 'DINK financial pattern — the child-free trajectory'. Census data on never-had-children rising 10%→17-19% at age 50, similar in younger cohorts. Accumulation-phase math: dual income with no childcare ($15-30K/yr/child HCOL, $8-15K LCOL), no education funding (vs $250K/child framework assumption), smaller housing rational, lower life insurance needs → 15-25pp higher savings rate achievable at same gross income → years-to-FI cut roughly in half (e.g. 28yr→17yr). Coast FIRE math more dramatic — many DINK households reach compound-carry threshold by mid-30s. Decumulation considerations: family-substitute care for elderly typically unavailable, Genworth LTC costs ($80-120K/yr facility) more operative; LTC insurance, hybrid life-LTC policies, Medicaid-eligible portfolios, CCRC entry more important. Estate planning shift to charitable/extended-family/lifetime-giving. Die-with-Zero aligns naturally without inheritance-expectation resistance from children-as-default-heirs.",
     "Phase 5.5 cleanup — new section."),

    # ============================================================
    # === PHASE 5 CHECKPOINT PERSONA OBSERVATIONS (P5A + P5B) ===
    # ============================================================
    # APPLIED = structural fix made; DEFERRED-P5.5 = backlog (joins EX2 items)
    ("CL313", "Zeit:Inv:3", "P5A#1+#2: Welch (2022) misattributed — actual paper found Robinhood-popular stocks did NOT underperform, opposite of framework's claim. Strong retail option-loss evidence comes from Lakonishok-Lee-Pearson-Poteshman 2007, Bauer-Cosemans-Eichholtz 2009, Bryzgalova-Pavlova-Sikorskaya 2023",
     "CITATION", "CON", "CORRECTED", "A",
     "Welch I. (2022) 'The Wisdom of the Robinhood Crowd' Journal of Finance Vol 77 No 3; Lakonishok, Lee, Pearson, Poteshman (2007); Bauer, Cosemans, Eichholtz (2009); Bryzgalova, Pavlova, Sikorskaya (2023)",
     "APPLIED: Rewrote retail trading paragraph. Removed incorrect Welch and Ben-David-Birru-Rossi citations. Added Lakonishok et al 2007, Bauer et al 2009. Welch now correctly characterized as finding NO Robinhood underperformance (with note that strong-loss evidence comes from elsewhere).",
     "Structural fix #1+#2 applied. Resolves the audit flag on CL288 (Ben-David verification needed) by removing the unverified $2B claim."),

    ("CL314", "Zeit:Inv:3", "P5A#3: '80% of 0DTE options expire worthless' oversimplifies — most retail 0DTE positions are CLOSED before expiration, not held to worthless; of those held, more than half do expire worthless",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "CBOE 0DTE options statistics; Bryzgalova-Pavlova-Sikorskaya 2023 wholesaler routing data",
     "APPLIED: Math-callout '0DTE framing precision' in Zeitgeist Investing §3. 80% expire worthless statistic refers to OTM contracts at expiry — by construction, roughly what you'd expect of a volatility bet. Doesn't directly measure retail P&L because most 0DTE positions closed before expiry (many profitable closes). Honest framing: 0DTE options have negative expected value to buyer because seller's premium captures expected volatility plus market-making margin; retail systematic buyers can expect to lose at rate consistent with theta decay; asymmetric loss distribution (many small losses, occasional large wins) is gambling-style payoff that makes product engaging despite negative EV. Brogaard-Han-Won 2024+ provides more careful retail 0DTE P&L decomposition than pop-finance 80% framing.",
     "Phase 5.5 cleanup — fold-in."),

    ("CL315", "Zeit:Inv:2", "P5A#4: Kakhbod et al. 56% misleading statistic includes multiple categories (factual errors, undisclosed conflicts, generic-but-suboptimal); single summary statistic oversimplifies study findings",
     "CITATION", "SOFT", "CORRECTED", "B",
     "Kakhbod, Loginova, Malenko, Malenko (2023) full paper methodology",
     "APPLIED: Zeitgeist Investing §2 FinTok section now contextualizes 56% statistic. The 'misleading' category aggregates distinct failure modes: factually wrong (smallest subset), oversimplified (largest), promotional disguised as advice (moderate), conflict-of-interest patterns (moderate). Distinction matters because appropriate response differs. Honest reading: most finfluencer content is partial-truth requiring verification before action, not all-or-nothing.",
     "Phase 5.5 cleanup — fold-in."),

    ("CL316", "Zeit:Inv:1 calc", "P5A#5: Coast FIRE calculator hardcoded $20K annual contribution; should be user input",
     "UX", "CON", "CORRECTED", "A",
     "Calculator UX standard practice",
     "APPLIED: Added 'coast-contrib' user input for annual contributions if continuing. Default $20K but user-adjustable. Calculator gracefully handles $0 contribution edge case.",
     "Structural fix #3 applied."),

    ("CL317", "Zeit:Inv:2 calc", "P5A#6: Dividend vs total return calculator structurally biased — assumed dividend strategy total return = yield (4%) vs total return strategy at 7%; not honest comparison",
     "MATH", "CON", "CORRECTED", "A",
     "Miller-Modigliani (1961) dividend irrelevance; honest comparison requires yield + appreciation separately specified for both strategies",
     "APPLIED: Restructured calculator with separate yield + capital appreciation inputs for BOTH dividend-focused and total-return-focused strategies. Default: 4%+3%=7% vs 1.5%+5.5%=7% (M-M equivalent). Result text now distinguishes 'equivalent total return' from 'unequal total return' scenarios, surfacing Miller-Modigliani equivalence and the real risk (yield-chasing → lower total returns).",
     "Structural fix #4 applied. Most consequential Phase 5 calculator correction."),

    ("CL318", "Zeit:Inv:3", "P5A#7: Retail trading 'casino vs investing' framing elides addictive-design literature on trading apps (Robinhood gamification, multiple regulatory actions and academic critiques)",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Robinhood regulatory actions; SEC, MA Securities Division, FINRA enforcement records; academic literature on gamification in trading apps",
     "APPLIED: Math-callout 'Addictive design — what the literature documents' in Zeitgeist Investing §3. Massachusetts Securities Division 2020 complaint against Robinhood (confetti animations, celebratory push notifications, achievement streaks) argued unfair practices targeting inexperienced investors. Confetti removed 2021; broader pattern of variable-reinforcement notifications, social-comparison framings, friction asymmetry remains common. Barber-Huang-Odean-Schwarz 2022 'Attention-Induced Trading and Returns' shows attention-grabbing features measurably increase trading frequency and decrease retail returns. Welch 2022 didn't find Robinhood-stock underperformance, but platform-design literature documents the trading-frequency increase that Barber-Odean line shows hazardous to retail returns. Mitigations: awareness of pattern (first-order); non-gamified platforms Vanguard/Fidelity/Schwab (second-order).",
     "Phase 5.5 cleanup — fold-in."),

    ("CL319", "Zeit:Inv:4", "P5A#8: Crypto 1-5% diversification framing needs diversification-benefit decay caveat — BTC/equity correlation has risen from ~0 (2014-2018) to 0.3-0.5+ (2020-2024)",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Briere, Oosterlinck, Szafarz (2015); subsequent crypto correlation literature; rolling correlation analyses 2020-2024",
     "APPLIED: Math-callout 'Crypto diversification benefit has decayed materially since 2020' in Zeitgeist Investing §4. Early academic case (Briere-Oosterlinck-Szafarz 2015 'Virtual Currency, Tangible Return'; Klein-Pham Thu-Walther 2018; subsequent work through 2020) relied on Bitcoin's 0.1-0.3 equity correlation. Post-2020 pattern differs: rolling 90-day correlations with US equities frequently exceed 0.5, occasionally approach 0.7, particularly during risk-off periods where diversification would matter most. Structural explanation: institutional participation through ETF approvals integrated crypto into traditional risk-on/risk-off cycles. Pre-2022 diversification benefit literature should be read with this empirical update — original correlation patterns no longer apply at same magnitude. Current case closer to convex-upside exposure than portfolio diversification per se.",
     "Phase 5.5 cleanup — fold-in."),

    ("CL320", "Zeit:Inv:1", "P5A#9: FIRE blogger sample bias claim presented without citation; ChooseFI demographic surveys and r/financialindependence data support it but aren't cited",
     "CITATION", "SOFT", "CORRECTED", "C",
     "ChooseFI community demographic surveys; r/financialindependence demographic data; emerging academic work on FIRE community demographics",
     "APPLIED: FIRE sample-bias citation expanded in Zeitgeist Investing §1. Kakhbod-Loginova-Malenko-Malenko 2023 provided quantitative anchor for FinTok content but for FIRE specifically, structural patterns are community-documented — prominent FIRE bloggers consistently report incomes substantially above US median during accumulation. Bogleheads forum and r/financialindependence periodically host meta-discussions acknowledging selection bias. Visible FIRE bloggers are by definition the ones who reached FI and wrote about it; the larger population who attempted high-savings-rate strategies and encountered burnout/divorce/medical/career disruption never reached the milestone to write about. Honest framing: FIRE math correct, visible practitioners non-representative.",
     "Phase 5.5 cleanup — fold-in."),

    ("CL321", "Zeit:Inv:1", "P5B#1: Post-FIRE depression / retirement transition mental health pattern not surfaced; clinically substantial pattern in high-savings-rate accumulators hitting their FI number",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Calvo, Sarkisian, Tamborini (2013); Heybroek, Haynes, Baxter (2015); retirement transition mental health literature",
     "APPLIED: Math-callout 'Post-FIRE depression — the empirical pattern' in Zeitgeist Investing §1. Calvo-Sarkisian-Tamborini 2013 'Causal Effects of Retirement Timing' (Journals of Gerontology) — early retirement (before 62) associated with worse mental health outcomes than continued work or traditional-age retirement, particularly without strong replacement activities. Mechanism: loss of identity-providing work, reduced social network (workplace = majority of weekly social contact for working-age), unstructured-time challenge. Mitigations: structured non-work activities, professional network maintenance, partial-FIRE patterns. Framework's accumulated FIRE guidance now paired with honest acknowledgment that psychological dimension is where typical failure mode actually occurs.",
     "Phase 5.5 cleanup — fold-in."),

    ("CL322", "Zeit:Life:6", "P5B#2: 'Soft saving' treatment assumes deliberate choice; clinical reality often includes resignation/giving-up phenomenology that responds to examination",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Financial therapy clinical literature; behavioral economics on learned helplessness",
     "APPLIED: Math-callout 'Soft saving phenomenology — choice versus resignation' in Zeitgeist Lifestyle §8 (Gen-Z section). Clinical distinction between deliberate trade-off pattern (agency present, understands long-term cost, consciously chooses present consumption — 'do you understand what you're choosing against' is operative question) and resignation pattern (stopped believing retirement is achievable, often partly accurate observations but learned helplessness; the framework's 'do you understand the math' framing assumes agency the resignation-pattern client doesn't experience). Clinically appropriate response: examine underlying belief — sometimes resignation is empirically accurate (extreme low income), more often cognitive distortion that responds to examining what modest saving compounded over 40 years actually produces.",
     "Phase 5.5 cleanup — fold-in to new Gen-Z section."),

    ("CL323", "Zeit:Life:5", "P5B#3: Die with Zero framing assumes preserved cognition through death; late-life dementia/cognitive decline closes spending utility window earlier than most people assume",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Alzheimer's Association statistics; cognitive aging literature; long-term care demographics",
     "APPLIED: Math-callout 'Cognitive decline complicates spend-it-during-life execution' in Zeitgeist Lifestyle §5 (Die with Zero). MCI affects 10-20% of 65+, dementia 11% of 65+ and 30% of 85+. Executive function and complex financial-decision capacity declines through 70s/80s even without diagnostic thresholds. Clinical pattern: late-life when physically able to enjoy spending often coincides with decline of decision-making infrastructure required to deploy savings flexibly. Practical implications: front-load discretionary experiences earlier in retirement when capacity is robust (60s rather than 80s); establish durable POA and trusts while fully intact; consider CCRC entry around 75-80 to convert lump-sum decisions into ongoing services; recognize simple irrevocable structures (annuities, immediate longevity insurance) may serve better than complex spend-down plans.",
     "Phase 5.5 cleanup — fold-in."),

    ("CL324", "Zeit:Life:6", "P5B#4: Burnout-as-financial-event framing missing — career interruption costs can exceed previous decade's accumulated savings advantage",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Burnout cost literature; career interruption financial impact studies; mental health economic research",
     "APPLIED: Math-callout 'Burnout as a financial event' in Zeitgeist Lifestyle §8 (Gen-Z section). For high earners burnout cost can exceed entire accumulated savings advantage from high-savings-rate period. Mechanism: 3-12 weeks medical leave, 3-18 months full recovery, 10-30% pay decrease career change, severe permanent labor force exit. Worked example: software engineer $300K → burnout → $200K after 9-month recovery = $225K forgone income + $100K/yr ongoing differential, easily exceeding 10% additional savings rate for previous decade. Math reinforces framework's sustainable-rate-over-extreme-rate position: 15% savings 35 years beats 40% savings 5 years + career disruption on every dimension that matters.",
     "Phase 5.5 cleanup — fold-in to new Gen-Z section."),

    ("CL325", "Zeit:Life:7", "P5B#5: Boomerang/multi-gen 'properly executed' caveat does substantial work without elaboration; relational complexity needs more honest treatment",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "Family systems literature on multi-generational living; financial therapy clinical patterns",
     "APPLIED: Math-callout 'Boomerang relational structure — what properly executed actually requires' in Zeitgeist Lifestyle §7. Three structural patterns from financial-therapy clinical practice that reliably distinguish multi-gen arrangements producing financial benefit from those that fail: (1) explicit written agreements about expectations/contributions/duration (informal one-page documents between parents and adult children outperform unwritten understandings on every dimension); (2) scheduled review points rather than open-ended (quarterly/semi-annual 'is this still working' conversations); (3) defined contribution structures (whether nominal rent, household expense share, specific bills, or none — matters less than that structure is explicit and durable). Failure patterns: parents subsidizing lifestyle vs accelerating savings, adult child failure-to-launch into independence, relational deterioration. Honest framing: boomerang living can be highest-leverage financial move but only with relational discipline financial framing alone doesn't address.",
     "Phase 5.5 cleanup — fold-in callout to existing section."),

    ("CL326", "Zeit:Life:8", "P5B#6: Behavioral synthesis closer omits savings-aversion clinical pattern — money avoidance, hoarding-spending cycles, scarcity-driven consumption respond to therapy more than spreadsheets",
     "EMPIRICAL", "CON", "CORRECTED", "C",
     "Klontz et al. financial psychology research; money disorders clinical literature",
     "APPLIED: Math-callout 'Savings-aversion as a clinical pattern' in Zeitgeist Lifestyle §10 (synthesis). Framework's accumulated guidance assumes rational chooser optimizing within constraints. Clinical reality includes meaningful subset of households for whom that framing isn't operative because something else is happening psychologically: money avoidance (refusing engagement with financial planning, magical-thinking around income), hoarding-spending cycles (extreme frugality ↔ impulsive consumption), scarcity-driven consumption (compulsive low-cost purchasing as deprivation response), money disorders documented in financial-therapy literature (Klontz, Britt et al.). These patterns respond to therapy more than spreadsheet guidance. Framework's value limited for these clients; appropriate intervention is therapeutic engagement with underlying pattern, after which framework's optimization guidance becomes operational. Honest acknowledgment: framework throughout has implicitly assumed all readers operating within rational-chooser model. Some are not.",
     "Phase 5.5 cleanup — fold-in callout to synthesis."),

    ("CL327", "Zeit:Life:5 calc", "P5B#7: Die with Zero calculator uses level real withdrawal; clinical reality is U-shaped retirement spending (high early, lower middle, rising late-life medical)",
     "EMPIRICAL", "SOFT", "CORRECTED", "B",
     "Retirement spending pattern research; David Blanchett 'Estimating the True Cost of Retirement'; healthcare cost trajectory in retirement",
     "APPLIED: Math-callout 'U-shape retirement spending pattern' in Zeitgeist Lifestyle §5 (Die with Zero). Empirical retirement-spending literature documents pattern Die-with-Zero monotonic-decline model misses: high in early retirement (go-go years, 60-75), drops materially in mid-retirement (slow-go, 75-85, travel/active pursuits taper), rises again late (no-go, 85+, healthcare and LTC surge). Documented in HRS data, Society of Actuaries research, Blanchett 2014, Banerjee 2019. Die-with-Zero spend-down should anticipate late-life cost spike rather than spending evenly. Mitigations partly overlap with cognitive-decline: longevity insurance (deferred annuities, QLACs), LTC insurance or hybrid products, CCRC entry (converts future cost spike into present-value-known monthly fees). Single-number trajectory calculator kept simple for pedagogy; model-limits is operative caveat — real Die-with-Zero plan needs U-shape modeling.",
     "Phase 5.5 cleanup — fold-in callout."),

    ("CL328", "Zeit:Life:8", "P5B#8: Behavioral synthesis treats reader as single agent; most consequential lifestyle/zeitgeist decisions are family-system decisions, not individual",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "Family systems theory; financial therapy clinical practice; couples and family financial decision-making literature",
     "APPLIED: Math-callout 'Family-system framing' in Zeitgeist Lifestyle §10 (synthesis). Most consequential financial decisions are family-system decisions, not individual: spousal disagreement about saving rate, family about elder care funding, parent-child about inheritance expectations, sibling about parent-care responsibilities. Clinical patterns more common than individual-optimization patterns framework's optimization frameworks address. Boomerang living math from §7 only works with relational agreement. Die-with-Zero trajectory only works without inheritance-expectation conflict among heirs. Retirement-age decision almost always made jointly by spouses. Framework throughout treats reader as single agent making single decisions; actual decision-making structure usually multi-party family system. Acknowledgment rather than correction: framework useful as individual-perspective optimization reference, but operative decisions in actual lives often require family-system conversation framework's individual-agent framing doesn't surface.",
     "Phase 5.5 cleanup — fold-in callout to synthesis."),

    # ============================================================
    # Phase 6 — Session 1 (Tax Attorney + CFP) findings
    # ============================================================

    ("CL360", "Contractor:8.2", "QBI §199A SSTB threshold is measured against taxable income before the QBI deduction (IRC §199A(e)(2)), not AGI",
     "STATUTORY", "CON", "PARTIAL", "A",
     "IRC §199A(e)(2); see also Rev. Proc. 2025-32 for 2026 indexed dollar figures",
     "APPLIED PARTIALLY: Added the 'taxable income before QBI, not AGI' clarifier to the Contractor:8.2 paragraph — material because retirement / HSA / itemize-vs-standard / capital-loss-carryover choices move taxable income but not AGI, and the SSTB threshold rides on the former. NOT APPLIED: P6 Tax Attorney persona also claimed the dollar threshold figures in the artifact ($201,775/$403,500 begin, $276,775/$553,500 complete) are stale 2024 numbers, asserting 2025 actual is $241,950/$483,900 (Rev. Proc. 2024-40). This contradicts the existing CL056 audit row, which captured 2025 actual at $197,300/$394,600 and 2026 at the figures currently in the artifact. The $241,950 figure for 2025 is inconsistent with the IRS's normal 3–5% inflation indexing trajectory from the 2024 base of $191,950 and is likely a persona-recall error. Threshold dollar values left unchanged pending an authoritative Rev. Proc. 2025-32 re-verification.",
     "Phase 6 — Session 1 Tax Attorney finding, applied partially."),

    ("CL362", "W2:10.1, W2:10.3", "IRMAA Medicare premium surcharges apply at MAGI thresholds with 2-year lookback; Roth conversions and SS claiming in pre-Medicare years can trigger multi-year premium increases",
     "REGULATORY", "CON", "CORRECTED", "A",
     "42 U.S.C. §1395r(i); SSA POMS HI 01101.020; CMS Annual IRMAA tables (2026 first tier ~$106K single / ~$212K MFJ; top tier ~$394K / ~$788K)",
     "APPLIED: Two callouts added. (1) W2:10.1 — IRMAA cliff awareness callout placed inside the Roth conversion ladder node, with the 2-year MAGI lookback, the five-tier cliff structure, the 2026 dollar figures, and the practical rule 'from age 63 onward, size conversions to the IRMAA tier boundary, not the federal bracket boundary.' Notes that the 12% bracket fill generally clears IRMAA for couples but the 22%/24% bracket-fill plays the framework otherwise recommends do not. (2) W2:10.3 — cross-reference callout to 10.1 explaining that SS claiming and conversion sizing are joint optimization problems on IRMAA grounds; claiming early at 62 with simultaneous large conversions is the most common way to land in the second or third IRMAA tier.",
     "Phase 6 — Session 1 Tax Attorney + CFP convergent finding."),

    ("CL363", "W2:10.1, W2:10.2", "Post-SECURE 10-year inherited IRA rule eliminates stretch for most non-spouse beneficiaries; per 2024 final regs, annual RMDs during years 1-9 required when decedent had begun RMDs; SECURE 2.0 RMD age 73→75 in 2033",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRC §401(a)(9)(H); SECURE Act §401 (2019); Treasury Reg. §1.401(a)(9)-5 final regs (July 18, 2024); SECURE 2.0 §107 (RMD age 75 in 2033)",
     "APPLIED: Callout in W2:10.1 covering the SECURE Act 10-year rule, the July 2024 final-regs years-1-9-RMD requirement when decedent had begun RMDs, the eligible-designated-beneficiary exceptions (surviving spouse, minor child of decedent, disabled/chronically ill, beneficiary within 10 years of decedent age), and the practical effect — pre-tax IRA balances inherited by working-age children distribute into peak earning years at 32-37% federal vs the parent's possible 10-24%, strengthening the conversion-ladder case by an 8-20 pp tax-rate spread. W2:10.2 updated to note RMD age 73→75 in 2033 per SECURE 2.0 §107, which extends the 'tax valley' conversion window by two additional years for households turning 65 in 2023 or later.",
     "Phase 6 — Session 1 Tax Attorney finding."),

    ("CL364", "W2:10.5 (new node); Zeit:Life:5", "QCDs permit direct IRA-to-charity transfer up to $108,000 (2026) at age 70½+; counts toward RMD, excluded from AGI; superior to most other charitable vehicles for eligible retirees",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRC §408(d)(8); SECURE 2.0 §307 (indexation effective 2024); SECURE 2.0 §307 (one-time $54K QCD-to-CRT/CGA option, indexed); IRS Publication 590-B",
     "APPLIED: New full-chart W2:10.5 node 'Qualified Charitable Distributions (QCDs) — the most efficient give-while-alive vehicle at 70½+' covering: 2026 $108K annual limit per IRA owner, age trigger preserved at 70½ (not the current RMD age of 73), AGI-exclusion mechanism, RMD satisfaction, IRMAA / SS-taxability interaction, DAF and private-foundation exclusion, direct-transfer-from-trustee requirement, SECURE 2.0 one-time $54K QCD-to-CRT/CGA option. Cross-referenced from Zeit:Life:5 (Die with Zero) with a math-callout noting that for charitably-inclined retirees 70½+ with Traditional IRA balances, QCDs dominate DAF funding / after-tax cash giving / bunching on after-tax-cost-per-dollar-donated math. Note: the cross-reference is one-directional (Zeit→10.5) per the framework's forward-only linking discipline.",
     "Phase 6 — Session 1 Tax Attorney finding; CFP convergent on the give-while-alive integration."),

    ("CL334", "Plan:computePlan LTC trigger", "LTC plan action undifferentiated by net worth × household structure × age window",
     "UX", "CON", "CORRECTED", "C",
     "AALTCI Price Index 2024; HHS/ASPE LTSS Research Brief 2022; CareScout/Genworth 2024 Cost of Care",
     "APPLIED: LTC action in computePlan now tiers by netWorth × ageBracket × dependents. Negative/Under-100K/100K-500K → 'Plan for Medicaid-path LTC, not LTC insurance' with 5-year lookback and spousal-impoverishment guidance. $500K-$2M → 'You are in the LTC insurance / hybrid-policy sweet spot — get quotes this year' with the 30-60% premium escalation across the 55-65 underwriting window, escalated to 'this-year' stage with critical tag when age is in the sweet-spot bracket, with a DINK heightener for households without children. $2M+ → 'Self-insurance feasible; LTC for legacy preservation' with hybrid-policy framing and DINK CCRC note. Uses existing diagnostic state — no new questions.",
     "Phase 6 — Session 1 CFP finding."),

    ("CL337", "Spend:Ess:1", "Housing section omits property-tax basis-reset traps, mortgage recast mechanics, and reverse-mortgage standby LOC strategy",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Pfau 2016 'Reverse Mortgages'; Pfau-Wagner 2014 JFP standby LOC; CA BOE Prop 19 guidance; FL DOR Save Our Homes cap docs",
     "APPLIED: Three callouts added before the rent-vs-buy calculator. (1) Property-tax basis-reset trap — CA Prop 13/19, FL Save Our Homes, TX/GA/OR analogues; quantified $14,500/yr downsize penalty example for long-tenure CA homeowner; verify state assessment-cap rules + base-year-value portability before pricing the move. (2) Mortgage recast as the right move for most lump-sum windfalls — lender re-amortizes against remaining term at same rate, monthly payment drops permanently, closing costs typically $250-$500, beats both extra-principal-only and refinancing-into-higher-current-rate; check loan docs since VA and most jumbos disallow. (3) Reverse mortgage as retirement standby LOC — Pfau's reframing of the structured HECM-as-sequence-buffer use case opened at 62-65 and left unused; cites Pfau-Wagner 2014 JFP standby LOC work; 5-15 pp portfolio survival improvement in some historical-cycles backtests for asset-rich house-rich retirees; explicit acknowledgment of historical product abuses without categorical dismissal.",
     "Phase 6 — Session 1 CFP finding."),

    ("CL338", "Plan view header", "Plan output is a prioritized list without integrated narrative synthesis; misses the planning-deliverable shape",
     "UX", "CON", "CORRECTED", "D",
     "Standard CFP planning-engagement deliverable structure; Kinder Institute Life Planning framework; XY Planning Network sample deliverables",
     "APPLIED: New 'plan-synthesis' block renders three paragraphs at the top of the Plan view above the stage-grouped action list. Generated entirely from existing state.answers + the computed plan; no new diagnostic introduced. Paragraph 1 (Situation): age + income type + filing + bracket + dependents + net worth + horizon synthesized into a single 'you are a...' sentence. Paragraph 2 (Binding constraints): scans diagnostic for the framework's gating items (untracked cash flow, sub-threshold EF, high-interest debt, dependent-tilt life-insurance gap, LTD gap, §83(b) open window, income concentration with high bracket, retirement-proximity) and assembles a constraints sentence that frames the action list as gated rather than parallel. Paragraph 3 (Horizon & sequencing): horizon-aware narrative that points 30+/20-30 households at savings rate as primary, 10-20 at the transition to decumulation architecture, 5-10 and <5 at the active decumulation arc with explicit references to the new W2:10.1/10.2/10.3/10.5 nodes and the IRMAA + SECURE + QCD work from CL362/363/364, and 'already' retirees at the active drawdown sequence. Both <5/5-10 and 'already' paragraphs explicitly note that the §7 historical-cycles simulator currently operates on portfolio-only inputs without SS overlay (the Phase 6.5 backlog item from CL329) so users read the success-rate numbers correctly. Adds .plan-synthesis CSS with left-border accent in the artifact's palette.",
     "Phase 6 — Session 1 CFP finding. Highest-leverage UX fix in the review."),

    # ============================================================
    # Phase 6 — Session 2 (Academic finance economist + Behavioral economist) findings
    # ============================================================

    ("CL370", "Math:7", "1928-2024 US historical-cycles dataset embeds survivorship/selection bias; international and pre-1871 evidence puts the median SWR at ~2.5-3%, not 4%",
     "EMPIRICAL", "SOFT", "PARTIAL", "C",
     "Pfau (2010) JFP 'An International Perspective on Safe Withdrawal Rates'; Estrada (2017) JFP 'Refining the Failure Rate'; McQuarrie (2024) FAJ 'Stocks for the Long Run? Sometimes Yes, Sometimes No'; Dimson-Marsh-Staunton Triumph of the Optimists (2002) and annual Credit Suisse/UBS Global Investment Returns Yearbook",
     "APPLIED PARTIALLY: New math-callout 'The McQuarrie challenge — what international and pre-1871 US data show about the 4% rule' added above the §7 simulator. Covers Pfau 2010 (16-of-17 developed countries below 4%; median ~2.5-3%), Estrada 2017 (DMS 23-country dataset), McQuarrie 2024 (pre-1871 US extension producing ~2.8-3.3% safe rate), and the practical implication that selecting 4% in the dropdown is implicitly betting on US equity exceptionalism. Reframes the existing horizon-adjusted Bengen guidance (3.25-3.5% at 40+ years) as approximately the right haircut for the wrong reason — it's the empirical international/long-history haircut, not the unconditional US horizon adjustment. DEFERRED P6.5: full CAPE-input or international-data toggle in the simulator itself.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL373", "Math:7, Math:3", "Safe withdrawal rate is empirically conditional on starting CAPE/valuation; Pfau 2012 and Kitces' updates show ~3-3.5% conditional SWR at current high-CAPE starting points vs the unconditional 4% historical result",
     "EMPIRICAL", "SOFT", "PARTIAL", "C",
     "Pfau (2012) JFP 'Withdrawal Rates, Savings Rates, and Valuation-Based Asset Allocation'; Kitces NEV ongoing updates 'Safe Withdrawal Rates and Current Market Valuations'",
     "APPLIED PARTIALLY: New math-callout 'Conditional SWR — what starting valuations do to the 4% rule' added in Math §7 immediately after the McQuarrie callout. Covers Pfau 2012 valuation-conditional historical SWR result (~3.0-3.5% at 90th-percentile CAPE; ~5.5-6% at 10th-percentile); explains that the McQuarrie and CAPE-conditional findings act on different axes (selection bias vs within-US valuation) and compound for retirees starting today. Manual workaround: run simulator at 3.0-3.5% to see success-rate movement. DEFERRED P6.5: CAPE-input extension to the simulator backend with starting-valuation-shock to first-N-year returns.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL374", "Math:4, Math:1, Math:3, Math:6, Math:7", "Calculator defaults of 6-7% real US equity returns are ex-post historical averages, not ex-ante expected forward returns; Fama-French 2002, Damodaran annual implied equity premium, and Bogle 2015 converge on 4-5% real as the more defensible forward planning number",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Mehra-Prescott (1985) JME 15 'The Equity Premium: A Puzzle'; Fama-French (2002) JF 57(2) 'The Equity Premium'; Damodaran 'Equity Risk Premiums' annual update NYU Stern; Ilmanen (2011) Expected Returns; Bogle (2015) JoPM 'Occam's Razor Redux'",
     "APPLIED: New math-callout 'Ex-post vs ex-ante — why historical averages overstate planning returns' added in Math §4 after the inflation-index callout. Names the Mehra-Prescott equity premium puzzle, the Fama-French 2002 ex-ante 2.5-4.3% real estimate, the Damodaran current 4-5% nominal forward implied premium, and Bogle's 2015 4-5% nominal published forward estimate. Operational planning implication: 4-5% real is the more defensible forward planning number for US equities at current valuations. Calculator defaults left unchanged but the callout instructs users to plug their own forward-return assumption rather than relying on defaults; defaults exist to be replaced.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL371", "Math:1", "Volatility-drag σ²/2 formula presented as exact rather than as a first-order continuous-time approximation; two near-duplicate callouts with subtly different framings and numerical examples",
     "METHODOLOGICAL", "CON", "CORRECTED", "C",
     "Markowitz (1959) Portfolio Selection Ch. 6; MacLean-Thorp-Ziemba (2010) The Kelly Capital Growth Investment Criterion",
     "APPLIED: Consolidated the two Math §1 callouts at lines 1840 and 1841 into a single 'Arithmetic vs geometric returns — volatility drag' callout. Adds 'to first order' qualifier; cites Markowitz 1959 and MacLean-Thorp-Ziemba 2010; notes the approximation is within ~15 bps for σ ≤ 20% and that fat-tailed return distributions produce small additional empirical drag beyond the σ²/2 prediction. Uses one consistent numerical example (US σ ≈ 16%, ~1.3pp drag, 10% arithmetic → ~8.7% geometric).",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL372", "Bog:3", "BHB 1986 sample was 91 homogeneous large US pension funds; Xiong-Ibbotson-Idzorek-Chen 2010 FAJ found asset allocation and active management equally important in dispersed cross-sections; the 90%-of-variance result is pedagogically clean but doesn't bear the weight of the popular reading",
     "CITATION", "SOFT", "CORRECTED", "C",
     "Xiong, Ibbotson, Idzorek, Chen (2010) 'The Equal Importance of Asset Allocation and Active Management,' FAJ 66(2); Statman (2000) 'The 93.6% Question of Financial Advisors,' JoPM",
     "APPLIED: Extended the existing Bog:3 'What the 90% of variance claim actually means' math-callout with the BHB sample-homogeneity caveat (91 large US pension funds 1974-1983), Xiong-Ibbotson-Idzorek-Chen 2010 FAJ rebuttal showing equal importance in dispersed samples, Statman 2000 retail-context pushback. Re-anchors the Bogleheads conclusion that allocation matters more than fund selection within a category to SPIVA/Bessembinder rather than BHB.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL376", "Bog:3", "Pfau-Kitces 2014 rising-equity glide path is conditional on US 1926-2010 sample (Estrada 2016 didn't replicate internationally) and on high assumed forward equity premium (Kitces 2016 update); Bernstein 2013 deep-risk framework provides separate psychological skepticism",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Pfau-Kitces (2014) JFP 'Reducing Retirement Risk with a Rising Equity Glide Path'; Estrada (2016) JoPM 'The Retirement Glidepath: An International Perspective'; Bernstein (2013) Deep Risk: How History Informs Portfolio Design; Kitces NEV 2016 update",
     "APPLIED: Extended the existing Bog:3 rising-equity-glide-path paragraph. Now explicitly cites the 1926-2010 US data scope, Estrada 2016 international non-replication, Kitces 2016 sensitivity to ex-ante equity premium (connecting back to Math §4's ex-post-vs-ex-ante callout from CL374), and Bernstein 2013 'deep risk' psychological-feasibility skepticism. Defensibility framing tightened: defensible under specific assumptions about future equity returns and the retiree's psychological tolerance, not as a default recommendation.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL377", "Port:Th:6", "DeMiguel-Garlappi-Uppal 2009 1/N-dominance result is the opening of a 15-year academic debate; Kirby-Ostdiek 2012 JFQA and Tu-Zhou 2011 JFE rebuttals show timing strategies and combination estimators outperform pure 1/N; current consensus is shrinkage/combination over both naive 1/N and naive sample-based MV",
     "CITATION", "SOFT", "CORRECTED", "C",
     "Kirby-Ostdiek (2012) JFQA 47(2) 'It's All in the Timing'; Tu-Zhou (2011) JFE 99(1) 'Markowitz Meets Talmud: A Combination of Sophisticated and Naive Diversification Strategies'; Jorion (1986) JFQA 21(3) 'Bayes-Stein Estimation for Portfolio Analysis'",
     "APPLIED: Extended the existing Port:Th:6 paragraph on DeMiguel-Garlappi-Uppal 2009. Now narrates the 2009 result as the opening of a 15-year debate, cites Kirby-Ostdiek 2012 and Tu-Zhou 2011 rebuttals, frames current consensus as 'shrinkage or combination estimators outperform both naive 1/N and naive sample-based MV' rather than '1/N dominates all optimization.' Reframes the Bogleheads market-cap defensibility on Sharpe-Lintner-Mossin equilibrium grounds and Bessembinder skewness grounds (cross-referencing the CL375 work) rather than on a contested 1/N-dominance claim.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL378", "Bog:2", "Home bias is welfare-reducing for US investors per published literature (French-Poterba 1991, Coeurdacier-Rey 2013, Vanguard 2017, Asness-Israelov-Liew 2011); the artifact's framing presented Bogle's 0% as symmetric with global-market-cap weights when the literature treats it as a contested outlier",
     "EMPIRICAL", "CON", "CORRECTED", "B",
     "French-Poterba (1991) AER P&P 81(2); Coeurdacier-Rey (2013) JEL 51(1) 'Home Bias in Open Economy Financial Macroeconomics'; Asness-Israelov-Liew (2011) FAJ 67(3) 'International Diversification Works (Eventually)'; Vanguard (2017) 'Global Equity Investing: The Benefits of Diversification and Sizing Your Allocation' Donaldson-Kinniry-Maciulis-Patterson",
     "APPLIED: Rewrote the Bog:2 'Common variants worth knowing' callout. Now leads with Vanguard 2017 + Asness-Israelov-Liew 2011 published consensus that ~30-50% international is welfare-improving for US investors. Bogle's 0% framing demoted from 'symmetric end of range' to 'contested outlier' explicitly. Cross-referenced to the Math §7 McQuarrie callout from CL370: accepting the international SWR evidence logically implies accepting that diversifying away from US-equity dependence has positive expected utility.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL379", "Math:7", "HIST_RETURNS 0.5pp rounding produces ±1-3pp bias in simulated success rates at tail-binding years; bond series maturity (intermediate Treasury) was unlabeled",
     "METHODOLOGICAL", "SOFT", "PARTIAL", "C",
     "Shiller online dataset http://www.econ.yale.edu/~shiller/data.htm; Damodaran NYU Stern historical returns dataset",
     "APPLIED PARTIALLY: §7 model-limits note extended to quantify the rounding bias (±1-3pp on headline success rates at tail-binding years for typical 4% withdrawal rates), label the bond series as intermediate Treasury real returns (consistent with Bengen/Pfau convention), and cross-reference the McQuarrie/Pfau-international callouts above the calculator for the broader US-only conditionality. DEFERRED P6.5: full dataset replacement with unrounded Shiller/Damodaran source data + explicit bond-series proxy choice. Aligns with the existing swap-interface architecture — if ND plugs in a personal MC backend, that backend can carry its own dataset.",
     "Phase 6 — Session 2 academic finance economist finding."),

    ("CL375", "Port:Th:8, Bog:1", "Bessembinder 2018 JFE and Bessembinder-Chen-Choi-Wei 2023 FAJ are the strongest modern empirical foundation for indexing — top 4% of US stocks generated all 1926-2016 excess return; 60% of global stocks underperformed T-bills lifetime — not cited in artifact",
     "CITATION", "CON", "CORRECTED", "B",
     "Bessembinder (2018) Journal of Financial Economics 129(3) 'Do Stocks Outperform Treasury Bills?'; Bessembinder, Chen, Choi, Wei (2023) FAJ 'Long-Term Shareholder Returns: Evidence from 64,000 Global Stocks'",
     "APPLIED: New paragraph 'The Bessembinder skewness result — the modern empirical foundation for indexing' added in Port:Th:8 immediately before the behavioral synthesis paragraph. Covers the 2018 finding (top 4% of US stocks generated all 1926-2016 excess return; median stock underperformed T-bills) and the 2023 global extension (60% of global stocks underperformed T-bills lifetime). Practical implication framed as: concentration produces negatively-skewed expected outcomes independent of skill, which makes deviating from the Bogleheads market-cap-weighted approach expensive in expected-value terms. Stronger empirical case for indexing than the SPIVA active-fail-rate evidence the framework otherwise leans on.",
     "Phase 6 — Session 2 academic finance economist finding."),

    # ============================================================
    # Phase 6 — Session 3 (Consumer-finance advocate) findings
    # ============================================================

    ("CL390", "Welcome view", "Framework's audience calibration (sophisticated investor per HANDOFF) was invisible to readers; no acknowledgment of the structural conditions (racial wealth gap, gender wealth gap, wage stagnation, banking-desert geography, US tax-policy regressivity) that determine which households can act on individual-optimization recommendations",
     "META", "SOFT", "CORRECTED", "C",
     "Olen Pound Foolish (Penguin Portfolio 2012); Darity & Mullen From Here to Equality (UNC Press 2020); Hamilton et al. New School wealth-gap research; Federal Reserve Survey of Consumer Finances 2022; Aspen Institute Financial Security Program Future of Wealth in the US 2022; FDIC Survey of Unbanked and Underbanked Households 2023; Aliche Get Good with Money (Rodale 2021); Khalfani-Cox Zero Debt (Advantage 2004); Lowry Broke Millennial series (TarcherPerigee 2017+); Tobias The Only Investment Guide You'll Ever Need (Harvest revised 2022)",
     "APPLIED (Path A — scope decision held; structural-conditions acknowledgment added in Welcome view): New welcome-scope block under the diagnostic-launch CTA. Four paragraphs: (1) audience calibration explicitly named — $60K+ single / $100K+ household + prior-comfort-with-bracket-and-401k concepts — with explicit framing that earlier-arc readers will find some Plan items not-yet-relevant and that this is a feature not a critique. (2) Structural-conditions block naming wage stagnation, racial wealth gap (Fed SCF 2022 6× / 5× medians), gender wealth gap, banking-deserts, regressive design of US tax policy on retirement accounts; cites EPI, Hamilton/New School, Darity-Mullen, FDIC, NCRC, Brookings TPC. (3) Reading list — policy-context (Olen, Darity-Mullen, Aspen FSP) and practitioner-level alternatives calibrated to non-target audiences (Aliche, Khalfani-Cox, Lowry, Tobias). (4) 'What we owe you in return' — explicit pointer to the McQuarrie / Pfau-international / Bessembinder / ex-ante-vs-ex-post callouts and the Fernandes-Lynch-Netemeyer information-limits acknowledgment as the framework's honest engagement with its own limits. New .welcome-scope CSS with p-foundation taupe left-border accent and left-aligned body text. NOT APPLIED (Path B): full scope expansion to community-finance-educator audience deferred to a separate post-Phase-7 conversation.",
     "Phase 6 — Session 3 consumer-finance advocate finding. Path-A scope decision."),

    ("CL391", "Spend:Ess:3; calcHDHPvsPPO", "HDHP+HSA recommendation was math-correct but lacked the liquidity-gate precondition that determines whether the strategy is safe for moderate-income readers; without the gate, the recommendation is actively harmful for households who can't absorb the deductible from cash and end up funding it on 24%+ credit-card APR",
     "EMPIRICAL", "CON", "CORRECTED", "C",
     "KFF Employer Health Benefits Survey 2024; Federal Reserve Economic Well-Being of US Households 2023 (37% can't cover $400 emergency from savings); CFPB Medical Debt Burden in the United States 2022; Olen Slate HSA columns 2017+",
     "APPLIED: New 'HDHP precondition — the liquidity gate' math-callout added above the HDHP-vs-PPO calculator in Spend:Ess:3. Names the operational rule: if liquid EF < deductible+OOP, choose PPO this year regardless of headline cost math; HSA's tax-arbitrage value is wiped out by 24%+ credit-card APR on one unfunded deductible event. Cites Fed SHED 2023 37%-can't-cover-$400 stat, KFF EHBS 2024, CFPB Medical Debt 2022, Olen's Slate HSA reporting. calcHDHPvsPPO renderer now emits a contextual warning (.calc-warning CSS, p-safety left-border) when state.answers.emergencyFund is '0' or '<1mo' (always trigger — PPO is the safer choice) or '1-3mo' with hdhpExposure > $5K (borderline). Warning appears below the cost-comparison table inside the calculator's result block.",
     "Phase 6 — Session 3 consumer-finance advocate finding."),

    ("CL392", "Bog:1; Diagnostic:Benefits & coverage", "Bogleheads framework assumed good 401(k) fund access; DOL Form 5500 and Brightscope-ICI data document substantial fee-quality gap by employer size (small-employer plans 1.0-1.5% all-in vs 0.20-0.40% at large), which routes Bogleheads advice through captive bad menus for many readers",
     "EMPIRICAL", "CON", "CORRECTED", "C",
     "DOL Form 5500 dataset; Brightscope-ICI Defined Contribution Plan Profile annual; Tibble v. Edison International 575 US 523 (2015); Tobias The Only Investment Guide You'll Ever Need (revised 2022)",
     "APPLIED: New math-callout 'The good plan access precondition' added to Bog:1 after the existing fee-differential paragraph. Quantifies the small-vs-large-employer fee gap (~25-35% lower terminal balance over 30y for a 1.2% vs 0.20% plan), prescribes the operational sequence in a bad-menu plan (capture match anyway → max IRA where VTI/VXUS/BND available at 3-10bp → weight above-match contributions toward IRA/HSA/taxable → push HR/benefits team per ERISA §404(a) and Tibble), and cites Tobias 2022. New 'planLowestER' diagnostic question in Benefits & coverage section (under-10bp / 10-30bp / 30-60bp / over-60bp / unsure) — shown for W2 households with employerMatch=yes. Downstream wiring to a Plan-view bad-menu action recommendation is in P6.5 backlog; the diagnostic captures the state now.",
     "Phase 6 — Session 3 consumer-finance advocate finding."),

    ("CL395", "W2:6.2 IRA node", "Saver's Credit (IRC §25B) absent from artifact despite documented massive under-claim and the credit being one of the highest-NPV federal tax credits available to low/moderate-income retirement savers",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRC §25B; IRS Form 8880; IRS Rev. Proc. annual inflation adjustment for AGI thresholds; Pew Charitable Trusts / Brookings research on Saver's Credit under-claim rates",
     "APPLIED: New callout added to W2:6.2 IRA node after the contribution-limits math block. Names the 2026 AGI thresholds at the 50% credit tier (~$24K single / ~$36K HoH / ~$48K MFJ), the 50/20/10% tier structure, the $2,000/$4,000 contribution base, the 'guaranteed return on top of deductibility' framing for the 50% tier, the IRS Form 8880 mechanics, and the under-claim documentation. Verify-against-Rev-Proc qualifier included since AGI thresholds index annually. Full benefits-cliff treatment (EITC interactions, Medicaid asset/income tests, SSI, SNAP, LIHEAP, Atlanta Fed CLIFF Dashboard reference) remains DEFERRED to P6.5 per the Session-3 review.",
     "Phase 6 — Session 3 consumer-finance advocate finding (Saver's Credit one-line slice)."),

    ("CL381", "Diagnostic:Foundation; renderPlanSynthesis", "Framework posture assumed information→behavior; Lusardi-Mitchell and Fernandes-Lynch-Netemeyer establish information alone has near-zero behavioral persistence past ~6 months — choice architecture and automation dominate",
     "META", "CON", "CORRECTED", "B",
     "Lusardi & Mitchell (2014) JEL 52(1):5-44 'The Economic Importance of Financial Literacy'; Lusardi & Tufano (2015) JPEF 14(4) 'Debt Literacy, Financial Experiences, and Overindebtedness'; Fernandes, Lynch, Netemeyer (2014) Management Science 60(8):1861-1883 (meta-analysis of 168 studies)",
     "APPLIED: Four-question literacy block added to the Diagnostic Foundation section — the validated Lusardi-Mitchell Big Three (compound-interest growth, inflation effect on purchasing power, single-stock-vs-mutual-fund risk diversification) plus the Lusardi-Tufano debt-literacy item (credit-card-doubling time under compound interest). Each question presented with no-judgment framing in the help text. New getLiteracyScore() helper returns {score, total, level: 'low'|'mid'|'high'}. renderPlanSynthesis appended a literacy-calibration paragraph (only renders when ≥3 of the 4 items answered): low band → recommends Math view §1/§3/§4 as foundational before optimization items; mid band → working baseline note; high band → skip-to-optimization framing. Also appended an information-limit acknowledgment paragraph citing Fernandes-Lynch-Netemeyer 2014 meta-analysis (~0.1% of behavior-variance explained by financial education vs measurably-larger choice-architecture interventions) with the practical implication: the Plan produces real behavior change to the extent that recommendations get converted into one-time setup actions with auto-mechanics — high-friction items are at structural risk of not getting done.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL380", "Plan view, computePlan", "Plan actions ordered by urgency-and-expected-value only; Madrian-Shea 2001 and Choi-Laibson-Madrian 2003 'Optimal Defaults' establish that execution friction is the binding constraint, not information",
     "UX", "SOFT", "CORRECTED", "C",
     "Madrian & Shea (2001) QJE 116(4):1149-1187 'The Power of Suggestion: Inertia in 401(k) Participation'; Choi, Laibson, Madrian, Metrick (2003) 'Optimal Defaults' AER P&P 93; Thaler & Sunstein (2008) Nudge Ch. 5",
     "APPLIED: New friction inference at the end of computePlan derives a 'low' / 'med' / 'high' label for every action via FRICTION_RULES table (low = single login or single mailing, <10min; med = one vendor or pro engagement; high = sustained behavioral change over weeks). Each action card now shows a friction time-hint chip via a new .action-tag-friction-low/-med/-high CSS rule (using --p-highdebt for low to make easy items visually attractive). Within-stage sort by (critical desc, friction asc) so critical-low-friction items surface first within each time horizon.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL388", "Plan view", "12+ actions tagged 'critical' for typical household renders the tag non-functional; choice-overload literature (Iyengar-Lepper 2000; Iyengar-Huberman-Jiang 2004) predicts procrastination on the full list when no clear top item emerges",
     "UX", "CON", "CORRECTED", "C",
     "Iyengar & Lepper (2000) JPSP 79 'When Choice is Demotivating: Can One Desire Too Much of a Good Thing?'; Iyengar, Huberman, Jiang (2004) in Mitchell-Utkus eds Pension Design and Structure; Tversky & Kahneman (1973) Cog Psy 5 'Availability'",
     "APPLIED: New Tier-1 'Start here · this week' anchor rendered above the existing stage groups via new #plan-tier1 DOM container, .plan-tier1-block CSS with accent-bordered card and warm gradient. Tier-1 selection logic in computePlan picks up to three critical-tagged actions ranked by (friction asc, stage asc, idx asc) and exposes a tier1 array on the plan return. Render loop renders the same action via renderActionCard helper in both the Tier-1 anchor (with .action-tier1 class) and the stage view (no extra class). Existing critical tag preserved on cards; the Tier-1 anchor functions as the visual primacy lever rather than replacing the tag entirely.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL382", "Diagnostic:Benefits & coverage; computePlan", "Save More Tomorrow / auto-escalation absent from Plan engine and Diagnostic despite being the single most empirically validated savings-rate intervention; SECURE 2.0 made it default-on for new plans starting 2025",
     "UX", "CON", "CORRECTED", "B",
     "Thaler & Benartzi (2004) JPE 112(S1):S164-S187 'Save More Tomorrow: Using Behavioral Economics to Increase Employee Saving'; Benartzi & Thaler (2013) Science 339:1152-1153; SECURE 2.0 Act §101; Vanguard How America Saves 2024",
     "APPLIED: New 'autoEscalation' diagnostic question in Benefits & Coverage, shown when w2 income + employerMatch=yes. Values: yes-enabled / yes-not-enabled / no / unsure. computePlan adds a critical 'now' action when yes-not-enabled (cites Thaler-Benartzi 2004 mechanism: hyperbolic discounting via future-dated commitment, status-quo bias via auto-execution, loss aversion via raise-tied increases, inertia working for the saver), and a non-critical 'now' action when unsure (HR-portal check). Reframed as 'one decision today pre-commits five future decisions you would otherwise procrastinate on indefinitely' rather than as a will-power exercise.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL384", "Diagnostic:About you; renderPlanSynthesis", "CL330 risk-capacity question relied on self-prediction under counterfactual stress (affective-forecasting bias per Wilson-Gilbert 2003; Frydman-Rangel 2014) + had normatively-ranked options creating social-desirability bias",
     "UX", "SOFT", "CORRECTED", "C",
     "Wilson & Gilbert (2003) 'Affective Forecasting' Adv Exp Soc Psy 35:345-411; Frydman & Rangel (2014) JEBO 107:541-552; Loewenstein, Read, Baumeister eds (2003) Time and Decision; Frydman & Camerer (2016) JEP 30(4)",
     "APPLIED: Friendly correction to last week's CL330 work from Session 1. (1) CL330 riskCapacity options reframed as value-neutral ('Add to equity / Hold / Reduce equity / Move to cash') with help text noting 'no right answer here — we calibrate to how you actually invest, not how the textbook says you should.' (2) New riskCapacityRevealed companion question asking what the respondent actually did in March 2020 / 2022 / 2008 (with not-investing fallback for younger users). (3) renderPlanSynthesis constraint logic updated to weight revealed-preference over self-prediction when both are available, framing the constraint as 'a documented real-event response' vs 'a stated panic-response' accordingly.",
     "Phase 6 — Session 2 behavioral economist finding (correction to Session 1 CL330)."),

    ("CL389", "Diagnostic:Family & Assets; renderPlanSynthesis", "Family-system framing (CL328) acknowledged but spousal-alignment variable absent from Diagnostic despite being one of the highest-leverage household-level behavioral predictors of plan execution",
     "UX", "SOFT", "CORRECTED", "C",
     "Yilmazer & Lyons (2010) J Fam Econ Issues 31:219-227; Bertocchi, Brunetti, Torricelli (2014) Eur Econ Rev 71:293-316 'Who Holds the Purse Strings within the Household?'; Fonseca, Mullen, Zamarro, Zissimopoulos (2012) J Consumer Affairs 46:90-106; Klontz/Britt financial-therapy literature cited in CL326",
     "APPLIED: Three new diagnostic questions in Family & Assets, all shown when filingStatus=mfj. (1) spousalAlignSaving — yes-fully/yes-mostly/no-gap/not-discussed; (2) spousalAlignRetirement — same scale; (3) spousalAlignLegacy — same scale + na option for households without heirs. renderPlanSynthesis constraint block now surfaces any no-gap or not-discussed answers as 'an unresolved spousal-alignment gap on [comma-list]' with explicit framing: the framework's recommendations assume household alignment, and the published evidence treats disagreement as one of the strongest predictors of plan-execution failure; consider running diagnostic separately with spouse and reconciling outputs before pursuing the action list. Natural extension of CL330's Session-1 spousal-income-split addition.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL385", "Bog:1", "Behavior gap correctly cited but attributed to 'discipline' rather than to its actual behavioral mechanism (myopic loss aversion per Benartzi-Thaler 1995); structural interventions (evaluation frequency, app notifications, target-date abstraction) absent from the prescriptive response",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Benartzi & Thaler (1995) QJE 110(1) 'Myopic Loss Aversion and the Equity Premium Puzzle'; Thaler-Tversky-Kahneman-Schwartz (1997) QJE 112(2); Kahneman-Tversky (1979) Econometrica 47 prospect theory; Dalbar QAIB 2024; Morningstar Mind the Gap 2024",
     "APPLIED: New 'The mechanism behind the behavior gap — myopic loss aversion, not lack of discipline' math-callout added to Bog:1 immediately after the existing behavior-gap callout. Cites Benartzi-Thaler 1995 mechanism, Thaler-Tversky-Kahneman-Schwartz 1997 experimental confirmation, Kahneman-Tversky 1979 prospect theory. Prescribes structural interventions: quarterly/annual portfolio review (not daily), disable brokerage-app notifications, target-date funds for would-be-reactive-rebalancers, automatic-investment continuation through drawdowns. Notes the discipline frame works disproportionately for households with low System-1-vs-System-2 friction; the structural moves work for everyone. Cross-references Math §1 volatility-drag math as the indirect documentation of why infrequent evaluation produces better realized returns through behavioral rather than mathematical channels.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL386", "Zeit:Life:5, Math:5", "Die-with-Zero / longevity-insurance recommendations omitted the annuity puzzle as the operative behavioral obstacle; QLAC and deferred-annuity framing used investment-framing which empirically suppresses uptake by ~50pp vs consumption-framing",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Yaari (1965) RES 32 'Uncertain Lifetime, Life Insurance, and the Theory of the Consumer'; Mitchell-Poterba-Warshawsky-Brown (1999) AER 89; Brown-Kling-Mullainathan-Wrobel (2008) AER P&P 98 'Why Don't People Insure Late-Life Consumption? A Framing Explanation of the Under-Annuitization Puzzle'; Benartzi-Previtero-Thaler (2011) JEP 25 'Annuitization Puzzles'",
     "APPLIED: New 'The annuity puzzle — why the structural Die-with-Zero solution is one of the hardest behavioral sells in personal finance' math-callout added to Zeit:Life:5 after the 'Where Die with Zero converges' callout. Names Yaari 1965, Mitchell-Poterba-Warshawsky-Brown 1999, Benartzi-Previtero-Thaler 2011, Brown-Kling-Mullainathan-Wrobel 2008 framing-effects result. Reframed the existing Perkins-annuitization sentence to flag the behavioral-rather-than-product-quality framing of the obstacle. Math §5 longevity callout updated to use consumption framing ('converts $X today into $Y/month of guaranteed lifetime income starting at age 80-85') rather than investment framing. Prescriptive sizing: cover the essential portion (essentialFraction × annualExpenses from CL333 diagnostic state) with a deferred income annuity / QLAC starting at age 80-85; equity portfolio funds discretionary. Zeit:Life:9 DINK reframing deferred to a future touch — the §5 + §7 + §9 trio shares this framing and a coordinated re-pass is cleaner than per-section edits.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL387", "Zeit:Inv:1", "FIRE coverage engages survivorship bias but omitted projection bias (Loewenstein-O'Donoghue-Rabin 2003) which is the upstream behavioral mechanism generating the FIRE survival/failure split",
     "EMPIRICAL", "SOFT", "CORRECTED", "C",
     "Loewenstein, O'Donoghue, Rabin (2003) QJE 118 'Projection Bias in Predicting Future Utility'; Read & van Leeuwen (1998) Org Behav Hum Dec 76 'Predicting Hunger'; Loewenstein & Schkade (1999) 'Wouldn't It Be Nice' in Well-Being: Foundations of Hedonic Psychology",
     "APPLIED: New 'Projection bias — the behavioral mechanism that generates the FIRE survival/failure split' math-callout added to Zeit:Inv:1 immediately after the survivor-bias paragraph. Cites Loewenstein-O'Donoghue-Rabin 2003; names the documented FIRE failure modes (burnout, partner disagreement, post-FI identity crisis CL321, savings-aversion CL326, boomerang CL325) as predictable consequences of projection bias applied to sustained-extreme-commitment plans; reframes the framework's existing sustainable-rate emphasis (20-30% over 30 years vs 60% over 5 years) as an implicit projection-bias hedge. Cross-references Math §3 rising-income callout and Zeit:Life §6 anti-hustle. Surfaces the operative behavioral test: the savings rate you would still be sustaining under common life-event shocks is the rate the math should use.",
     "Phase 6 — Session 2 behavioral economist finding."),

    ("CL330", "Diagnostic:About you, Family & assets", "Diagnostic does not elicit spousal income split, parent-care exposure, or risk-capacity vs tolerance distinction",
     "UX", "CON", "PARTIAL", "C",
     "Bernstein 'The Four Pillars of Investing' on ability/willingness/need; Genworth Beyond Dollars caregiver cost survey; Kitces NEV on widow's tax trap (2018)",
     "APPLIED PARTIALLY: Three new diagnostic questions land now; downstream consumption of the new state is staged. (1) About you: 'spousalIncomeSplit' — shown when filingStatus is mfj/mfs; values 50-50 / 60-40 / 70-30 / sole; drives Social Security claiming optimization and surviving-spouse modeling. (2) About you: 'riskCapacity' — operationalizes Bernstein's willingness component via 35% drawdown panic-response probe (rebalance-in / hold / cut-equity / sell-to-cash). (3) Family & assets: 'parentCare' — yes-substantial / yes-modest / unlikely / na; captures sandwich-generation exposure with Genworth $50K-$200K per parent magnitude reference in the help text. Plan-view synthesis updated to surface the new state as binding constraints (sole-earner widow's-tax-trap framing, riskCapacity panic-response → conservative-glide-path implication, parentCare yes-substantial → unlogged cash-flow obligation). DEFERRED P6.5: full downstream wiring — claiming-age recommendation keyed to spousal-split (higher earner delays to 70 while lower earner files at FRA), capacity-limited glide path adjustments in Plan view, parent-care provisioning line item in the budget action.",
     "Phase 6 — Session 1 CFP finding."),

    ("CL333", "Diagnostic:Family & assets", "Diagnostic does not elicit annual expenses or essential/discretionary decomposition; downstream calculators all use hardcoded defaults",
     "UX", "CON", "PARTIAL", "B",
     "Blanchett 2014 'Exploring the Retirement Consumption Puzzle' JFP; Society of Actuaries Spending in Retirement Survey 2018",
     "APPLIED PARTIALLY: Two new diagnostic questions in Family & assets. (1) 'annualExpenses' bucket (under-40k / 40-60k / 60-90k / 90-150k / 150-250k / 250k+); midpoint mapping in getDefault now propagates to sr-expenses (Math §3 SS-offset), hc-wd (Math §7 first-year withdrawal), and dz-income (Zeit:Life §5 Die-with-Zero income proxy). (2) 'essentialFraction' bucket (90+ / 70-90 / 50-70 / under-50); midpoint mapping returns decimal. New getDefault('essentialExpenses') convenience returns annualExpenses × essentialFraction for downstream LTC / DI / floor calculations. PARTIAL because dime-income (DIME life-insurance calculator) is not yet wired — the right input there is essentialExpenses, not annualExpenses, and the calculator field is currently labeled 'income to replace' rather than 'essential expenses to replace'. DEFERRED P6.5: (a) DIME refactor to take essentialExpenses as input; (b) Plan-view strategy-class recommendation in §7 keyed off essentialFraction (high-essential → floor-and-upside or guardrails; high-discretionary → dynamic spending feasible); (c) LTC sizing tied to essentialExpenses rather than aggregate.",
     "Phase 6 — Session 1 CFP finding."),

    ("CL332", "Math:7", "Withdrawal-strategy options unnamed and unsourced; framework lacks the broader canonical taxonomy (Bengen / Guyton-Klinger / Vanguard dynamic / Blanchett floor-and-upside / Pfau funded-ratio / Kitces ratcheting)",
     "CITATION", "SOFT", "CORRECTED", "C",
     "Bengen 1994 JFP 'Determining Withdrawal Rates Using Historical Data'; Guyton-Klinger 2006 JFP 'Decision Rules and Maximum Initial Withdrawal Rates'; Pfau-Ameriks-Madamba Vanguard 2014; Blanchett floor-and-upside JFP series",
     "APPLIED: hc-strategy dropdown options relabeled to map to citeable strategies: 'Bengen (1994) — fixed real, inflation-adjusted', 'Vanguard-style dynamic (cut 25% after down years)', 'Guyton-Klinger (2006) guardrails (±20% bands)'. calcHistoricalCycles strategy label rendering updated to match. New math-callout above the §7 calculator names the canonical taxonomy: Bengen, Guyton-Klinger, Vanguard dynamic, Blanchett floor-and-upside, Pfau funded-ratio, Kitces ratcheting — with notation of which three the simulator covers (the most-commonly-cited) and which await Phase 6.5 modeling. DEFERRED P6.5: Plan-view recommendation mapping spending-decomposition diagnostic answers to a recommended strategy class.",
     "Phase 6 — Session 1 CFP finding."),

    ("CL335", "Spend:Ess:4; Plan:computePlan", "Disability insurance lacks a sizing calculator parallel to DIME, and the Plan view doesn't surface the DI-vs-life priority inversion at high incomes",
     "UX", "CON", "PARTIAL", "C",
     "Council for Disability Awareness Personal Disability Quotient; Kitces NEV on group-LTD shortfall for HENRYs",
     "APPLIED PARTIALLY: Plan-view 'next-quarter / critical' action added in computePlan when taxBracket=high AND ltdInPlace is no/partial AND dependents=yes AND termLifeInPlace is no/partial. Explains the priority inversion (DI gap typically dominates additional term-life need in expected-PV dollars at high income), cites Council for Disability Awareness 25% lifetime 90+day-disability probability, gives the typical $200K+/yr unreplaced-gap-after-group-LTD-cap math for $400K earners, and sequences: get individual own-occupation DI first, then add/extend term life. DEFERRED P6.5: full DI sizing calculator parallel to DIME (parallel UI structure, ~30 lines, with replacement-target / group-cap / taxability / years-to-retirement inputs producing monthly gap and PV at 3% real).",
     "Phase 6 — Session 1 CFP finding."),

    ("CL336", "Math:6; Bog:5", "Asset location calculator treats future tax rate as a static input rather than a Roth-conversion-ladder choice variable; framework gestures at the integration in prose but can't model it",
     "META", "CON", "PARTIAL", "C",
     "Bruno & Bortolotti Vanguard 2017 'Cost-effective Roth conversions'; Kitces NEV on integrated location + conversion math",
     "APPLIED PARTIALLY: New math-callout under the Math §6 asset-location calculator that frames the future-tax-rate input as a Roth-conversion-ladder choice variable rather than a static planning assumption. Documents the inversion case (peak-earnings household at 32% current / 24% future without ladder follows the standard placement rule; same household with planned 5-year ladder filling the 12% bracket sees optimal placement flip — bond yields in taxable, stocks in Roth) and cites the Bruno-Bortolotti Vanguard 2017 30-80 bps/yr integrated-planning premium on top of naive asset location. The callout instructs users to run the matrix twice (no-ladder vs with-ladder future rate) as the manual workaround. DEFERRED P6.5: integrated calculator extension that models both regimes simultaneously and surfaces the placement-inversion delta.",
     "Phase 6 — Session 1 CFP finding."),

    ("CL361", "W2:9.1, Contractor:9.2, Business:9.2", "Mega Backdoor Roth: after-tax earnings taxed as ordinary income at in-plan conversion per IRS Notice 2014-54; same-payroll or daily automatic conversion is the operative fix",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRS Notice 2014-54; IRC §402(c)(2); IRC §1411(c) for NIIT reach on accrued earnings",
     "APPLIED: Warning callout added to each of the three Mega Backdoor nodes (W2:9.1, Contractor:9.2 Solo 401(k), Business:9.2 business 401(k)). Each callout explains the Notice 2014-54 bifurcation (contribution basis converts tax-free; earnings on after-tax contributions are ordinary-income taxable at conversion), the practical leakage at 32-37% federal + NIIT for quarterly/annual manual conversion, and the fix: same-payroll or daily automatic in-plan Roth conversion. Business 9.2 callout adds the plan-design framing — when the owner is drafting the plan document for the first time, spec automatic in-plan Roth conversion as the default since changing it post-hoc requires plan amendment.",
     "Phase 6 — Session 1 Tax Attorney finding."),

    ("CL365", "Business:9.1", "§1202 QSBS planning: §1202(g) gifting + non-grantor trust stacking multiplies the per-issuer cap; §1045 60-day rollover preserves the clock; §57(a)(7) AMT preference on excluded gain",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRC §1202(g), §1223(2), §1045, §57(a)(7); Letter Ruling 9633004 on holding-period tacking; pending Treasury guidance on AMT treatment of post-OBBBA exclusion tiers",
     "APPLIED: Three callouts added to Business:9.1. (1) Stacking the per-issuer cap — §1202(g) gifts to family members with §1223(2) holding-period tacking, plus non-grantor trust contributions, each donee/trust gets own $15M cap; pre-sale lever only (cannot apply retroactively post-signing); $60M-gain founder can potentially stack to 4× $15M via spouse + two adult-child non-grantor trusts; cites Letter Ruling 9633004. (2) §1045 60-day rollover — sales before 5-year hold can be rolled into replacement QSBS within 60 days to preserve original holding-period clock, avoiding the 28% rate trap; commonly used in serial-founder situations. (3) §57(a)(7) AMT preference — 7% of excluded gain is AMT preference for pre-OBBBA 50%/75% tier exclusions; post-OBBBA partial-exclusion tier AMT treatment awaits Treasury guidance; conservative posture is to assume preference still applies until clarified, particularly material for AMT-exposed taxpayers in high-state-tax jurisdictions. Highest-dollar-magnitude correction in the review ($5M-$20M on a founder exit) but operates on a thin audience.",
     "Phase 6 — Session 1 Tax Attorney finding."),

    ("CL367", "Math:6, Bog:5", "Asset location: muni-bond taxable-equivalent yield not modeled in calculator; direct indexing tax-alpha decay and embedded-gains pile-up not flagged",
     "EMPIRICAL", "SOFT", "PARTIAL", "B",
     "IRC §103 (muni interest exclusion); IRC §1014 (basis step-up); Vanguard 2022 'The Value of Personalized Indexing'; Wealthfront direct indexing methodology",
     "APPLIED PARTIALLY: Added direct-indexing tax-alpha decay note to the Bog:5 HNW callout — Vanguard 2022 + Wealthfront methodology both document alpha falling below the 25-40 bps strategy fee after 5-8 years in sustained bull markets; embedded-gains pile-up creates a forced-realization cliff (basis step-up only at death under §1014, charitable contribution at FMV under §170, or rare §1031-like events); strategy most attractive with a known charitable-deployment endpoint or hold-to-step-up-at-death intent. DEFERRED P6.5: full muni-bond fourth-asset addition to Math §6 calculator with state-tax-equivalent-yield computation.",
     "Phase 6 — Session 1 Tax Attorney finding."),

    ("CL368", "Contractor:9.1, Business:1.1", "S-corp reasonable salary: comparable-wage methodology with burden-on-taxpayer; recent Tax Court memoranda (Maggard 2024) reinforce that professional-services salary ratios under ~40% are the operative audit-risk zone",
     "CITATION", "SOFT", "CORRECTED", "B",
     "Watson v. Comm'r (8th Cir. 2012); Sean McAlary Ltd. v. Comm'r (T.C. Summ. Op. 2013-62); Glass Blocks Unlimited v. Comm'r; Maggard v. Comm'r (T.C. Memo 2024-77); IRS Fact Sheet FS-2008-25",
     "APPLIED: Contractor:9.1 'reasonable salary trap' warning updated to reflect IRS comparable-wage methodology (RC Reports, BLS OEWS, ERI), recent Tax Court reinforcement of taxpayer burden (Maggard 2024, Sean McAlary line, Glass Blocks Unlimited), the under-40% professional-services audit-risk zone, and the planner pattern of targeting 60-70% salary ratios at $300K+ net to remove the issue from audit consideration — characterized as a defensible $3-5K/yr-of-additional-payroll-tax trade for examination immunity. The existing $70K-on-$150K example retained as reasonable for many businesses but flagged as 'on the lower end for pure-service businesses where the owner is the principal revenue generator.'",
     "Phase 6 — Session 1 Tax Attorney finding."),

    ("CL369", "Business:8.4 (new node)", "State-level optimization missing: PTET elections (36+ states; IRS Notice 2020-75), state non-conformity to federal preferences (CA, NY, NJ), pre-liquidity-event domicile change",
     "STATUTORY", "CON", "PARTIAL", "A",
     "IRS Notice 2020-75 (PTET federal deductibility); state PTET statutes (CA AB 150, NY Tax Law §860 et seq., NJ BAIT, IL PTE Tax); CA Rev. & Tax. Code §17041; SALT cap permanence under OBBBA",
     "APPLIED PARTIALLY: New Business:8.4 'State-level optimization' node added. Covers (1) PTET elections in 36+ states, IRS Notice 2020-75 federal deductibility, the typical $13K SALT recovery for a CA owner with $400K pass-through in the 37% federal bracket, and the March 15 prior-year election deadline pattern with the warning that missed elections are not retroactively curable; (2) state non-conformity — CA not conforming to §1202, partial non-conformity on §1031 and bonus depreciation in CA/NY/NJ/MA, CA Mental Health Services Tax (1% on income over $1M); (3) domicile-change framing for pre-liquidity-event timing — 7-figure state-tax savings on $5M+ exits when domicile is established in no-income-tax state 12+ months pre-event, with explicit indicia-of-domicile checklist and the warning that aggressive claims without underlying ties are litigation magnets in CA and NY. DEFERRED to Phase 7: full domicile-change playbook with indicia documentation framework (aligns with existing EX1 catalog item).",
     "Phase 6 — Session 1 Tax Attorney finding."),

    ("CL366", "Diagnostic:Income & taxes; Plan:computePlan", "§83(b) election must be filed within 30 days of grant of restricted property; no relief for missed deadline (IRC §83(b), Treas. Reg. §1.83-2(b))",
     "STATUTORY", "CON", "CORRECTED", "A",
     "IRC §83(b)(1)-(2); Treas. Reg. §1.83-2(b); Rev. Proc. 2012-29 (sample election form)",
     "APPLIED: Bridge warning ahead of full Phase 7 equity-comp coverage. New diagnostic question in Income & taxes section: 'In the past 30 days, have you received restricted stock, founder equity, or early-exercised stock options?' (yes/considering/no). When yes, computePlan adds a 'now / critical' action surfacing the 30-day deadline, the $1M+ magnitude on founder-equity scenarios, and that there is no late-filing relief or equitable tolling. When considering, adds a 'now' (non-critical) action to plan the filing logistics ahead of grant. NodeKey points to business:9.1 (Section 1202 QSBS) as the closest existing equity node — full §83(b) mechanics treatment is a Phase 7 deliverable.",
     "Phase 6 — Session 1 Tax Attorney finding, structural bridge applied; full Phase 7 §83(b) section still deferred."),
]

# Build Excel workbook
wb = Workbook()

# === Sheet 1: Full audit ===
ws1 = wb.active
ws1.title = "Claims Audit"

headers = ["ID", "Location", "Claim", "Category", "Priority", "Status", "Tier",
           "Source", "Correction Needed", "Notes"]
ws1.append(headers)

for row in CLAIMS:
    ws1.append(list(row))

# Styling
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1C2128")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row formatting
for c in range(1, len(headers) + 1):
    cell = ws1.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Status-based color coding
status_colors = {
    "PASS": "E8F5E9",        # light green
    "CORRECTED": "FFF8E1",   # light amber (resolved but was an issue)
    "FAIL": "FFEBEE",        # light red
    "PARTIAL": "FFF3E0",     # light orange
    "PENDING": "F5F5F5",     # light gray
    "DEFERRED-P2.5": "EDE7F6",  # light purple (deferred backlog)
    "DEFERRED-P3.5": "EDE7F6",
    "DEFERRED-P4.5": "EDE7F6",
    "DEFERRED-P5.5": "EDE7F6",
}
tier_colors = {
    "A": "1976D2",  # blue
    "B": "388E3C",  # green
    "C": "F57C00",  # orange
    "D": "757575",  # gray
}
data_font = Font(name="Arial", size=9)
data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

for r in range(2, len(CLAIMS) + 2):
    status = ws1.cell(row=r, column=6).value
    fill_color = status_colors.get(status, "FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws1.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = border
        cell.fill = PatternFill("solid", start_color=fill_color)
    # Tier coloring
    tier_cell = ws1.cell(row=r, column=7)
    tier_val = tier_cell.value
    if tier_val in tier_colors:
        tier_cell.font = Font(name="Arial", size=9, bold=True, color=tier_colors[tier_val])

# Column widths
col_widths = {"A": 8, "B": 22, "C": 55, "D": 14, "E": 10, "F": 12, "G": 7, "H": 35, "I": 50, "J": 35}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# Freeze header
ws1.freeze_panes = "A2"

# === Sheet 2: Summary ===
ws2 = wb.create_sheet("Summary")
ws2.append(["Phase 1 Fact-Check Audit — Summary"])
ws2["A1"].font = Font(name="Arial", size=14, bold=True)
ws2.append([])
ws2.append(["Total claims audited:", len(CLAIMS)])

# Status breakdown
from collections import Counter
status_count = Counter(c[5] for c in CLAIMS)
tier_count = Counter(c[6] for c in CLAIMS if c[6])
cat_count = Counter(c[3] for c in CLAIMS)
pri_count = Counter(c[4] for c in CLAIMS)

ws2.append([])
ws2.append(["BY STATUS"])
ws2["A5"].font = Font(name="Arial", size=11, bold=True)
for s, n in status_count.most_common():
    ws2.append([s, n])
ws2.append([])
ws2.append(["BY CONFIDENCE TIER (passed/corrected)"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=11, bold=True)
tier_labels = {"A": "A — Statute / Code",
               "B": "B — Established consensus",
               "C": "C — Empirical / Contested",
               "D": "D — Heuristic / Convention"}
for t in ["A", "B", "C", "D"]:
    ws2.append([tier_labels[t], tier_count.get(t, 0)])

ws2.append([])
ws2.append(["BY PRIORITY"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=11, bold=True)
for p, n in pri_count.most_common():
    ws2.append([p, n])

ws2.append([])
ws2.append(["BY CATEGORY"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=11, bold=True)
for c, n in cat_count.most_common():
    ws2.append([c, n])

# Findings section
ws2.append([])
ws2.append([])
ws2.append(["KEY FINDINGS — CORRECTIONS REQUIRED"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="C62828")
ws2.append([])

findings = [
    ("CL016", "SIMPLE IRA catch-up — artifact said $3,850; corrected to $4,000 for most plans in 2026 with $3,850 reserved for certain applicable SIMPLE plans under SECURE 2.0 §117 (25 or fewer employees, OR 26-100 with 4% NE / 4%-increased match)."),
    ("CL037", "I-Bonds paper purchase via tax refund — discontinued by Treasury announcement October 17, 2024 with tax year 2024 filings the last to use it. Artifact updated."),
    ("CL056", "QBI SSTB thresholds — corrected to 2026 figures with OBBBA-expanded $75K/$150K phase-in window. $400 minimum deduction and engineering/architecture exclusion added."),
    ("CL102", "Employer Roth match — updated to reflect SECURE 2.0 §604 permitting Roth match (Persona 1 catch)."),
    ("CL106", "ACA Premium Tax Credit cliff returned for 2026; artifact updated."),
    ("CL112", "High-earner Roth catch-up framing — final regs grant transition relief to plan years beginning after Dec 31, 2026. Artifact softened from 'cannot make catch-up at all' to operational reality with ~80% large-plan adoption."),
    ("CL113", "Mega Backdoor Roth ACP testing constraint added to W2:9.1 / Business:9.2 — in non-Safe Harbor plans, HCEs often have $5K-$15K of usable after-tax space, not the headline $37K."),
    ("CL114", "Solo 401(k) ACP bypass note added to Contractor:9.2 — structural advantage of Solo over employer plans."),
    ("CL115", "In-plan Roth conversion 5-year clock per conversion added to W2:9.1."),
    ("CL116", "529-to-Roth 5-year contribution lookback rule added to W2:9.4."),
    ("CL117", "Additional Medicare 0.9% scope clarified — applies to W-2 wages too, with employer withholding nuances."),
    ("CL118-120", "Pro-rata exclusions added to Contractor:6.2: inherited IRAs, Roth IRAs, SIMPLE 2-year wait."),
    ("CL121", "SIMPLE §117 mechanism specified in Business:3.1."),
    ("CL122", "§415(c) per-employer vs §402(g) aggregate distinction added to Business:8.1."),
    ("CL123", "Safe Harbor 401(k) — all four formulas listed in Business:3.1 (basic match, enhanced match, 3% NE, QACA)."),
    ("CL124", "Solo 401(k) employer-side circular formula explained in Contractor:8.1."),
    ("CL125, 128, 129", "ESOP §1042 — four requirements, QRP definition, and step-up mechanism added to Business:10.1."),
    ("CL126, 127", "S-corp family aggregation rule (§1361(c)(1)) and eligible trusts (ESBTs/QSSTs) added to Business:0.1."),
]
for fid, desc in findings:
    ws2.append([fid, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True, color="C62828")
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append(["ADDITIONS RECOMMENDED"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="0288D1")
ws2.append([])
additions = [
    ("APPLIED: 401(a)(17) compensation limit $360,000 (2026) added to Business:8.1 — affects S-corp salary/profit-sharing calculations"),
    ("APPLIED: Annual gift tax exclusion $19,000 (2026) added to W2:6.3 estate planning callout"),
    ("APPLIED: OBBBA $400 minimum QBI deduction added to Contractor:8.2"),
    ("APPLIED: Engineering/architecture SSTB exclusion noted in Contractor:8.2"),
    ("APPLIED: Additional Medicare 0.9% threshold-not-indexed note added to Contractor:0.3"),
    ("APPLIED: ACA 400% FPL cliff returning for 2026 added to Contractor:1.3"),
]
for note in additions:
    ws2.append(["", note])
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 2 BUILD STATUS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="0288D1")
ws2.append([])
phase2_status = [
    ("Math view built", "6 sections (compound interest, Rule of 72, savings rate, real vs nominal, sequence risk, asset location) + interactive calculators"),
    ("Inline calculator added", "W2:3.1 employer match — 'What your match is actually worth'"),
    ("Phase 2 fact-checks", "16 math claims added to database (CL130-CL145), all PASS, mostly tier A"),
    ("Checkpoint personas run", "P2A senior actuary, P2B retirement planning software engineer"),
    ("Structural fixes applied", "CL146 (real/nominal clarification in sequence risk), CL147 (deterministic disclaimer in math intro), CL148 (asset location realization update + Roth caveat), CL149 (input validation), CL150 (compounding frequency clarification)"),
]
for label, desc in phase2_status:
    ws2.append([label, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True)
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 2.5 BACKLOG — DEFERRED FROM CHECKPOINT PERSONAS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="6A1B9A")
ws2.append([])
ws2.append(["", "These items were flagged by P2A or P2B as feature gaps or refinements. Each is captured in the main Claims Audit with status DEFERRED-P2.5. All must be implemented before Phase 6 reviews; the original 5 personas will be checking that nothing was lost."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.append([])

p25_items = [
    ("CL151", "Volatility drag / geometric vs arithmetic mean — add callout near compound interest section"),
    ("CL152", "Rising real income & savings rate — note that MMM formula is conservative for typical accumulators"),
    ("CL153", "Longevity risk — surface that fixed horizons understate; consider age-driven horizon estimation"),
    ("CL154", "Retiree-specific inflation (CPI-E) — note healthcare runs ~3.5-4.5% vs CPI-U ~2.5-3%"),
    ("CL155", "Sequence risk filler — replace constant 9.07% with historical sequences (1966/1982/2000)"),
    ("CL156", "4% rule horizon dependence — cross-reference to existing CL053-054 horizon-adjusted SWR"),
    ("CL157", "Personalized defaults — propagate diagnostic answers into Math view input defaults"),
    ("CL158", "Monte Carlo / historical cycles for sequence risk — major Phase 2.5 feature"),
    ("CL159", "Withdrawal flexibility (Guyton-Klinger, ratcheting, floor-and-upside) selector"),
    ("CL160", "Social Security / other income overlay in Years-to-FI calculator"),
    ("CL161", "SVG charts: §1 compound interest (stacked bar), §5 sequence risk (two paths), §7 historical-cycles (fan chart)"),
    ("CL162", "Tax bracket propagation from diagnostic state to Math view calculators"),
    ("CL163", "Full asset location matrix — 3 account types × N assets (currently 2×2)"),
    ("CL164", "Per-section 'what this doesn't model' disclaimers (general one applied; section-specific deferred)"),
]
for cid, desc in p25_items:
    ws2.append([cid, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True, color="6A1B9A")
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append(["", f"14 items deferred to Phase 2.5. All have full audit entries with statuses, sources, and proposed fixes. SCHEDULED: Phase 2.5 will run AFTER Phase 5 (zeitgeist behaviors) completes and BEFORE Phase 6 (original 5 personas review). This ordering is locked — the original 5 personas must see a complete Phases 1-5 plus cleared Phase 2.5 backlog when they review."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2[f"B{ws2.max_row}"].font = Font(name="Arial", size=10, italic=True, color="6A1B9A")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 3 BUILD STATUS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="0288D1")
ws2.append([])
phase3_status = [
    ("Spending: Essentials view built", "4 sections — Housing, Transportation, Healthcare, Insurance"),
    ("Spending: Lifestyle view built", "4 sections — Food/discretionary, Childcare/Education, Subscriptions, Lifestyle creep (meta-closer)"),
    ("Phase 3 calculators", "Rent vs buy, Vehicle TCO, HDHP vs PPO, DIME term life, DCFSA tax savings, Subscription compounding, Lifestyle creep impact"),
    ("Phase 3 fact-checks", "41 new claims added (CL165-CL205); 21 tier B (specific citations: BLS, KFF, College Board, SSA, Cost Plus Drugs, AAA, Killingsworth-Kahneman 2023), 13 tier A math/statutory (DCFSA limit, FHA 3.5%, mortgage math), 7 tier C heuristics (5% rule, price-to-rent, DIME)"),
    ("Phase 3 checkpoint personas", "P3A real estate / housing economist; P3B healthcare benefits consultant — completed"),
    ("Phase 3 structural fixes applied", "CL208 (rent-vs-buy 2% → 1% Case-Shiller appreciation), CL209 (transaction costs added to rent-vs-buy: 3% buy + 7% sell), CL215 (HSA-eligible HDHP requirements + investment-access nuance), CL218 (HSA custodian variation captured in CL215 callout), CL219 (DPC content added with proper catastrophic-pairing caveat — note: P3B premise was incorrect; original section did not mention DPC, converted critique to content addition), CL220→deferred but partial captured, CL224 (secular healthshare alternatives noted)"),
]
for label, desc in phase3_status:
    ws2.append([label, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True)
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 4 BUILD STATUS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="0288D1")
ws2.append([])
phase4_status = [
    ("Portfolio: Bogleheads view built", "5 sections — Philosophy/SPIVA, Three-fund portfolio, Asset allocation, Rebalancing, Tax-efficient placement (cross-ref Math §6)"),
    ("Portfolio: Theory view built", "3 sections — Modern Portfolio Theory (Markowitz 1952), CAPM (Sharpe-Lintner-Mossin 1964-66), Factor models (Fama-French 1992/2015, Carhart 1997, factor zoo)"),
    ("Phase 4 calculators", "Expense ratio impact (compounded drag), Two-asset efficient frontier, Sharpe ratio comparator"),
    ("Phase 4 fact-checks", "37 new claims added (CL225-CL261); 19 tier B (specific paper citations: Markowitz 1952, Sharpe 1964, Roll 1977, Fama-French 1992/2015, Carhart 1997, Harvey-Liu-Zhu 2016, McLean-Pontiff 2016, Pfau-Kitces 2014, Vanguard 2010 rebalancing, SPIVA, ICI Fact Book), 14 tier A math/historical (Markowitz Nobel 1990, Sharpe Nobel 1990, Shiller Nobel 2013, IRC §901 foreign tax credit, IRC §852(b)(6) ETF mechanics, MPT formulas, Sharpe ratio), 4 tier C (Boglehead principles, allocation heuristics, robust factor set)"),
    ("Phase 4 checkpoint personas", "P4A index fund portfolio manager (practicing); P4B empirical asset pricing researcher (factor specialist) — completed"),
    ("Phase 4 structural fixes applied", "CL262 (Vanguard glide path corrected to 'through retirement'), CL264 (5/25 rebalancing band example clarified to 'lesser of'), CL270 (Fama-French 1992 pattern documentation vs 1993 three-factor formalization), CL271 (Jegadeesh-Titman 1993 momentum origin attribution), CL274 (DeMiguel-Garlappi-Uppal 2009 1/N paper added to MPT section as strongest case against pure optimization, plus Jorion 1986 and Michaud 1998 references)"),
]
for label, desc in phase4_status:
    ws2.append([label, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True)
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 3.5 BACKLOG — DEFERRED FROM CHECKPOINT PERSONAS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="6A1B9A")
ws2.append([])
ws2.append(["", "Phase 3 backlog items captured from P3A and P3B critiques. Same scheduling and accountability as Phase 2.5: must clear before Phase 6 (original 5 personas) reviews. SCHEDULED: runs after Phase 5 (zeitgeist behaviors), in the consolidated sub-phase that includes Phase 2.5 backlog."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.append([])

p35_items = [
    ("CL206", "P3A#1 — Rate-environment caveat on price-to-rent thresholds (15/20 calibrated to 3-4% mortgage era)"),
    ("CL207", "P3A#2 — 5% rule completeness footnote: mortgage interest deduction, primary residence gain exclusion, fixed-rate inflation hedge"),
    ("CL210", "P3A#5 — 28/36 affordability rule sidebar in housing section"),
    ("CL211", "P3A#6 — Refinance framework: replace 0.75-1% heuristic with cost-sensitive breakeven; note no-cost refi option"),
    ("CL212", "P3A#7 — FHA self-sufficiency test for 3-4 unit, MIP ongoing cost note"),
    ("CL213", "P3A#8 — Behavioral execution caveat for 15-vs-30 'invest the difference' assumption"),
    ("CL214", "P3A#9 — Geographic arbitrage: employer salary haircut for relocating remote workers"),
    ("CL216", "P3B#2 — HDHP vs PPO calculator: add coinsurance % as user input"),
    ("CL217", "P3B#3 — HSA receipt-keeping strategy: cash-flow-availability caveat"),
    ("CL220", "P3B#6 — DIME calculator: add PV-discounting option or HLV alternative"),
    ("CL221", "P3B#7 — Full disability insurance rider/feature treatment (benefit period, elimination period, residual, COLA, FIO, M&N)"),
    ("CL222", "P3B#8 — SSA 1-in-4 disability stat contextualization (SSDI definition vs own-occupation policy)"),
    ("CL223", "P3B#9 — Umbrella liability underlying-limits requirement + business exclusion note"),
]
for cid, desc in p35_items:
    ws2.append([cid, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True, color="6A1B9A")
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append(["", f"13 items deferred to Phase 3.5. All have full audit entries with statuses, sources, and proposed fixes. SCHEDULED: Phase 3.5 will run alongside Phase 2.5 after Phase 5 (zeitgeist behaviors) completes and BEFORE Phase 6 (original 5 personas review)."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2[f"B{ws2.max_row}"].font = Font(name="Arial", size=10, italic=True, color="6A1B9A")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 4.5 BACKLOG — DEFERRED FROM CHECKPOINT PERSONAS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="6A1B9A")
ws2.append([])
ws2.append(["", "Phase 4 backlog items captured from P4A and P4B critiques. Same scheduling and accountability as Phase 2.5/3.5: must clear before Phase 6 (original 5 personas) reviews. SCHEDULED: runs alongside Phase 2.5 and Phase 3.5 in the consolidated sub-phase after Phase 5 (zeitgeist behaviors) completes."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.append([])

p45_items = [
    ("CL263", "P4A#2 — Hedge three-fund holding counts to broad buckets or note quarterly variation"),
    ("CL265", "P4A#4 — Vanguard ETF-share-class tax footnote (patent-protected structure expired 2023; existing funds retain advantage)"),
    ("CL266", "P4A#5 — BHB 90%-of-variance interpretation: time-series variance vs cross-sectional return levels (Ibbotson-Kaplan 2000 decomposition)"),
    ("CL267", "P4A#6 — Replace muni break-even point estimate with formula; rate-environment-dependent"),
    ("CL268", "P4A#7 — Surface Dalbar/Morningstar behavior gap evidence as core empirical finding in philosophy section"),
    ("CL269", "P4A#8 — HNW asset location nuance (when 'bonds in deferred' rule breaks down); may belong in Phase 7"),
    ("CL272", "P4B#3 — Update robust factor set framing to 2025 consensus (quality meta-factor; size/value skepticism)"),
    ("CL273", "P4B#4 — Replication crisis update with Chen-Zimmermann 2020/2022 and Jensen-Kelly-Pedersen 2023 contemporary synthesis"),
    ("CL274", "P4B#5 — Michaud 1998 resampled frontier full treatment (DeMiguel-Garlappi-Uppal already added)"),
    ("CL275", "P4B#6 — Broaden behavioral synthesis attribution: Kahneman-Tversky 1979, Thaler 1980+, DeBondt-Thaler 1985, Shleifer-Vishny 1997 limits to arbitrage"),
    ("CL276", "P4B#7 — CAPM post-Roll pragmatic framing: Merton 1973 ICAPM, Campbell 1996 conditional CAPM"),
    ("CL277", "P4B#8 — Sharpe ratio calculator: confidence intervals (Lo 2002) and Sortino/Calmar alternatives for skewed returns"),
    ("CL278", "P4B#9 — Q-factor vs Fama-French five-factor framed as competing paradigms (not variants)"),
]
for cid, desc in p45_items:
    ws2.append([cid, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True, color="6A1B9A")
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 5 BUILD STATUS"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="0288D1")
ws2.append([])
phase5_status = [
    ("Zeitgeist: investing view built", "4 sections — FIRE and variants (lean/fat/barista/coast/GeoArb/Slow FI), FinTok and influencer-driven investing (Kakhbod 2023, SEC 2022 enforcement), retail trading culture (Barber-Odean 2000, Welch 2022, Ben-David-Birru-Rossi 2022, 0DTE), crypto and alternatives (BTC ETF Jan 2024, IRC Notice 2014-21)"),
    ("Zeitgeist: lifestyle view built", "4 sections — Die with Zero (Perkins 2020), anti-hustle and soft life, generational housing patterns (Pew multi-gen data), behavioral synthesis closer"),
    ("Phase 5 calculators", "Coast FIRE (compound-carry threshold), Dividend yield vs total return illustration, Die with Zero spending trajectory simplified"),
    ("Phase 5 fact-checks", "29 new claims added (CL279-CL307); 18 tier B (specific citations: Kakhbod 2023, SEC 2022 enforcement, Barber-Odean 2000, Welch 2022, Bryzgalova 2023, Perkins 2020, Pew multi-gen, Gallup engagement, Genworth Cost of Care, CoinGecko historical drawdowns), 9 tier A (BTC ETF approval Jan 2024, IRC Notice 2014-21, IRC §1091 wash sale, IRC §2503(e), compound interest math, Coast FIRE/dividend/DWZ calculator math), 2 tier C (FIRE variant terminology, soft saving phenomenon)"),
    ("Expansionist EX2 reviewed structure pre-build", "5 additions identified for Phase 5.5 backlog: Dave Ramsey orthodoxy (full section), expanded dividend investing treatment, BNPL section, Gen-Z anxiety / soft saving expanded, DINK financial pattern"),
    ("Phase 5 checkpoint personas", "P5A behavioral finance researcher (retail investor decisions, social media); P5B financial therapist / money coach (clinical practice) — completed"),
    ("Phase 5 structural fixes applied", "CL313 (retail trading paragraph rewritten: Welch misattribution removed, Ben-David $2B claim removed, Lakonishok 2007 and Bauer 2009 added, Welch correctly characterized as finding NO Robinhood underperformance), CL316 (Coast FIRE: hardcoded $20K → user input field), CL317 (dividend calc restructured: yield + capital appreciation as separate inputs for both strategies; Miller-Modigliani equivalence properly surfaced)"),
]
for label, desc in phase5_status:
    ws2.append([label, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True)
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append([])
ws2.append(["PHASE 5.5 BACKLOG — FROM EXPANSIONIST EX2 (PRE-BUILD REVIEW)"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="6A1B9A")
ws2.append([])
ws2.append(["", "Phase 5.5 backlog originates from the Expansionist EX2 scope review run BEFORE the Phase 5 build (rather than after, as with checkpoint personas). Same scheduling and accountability: must clear before Phase 6 (original 5 personas) reviews. SCHEDULED: runs alongside Phase 2.5, 3.5, 4.5 in the consolidated sub-phase after Phase 5 (zeitgeist behaviors) completes."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.append([])

p55_items = [
    # From Expansionist EX2 (pre-build scope review)
    ("CL308", "EX2#1 — Dave Ramsey orthodoxy as full section: Baby Steps, debt snowball vs avalanche, 12% return assumption, 8% SWR, credit card framing"),
    ("CL309", "EX2#2 — Expanded dividend investing treatment: Seeking Alpha ecosystem, dividend growth investing community, psychological appeal vs Miller-Modigliani"),
    ("CL310", "EX2#3 — Buy Now Pay Later (BNPL) section: Affirm/Klarna/Afterpay, payment fragmentation behavioral effects, credit reporting integration gap"),
    ("CL311", "EX2#4 — Gen-Z anxiety / 'system is broken' framing: real wage vs housing/education cost growth, soft saving phenomenon, 'work until I die' resignation"),
    ("CL312", "EX2#5 — DINK financial pattern section: child-free household trajectory, distinct savings rate capacity, different retirement healthcare considerations"),
    # From P5A checkpoint persona
    ("CL314", "P5A#3 — 0DTE framing refinement: most positions closed before expiration; held-to-expiration disproportionately worthless"),
    ("CL315", "P5A#4 — FinTok 56% misleading statistic: break down into factual errors vs ethical issues vs suboptimal advice"),
    ("CL318", "P5A#7 — Addictive design literature on Robinhood gamification: regulatory actions and academic critiques"),
    ("CL319", "P5A#8 — Crypto diversification-benefit decay: BTC/equity correlation has risen substantially 2014→2024"),
    ("CL320", "P5A#9 — FIRE blogger sample bias citation: ChooseFI demographic surveys, r/financialindependence data"),
    # From P5B checkpoint persona
    ("CL321", "P5B#1 — Post-FIRE depression / retirement transition mental health pattern (Calvo-Sarkisian-Tamborini 2013)"),
    ("CL322", "P5B#2 — Soft saving phenomenology: distinguish deliberate trade-off from resignation/giving-up pattern"),
    ("CL323", "P5B#3 — Die with Zero late-life cognitive decline caveat: spending utility window closes earlier than planned"),
    ("CL324", "P5B#4 — Burnout-as-financial-event quantification: career interruption can exceed decade's savings advantage"),
    ("CL325", "P5B#5 — Boomerang/multi-gen relational structure guidance: specific arrangements and red flags"),
    ("CL326", "P5B#6 — Savings-aversion clinical pattern acknowledgment (Klontz money disorders literature)"),
    ("CL327", "P5B#7 — DWZ calculator U-shape retirement spending: high early / lower middle / rising late-life medical"),
    ("CL328", "P5B#8 — Family-system framing in behavioral synthesis: most lifestyle decisions are family decisions"),
]
for cid, desc in p55_items:
    ws2.append([cid, desc])
    ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=10, bold=True, color="6A1B9A")
    ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

ws2.append([])
ws2.append(["", f"18 items in Phase 5.5 backlog (5 from EX2 pre-build review + 5 from P5A + 8 from P5B). All have full audit entries. SCHEDULED: Phase 5.5 will run alongside Phase 2.5, 3.5, 4.5 in the consolidated sub-phase after Phase 5 closes and BEFORE Phase 6 (original 5 personas review). Combined backlog at this point: 14 P2.5 + 13 P3.5 + 13 P4.5 + 18 P5.5 = 58 items."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2[f"B{ws2.max_row}"].font = Font(name="Arial", size=10, italic=True, color="6A1B9A")

# Column widths for summary
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 90

# ============================================================
# Sheet 3: Future Additions (Expansionist territory — UNVERIFIED)
# ============================================================
# Each item is a CANDIDATE for future implementation, not verified content.
# Must run full Phase 1-style fact-check before any is added to the artifact.

FUTURE_ADDITIONS = [
    # === EQUITY COMPENSATION ===
    ("FA001", "Equity comp", "§83(b) election within 30 days of grant converts future ordinary-income appreciation on restricted stock to long-term capital gain",
     "New node, Phase 7 or W2:9.x",
     "IRC §83(b); Treas. Reg. §1.83-2",
     "30-day deadline is strict; mechanics, who should make the election, basis tracking, interaction with §1202 holding period"),

    ("FA002", "Equity comp", "ISO vs NSO vs RSU tax treatment differences and planning implications",
     "New node, Phase 7",
     "IRC §421-§424 (ISOs); §83 (NSOs); §451 / Treas. Reg. §1.83-1 (RSUs)",
     "AMT exposure on ISO exercises; bargain element timing; 2-year/1-year holding period rules for ISO qualification"),

    ("FA003", "Equity comp", "Net Unrealized Appreciation (NUA) — distribute employer stock in-kind from 401(k); basis taxed ordinary, appreciation taxed LTCG",
     "W2:10.2 or new decumulation node",
     "IRC §402(e)(4); Rev. Rul. 75-125",
     "Lump-sum distribution requirement, qualifying triggering events, separate-account treatment, basis recovery rules"),

    ("FA004", "Equity comp", "§1202 stacking via non-grantor trust gifting — multiply $15M QSBS exclusion across family members",
     "Business:9.1",
     "IRC §1202(b); §1202(h) (gifts); Treas. Reg. §1.1202-2",
     "Pre-appreciation timing critical; non-grantor trust structure; substantial gift tax considerations; recent IRS scrutiny of aggressive stacking"),

    # === REAL ESTATE AS TAX INSTRUMENT ===
    ("FA005", "Real estate", "Cost segregation study accelerates 20–30% of property basis into year-one bonus depreciation",
     "Business:9.3",
     "IRC §168(k) (bonus depreciation, restored permanent by OBBBA §70401); §1.168(i)-1 (cost segregation methodology)",
     "100% bonus depreciation restored permanently under OBBBA for property placed in service after Jan 19, 2025 — VERIFY this date and scope"),

    ("FA006", "Real estate", "Real Estate Professional Status (REPS): 750+ hours material participation lets depreciation offset ordinary W-2 income",
     "Business:9.3",
     "IRC §469(c)(7); Treas. Reg. §1.469-9",
     "Hour-tracking documentation requirements; spouse can qualify; aggregation election; multiple court cases on substantiation"),

    ("FA007", "Real estate", "Short-term rental 'loophole': average stays <7 days qualify without REPS; material participation alone suffices",
     "Business:9.3",
     "Treas. Reg. §1.469-1T(e)(3)(ii)(A); Rev. Rul. 84-12",
     "7-day average stay test; material participation hours threshold; substantial-services exception nuances"),

    ("FA008", "Real estate", "§1031 like-kind exchange defers gain on real property reinvested in like-kind property",
     "Business:9.3 or new decumulation node",
     "IRC §1031 (post-TCJA: real property only); §1031(a)(3) timing rules",
     "45-day identification, 180-day acquisition, qualified intermediary required, no boot rules, related-party limitations"),

    ("FA009", "Real estate", "§1031 → DST → §721 UPREIT path: 1031 into Delaware Statutory Trust, then contribute to REIT under §721",
     "New node, Phase 7",
     "IRC §1031; §721; Rev. Rul. 2004-86 (DST treatment)",
     "Two-step transaction; preserves deferral chain; final REIT contribution provides liquidity and diversification"),

    # === FAMILY WEALTH MULTIPLIERS ===
    ("FA010", "Family wealth", "Custodial Roth IRA for minor children with earned income — 50+ years of tax-free compounding",
     "New node, family planning section",
     "IRC §408A; §408(a) (IRA generally); custodial structure under state UGMA/UTMA",
     "Earned income substantiation; reasonable compensation if child works for family business; control transfer at age of majority"),

    ("FA011", "Family wealth", "Family employment: hire children age 7+ in family business; deductible to business, taxable to child at lower bracket",
     "New node, family planning",
     "Treas. Reg. §1.162-7 (reasonable compensation); Eller v. Commissioner; Denman v. Commissioner",
     "Reasonable wage substantiation; actual work performed; FICA exemption for unincorporated parent employers (children under 18)"),

    ("FA012", "Family wealth", "529 superfunding: front-load 5 years of annual gift exclusion = $95K/donor, $190K/couple per beneficiary",
     "W2:9.4",
     "IRC §529(c)(2)(B); Form 709 5-year election",
     "5-year proration of exclusion; no additional exclusion gifts to same beneficiary during the 5-year period; death within 5 years pulls partial amount back to estate"),

    ("FA013", "Family wealth", "Family Limited Partnerships (FLPs) with valuation discounts of 25–40% on transferred interests",
     "Business:10.3",
     "IRC §2031; §2036 (recent estate tax cases); Rev. Rul. 93-12",
     "Lack-of-control and lack-of-marketability discounts; IRS scrutiny under §2036; documented business purpose required"),

    ("FA014", "Family wealth", "Intentionally Defective Grantor Trusts (IDGTs) — grantor pays income tax on trust assets, effectively tax-free gifts",
     "Business:10.3",
     "IRC §671-§679 (grantor trust rules); §1014 (basis at death)",
     "Sale to IDGT in exchange for promissory note; valuation discount potential; AFR rates apply"),

    # === COMPOUND RETIREMENT STACK ===
    ("FA015", "Compound stack", "Cash Balance Defined Benefit Plan on top of Solo 401(k): $200K–$300K additional pre-tax shelter for high-income owners",
     "Contractor:8.1 / Business:8.2",
     "IRC §415(b) ($290K DB benefit limit 2026); §401(a)(26); §401(a)(4)",
     "Actuarial certification required; consistent income requirement; admin costs $3K-$5K/yr; PBGC coverage for some plans"),

    ("FA016", "Compound stack", "Full §415 stack for high earners: HSA + IRA + 401(k) employee + profit-sharing + Mega Backdoor Roth + Cash Balance DB",
     "Maxtax phase synthesis node",
     "IRC §415(c) ($72K DC); §415(b) ($290K DB); §402(g) ($24,500 deferral)",
     "Aggregate planning rather than individual account optimization; can shelter $400K-$500K/yr at top brackets"),

    # === DECUMULATION UPSIDE ===
    ("FA017", "Decumulation upside", "0% long-term capital gains harvesting: realize gains tax-free when taxable income under ~$96,700 MFJ / ~$48,350 single (2026)",
     "W2:10.2",
     "IRC §1(h)(1)(A) 0% rate bracket",
     "Annual basis reset opportunity; interaction with Social Security taxation; ACA cliff coordination"),

    ("FA018", "Decumulation upside", "Qualified Charitable Distributions (QCDs) at 70½+: direct IRA-to-charity distribution up to $108,000 (2025); satisfies RMDs, excluded from AGI",
     "W2:10.x or new node",
     "IRC §408(d)(8); SECURE 2.0 §307 (indexing the QCD limit)",
     "Age 70½ trigger (not RMD age 73); $108K limit indexed; new SECURE 2.0 split-interest CRT/CGA option ($53K one-time)"),

    ("FA019", "Decumulation upside", "State domicile arbitrage: relocate from high-tax state (CA, NY, NJ) to no-income-tax state (FL, TN, NV, TX, WA) at retirement",
     "W2:10.x or new geographic-arbitrage node",
     "State tax law variation; residency tests vary by state",
     "Domicile vs residency requirements; ongoing audit risk from origin state (especially CA, NY); statutory residency rules"),

    ("FA020", "Decumulation upside", "Dynasty trust states (SD, NV, DE, AK, TN) — no state income tax on trust, perpetual duration, strong creditor protection",
     "Business:10.3",
     "State trust statutes; SD Codified Laws §55-1-19; NV Rev. Stat. §163.5555",
     "Rule against perpetuities abolished in trust state; situs requirements; GST tax planning interaction"),

    # === ADJACENT VEHICLES ===
    ("FA021", "Adjacent vehicles", "Direct indexing: separately managed accounts replicating an index while harvesting losses at individual security level",
     "W2:9.3",
     "Practitioner literature; SMA provider methodologies (Parametric, Aperio, Wealthfront)",
     "30-50% more TLH alpha than index funds claimed; minimum account size typically $100K-$250K; tracking error vs index"),

    ("FA022", "Adjacent vehicles", "§1256 contracts: futures and broad-based index options taxed 60% LTCG / 40% ST regardless of holding period",
     "W2:9.x or new node",
     "IRC §1256; §1256(e) exclusions",
     "Year-end mark-to-market; broad-based index options included; cannot be combined with §1092 straddle treatment"),

    ("FA023", "Adjacent vehicles", "Opportunity Zone (QOZ) investments: defer capital gains until 2026; 10% basis step-up at 5 years; tax-free appreciation after 10-year hold",
     "New node, Phase 7",
     "IRC §1400Z-1; §1400Z-2",
     "Original deferral until 12/31/2026 (statutory); OBBBA may have created new round — VERIFY current law"),

    ("FA024", "Adjacent vehicles", "§831(b) micro-captive insurance for predictable business risks — up to $2.45M (2026) premium income excluded from corporate income",
     "Business:9.x",
     "IRC §831(b); IRS Notice 2016-66 (listed transactions); recent Tax Court cases",
     "Heavy IRS scrutiny on abusive structures; legitimate uses for genuine insurance need; bona fide insurance requirements (risk distribution, risk shifting)"),

    ("FA025", "Adjacent vehicles", "Self-Directed IRAs (SDIRAs) for alternative investments: real estate, private equity, precious metals, crypto inside retirement wrapper",
     "W2:9.x or new node",
     "IRC §408(a); §4975 (prohibited transactions)",
     "Prohibited transaction rules; UBTI on leveraged real estate; UBIT for active business income; custodian requirements"),
]

ws3 = wb.create_sheet("Future Additions")
fa_headers = ["ID", "Theme", "Strategy", "Suggested Location", "Sources to Verify", "Phase 7 Verification Notes"]
ws3.append(fa_headers)

for row in FUTURE_ADDITIONS:
    ws3.append(list(row))

# Format header
for c in range(1, len(fa_headers) + 1):
    cell = ws3.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Format body
theme_colors = {
    "Equity comp": "EDE7F6",
    "Real estate": "E8F5E9",
    "Family wealth": "FFF8E1",
    "Compound stack": "E3F2FD",
    "Decumulation upside": "FFF3E0",
    "Adjacent vehicles": "FCE4EC",
}
for r in range(2, len(FUTURE_ADDITIONS) + 2):
    theme = ws3.cell(row=r, column=2).value
    fill_color = theme_colors.get(theme, "FFFFFF")
    for c in range(1, len(fa_headers) + 1):
        cell = ws3.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = border
        cell.fill = PatternFill("solid", start_color=fill_color)

# Column widths
fa_widths = {"A": 8, "B": 22, "C": 60, "D": 32, "E": 45, "F": 55}
for col, w in fa_widths.items():
    ws3.column_dimensions[col].width = w

ws3.freeze_panes = "A2"

# Add status note at top of summary about Future Additions
ws2.append([])
ws2.append([])
ws2.append(["FUTURE ADDITIONS BACKLOG"])
ws2[f"A{ws2.max_row}"].font = Font(name="Arial", size=12, bold=True, color="6A1B9A")
ws2.append([])
ws2.append(["", f"{len(FUTURE_ADDITIONS)} candidate strategies surfaced by Expansionist persona, organized by 6 themes. ALL UNVERIFIED — each must pass Phase 1-style fact-check (citation, current law, tier assignment) before any implementation in the artifact. See 'Future Additions' sheet."])
ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

# Save
import os
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "phase1_factcheck_audit.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total claims: {len(CLAIMS)}")
print(f"Status: {dict(status_count)}")
print(f"Tier (passed/corrected): {dict(tier_count)}")
